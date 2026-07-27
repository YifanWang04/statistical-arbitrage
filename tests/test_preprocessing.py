from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

import duckdb
import numpy as np
import pandas as pd

from stat_arb_data.database import DuckDBDataset
from stat_arb_preprocessing import (
    PreprocessingConfig,
    build_preprocessing,
    get_snapshot,
)
from stat_arb_preprocessing.calculations import build_snapshot_from_frames
from stat_arb_preprocessing.excel import export_snapshot_workbook
from stat_arb_preprocessing.repository import EXPECTED_RETURN_BASIS


class PreprocessingIntegrationTests(unittest.TestCase):
    def test_vectorized_snapshot_matches_scalar_reference(self) -> None:
        window_dates = tuple(
            value.date()
            for value in pd.bdate_range("2026-01-05", periods=5)
        )
        membership = pd.DataFrame(
            [
                {"ticker": "BBB", "market_cap_rank": 2},
                {"ticker": "FFF", "market_cap_rank": 6},
                {"ticker": "AAA", "market_cap_rank": 1},
                {"ticker": "DDD", "market_cap_rank": 4},
                {"ticker": "CCC", "market_cap_rank": 3},
                {"ticker": "EEE", "market_cap_rank": 5},
            ]
        )
        residual_patterns = {
            "AAA": (0.01, 0.02, 0.03, 0.04, 0.05),
            "BBB": (0.01, 0.04, 0.02, 0.05, 0.03),
            "CCC": (0.02, 0.02, 0.02, 0.02, 0.02),
            "DDD": (0.03, 0.01, 0.04, 0.02, 0.05),
            "EEE": (0.05, 0.03, 0.01, 0.04, 0.02),
            "FFF": (0.04, 0.02, 0.05, 0.01, 0.03),
        }
        rows: list[dict[str, object]] = []
        for ticker, residual_values in residual_patterns.items():
            for index, (trade_date, residual) in enumerate(
                zip(window_dates, residual_values, strict=True)
            ):
                if ticker == "DDD" and index == 2:
                    continue
                is_valid = True
                reason = None
                if ticker == "EEE" and index in (1, 3):
                    is_valid = False
                    reason = "later_reason" if index == 3 else None
                if ticker == "FFF" and index in (0, 4):
                    is_valid = False
                    reason = "early_reason" if index == 0 else "late_reason"
                rows.append(
                    {
                        "trade_date": trade_date,
                        "ticker": ticker,
                        "market_cap_rank": next(
                            int(row["market_cap_rank"])
                            for row in membership.to_dict("records")
                            if row["ticker"] == ticker
                        ),
                        "stock_return": residual + 0.001,
                        "market_return": 0.001 * (index + 1),
                        "beta": 1.0 + index / 10.0,
                        "market_residual_return": residual,
                        "is_valid": is_valid,
                        "exclusion_reason": reason,
                    }
                )
        daily_residuals = pd.DataFrame(rows).sample(
            frac=1.0,
            random_state=7,
        )
        config = PreprocessingConfig(
            database_path=Path("unused.duckdb"),
            correlation_window=5,
        )

        expected = _scalar_snapshot_expectations(
            window_dates,
            membership,
            daily_residuals,
            config.variance_epsilon,
        )
        snapshot = build_snapshot_from_frames(
            preprocessing_run_id="test-run",
            return_basis=EXPECTED_RETURN_BASIS,
            as_of_date=date(2026, 1, 12),
            window_dates=window_dates,
            membership=membership,
            daily_residuals=daily_residuals,
            config=config,
        )

        self.assertEqual(snapshot.tickers, expected["tickers"])
        self.assertEqual(snapshot.market_cap_ranks, expected["ranks"])
        pd.testing.assert_frame_equal(snapshot.exclusions, expected["exclusions"])
        pd.testing.assert_frame_equal(snapshot.beta_matrix, expected["beta"])
        pd.testing.assert_frame_equal(
            snapshot.stock_return_matrix,
            expected["stock_return"],
        )
        pd.testing.assert_frame_equal(
            snapshot.residual_matrix,
            expected["market_residual_return"],
        )
        pd.testing.assert_series_equal(
            snapshot.market_returns,
            expected["market_return"],
        )
        pd.testing.assert_frame_equal(
            snapshot.correlation_matrix,
            expected["correlation"],
        )
        self.assertEqual(
            snapshot.exclusions["reason"].tolist(),
            [
                "zero_residual_variance",
                "incomplete_residual_window:missing_residual_row",
                "incomplete_residual_window:later_reason",
                "incomplete_residual_window:early_reason",
            ],
        )

    def test_vectorized_snapshot_still_rejects_inconsistent_market_returns(
        self,
    ) -> None:
        window_dates = tuple(
            value.date()
            for value in pd.bdate_range("2026-01-05", periods=5)
        )
        membership = pd.DataFrame(
            [
                {"ticker": "AAA", "market_cap_rank": 1},
                {"ticker": "BBB", "market_cap_rank": 2},
            ]
        )
        rows = [
            {
                "trade_date": trade_date,
                "ticker": ticker,
                "market_cap_rank": rank,
                "stock_return": residual,
                "market_return": (
                    0.02
                    if ticker == "BBB" and index == 2
                    else 0.01
                ),
                "beta": 1.0,
                "market_residual_return": residual,
                "is_valid": True,
                "exclusion_reason": None,
            }
            for rank, ticker in enumerate(("AAA", "BBB"), start=1)
            for index, (trade_date, residual) in enumerate(
                zip(
                    window_dates,
                    (
                        (0.01, 0.02, 0.03, 0.04, 0.05)
                        if ticker == "AAA"
                        else (0.02, 0.01, 0.04, 0.03, 0.05)
                    ),
                    strict=True,
                )
            )
        ]

        with self.assertRaisesRegex(
            ValueError,
            "Market returns are inconsistent",
        ):
            build_snapshot_from_frames(
                preprocessing_run_id="test-run",
                return_basis=EXPECTED_RETURN_BASIS,
                as_of_date=date(2026, 1, 12),
                window_dates=window_dates,
                membership=membership,
                daily_residuals=pd.DataFrame(rows),
                config=PreprocessingConfig(
                    database_path=Path("unused.duckdb"),
                    correlation_window=5,
                ),
            )

    def test_build_computes_inclusive_beta_and_preserves_warmup_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            config = PreprocessingConfig(database_path=database)

            run_id = build_preprocessing(config)

            with duckdb.connect(str(database), read_only=True) as connection:
                rows = connection.execute(
                    """
                    SELECT trade_date, beta, market_residual_return,
                           beta_observation_count, is_valid, exclusion_reason
                    FROM preprocessing.daily_market_residuals
                    WHERE ticker = 'AAA' AND trade_date IN (?, ?)
                    ORDER BY trade_date
                    """,
                    [dates[58].date(), dates[59].date()],
                ).fetchall()
                completed = connection.execute(
                    "SELECT status FROM audit.preprocessing_runs WHERE run_id = ?",
                    [run_id],
                ).fetchone()

            self.assertEqual(rows[0][3:], (59, False, "insufficient_beta_window"))
            self.assertAlmostEqual(rows[1][1], 2.0, places=12)
            self.assertAlmostEqual(rows[1][2], 0.0, places=12)
            self.assertEqual(rows[1][3:], (60, True, None))
            self.assertEqual(completed, ("completed",))

    def test_snapshot_computes_correlation_excludes_zero_variance_and_reuses_cache(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            config = PreprocessingConfig(database_path=database)
            build_preprocessing(config)
            as_of_date = dates[-1].date()

            first = get_snapshot(config, as_of_date, cache=True)
            second = get_snapshot(config, as_of_date, cache=True)

            self.assertEqual(first.snapshot_id, second.snapshot_id)
            self.assertEqual(first.tickers, ("BBB", "CCC", "DDD"))
            self.assertEqual(first.residual_matrix.shape, (5, 3))
            self.assertEqual(first.correlation_matrix.shape, (3, 3))
            self.assertTrue(
                np.allclose(
                    first.correlation_matrix.to_numpy(),
                    np.corrcoef(first.residual_matrix.to_numpy(), rowvar=False),
                )
            )
            self.assertEqual(
                first.exclusions.set_index("ticker").loc["AAA", "reason"],
                "zero_residual_variance",
            )

            with duckdb.connect(str(database), read_only=True) as connection:
                correlation_count = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.snapshot_correlations"
                ).fetchone()[0]
            self.assertEqual(correlation_count, 6)

    def test_snapshot_replaces_same_date_cache_when_correlation_window_changes(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            five_day_config = PreprocessingConfig(
                database_path=database,
                correlation_window=5,
            )
            build_preprocessing(five_day_config)
            as_of_date = dates[-1].date()
            five_day = get_snapshot(five_day_config, as_of_date, cache=True)

            ten_day_config = PreprocessingConfig(
                database_path=database,
                correlation_window=10,
            )
            ten_day = get_snapshot(ten_day_config, as_of_date, cache=True)
            cached_ten_day = get_snapshot(ten_day_config, as_of_date, cache=True)

            self.assertEqual(len(five_day.residual_matrix), 5)
            self.assertEqual(len(ten_day.residual_matrix), 10)
            self.assertNotEqual(five_day.snapshot_id, ten_day.snapshot_id)
            self.assertEqual(ten_day.snapshot_id, cached_ten_day.snapshot_id)
            with duckdb.connect(str(database), read_only=True) as connection:
                snapshots = connection.execute(
                    """
                    SELECT snapshot_id, correlation_window
                    FROM preprocessing.correlation_snapshots
                    WHERE as_of_date = ?
                    """,
                    [as_of_date],
                ).fetchall()
            self.assertEqual(snapshots, [(ten_day.snapshot_id, 10)])

    def test_snapshot_rejects_config_that_does_not_match_built_residuals(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            build_preprocessing(PreprocessingConfig(database_path=database))
            mismatched_config = PreprocessingConfig(
                database_path=database,
                beta_window=20,
            )

            with self.assertRaisesRegex(RuntimeError, "run preprocessing build"):
                get_snapshot(mismatched_config, dates[-1].date(), cache=True)

    def test_failed_same_date_cache_replacement_preserves_previous_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            five_day_config = PreprocessingConfig(
                database_path=database,
                correlation_window=5,
            )
            build_preprocessing(five_day_config)
            as_of_date = dates[-1].date()
            original = get_snapshot(five_day_config, as_of_date, cache=True)
            with duckdb.connect(str(database)) as connection:
                correlation_count = connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM preprocessing.snapshot_correlations
                    WHERE snapshot_id = ?
                    """,
                    [original.snapshot_id],
                ).fetchone()[0]
                make_new_snapshot_correlations_fail(connection, original.snapshot_id)

            ten_day_config = PreprocessingConfig(
                database_path=database,
                correlation_window=10,
            )
            with self.assertRaises(duckdb.ConstraintException):
                get_snapshot(ten_day_config, as_of_date, cache=True)

            restored = get_snapshot(five_day_config, as_of_date, cache=True)
            self.assertEqual(restored.snapshot_id, original.snapshot_id)
            with duckdb.connect(str(database), read_only=True) as connection:
                snapshot_rows = connection.execute(
                    """
                    SELECT snapshot_id, correlation_window
                    FROM preprocessing.correlation_snapshots
                    WHERE as_of_date = ?
                    """,
                    [as_of_date],
                ).fetchall()
                restored_correlation_count = connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM preprocessing.snapshot_correlations
                    WHERE snapshot_id = ?
                    """,
                    [original.snapshot_id],
                ).fetchone()[0]
            self.assertEqual(snapshot_rows, [(original.snapshot_id, 5)])
            self.assertEqual(restored_correlation_count, correlation_count)

    def test_legacy_preprocessing_schema_is_migrated_once_without_losing_residuals(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            config = PreprocessingConfig(database_path=database)
            build_preprocessing(config)
            as_of_date = dates[-1].date()
            with duckdb.connect(str(database)) as connection:
                residual_count = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.daily_market_residuals"
                ).fetchone()[0]
                downgrade_preprocessing_schema(connection, as_of_date)

            uncached = get_snapshot(config, as_of_date, cache=False)

            with duckdb.connect(str(database), read_only=True) as connection:
                daily_columns = {
                    row[0]
                    for row in connection.execute(
                        "DESCRIBE preprocessing.daily_market_residuals"
                    ).fetchall()
                }
                snapshot_columns = {
                    row[0]
                    for row in connection.execute(
                        "DESCRIBE preprocessing.correlation_snapshots"
                    ).fetchall()
                }
                migrated_count = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.daily_market_residuals"
                ).fetchone()[0]
                cached_count = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.correlation_snapshots"
                ).fetchone()[0]
                epsilon = connection.execute(
                    """
                    SELECT variance_epsilon
                    FROM audit.preprocessing_runs
                    WHERE run_id = ?
                    """,
                    [uncached.preprocessing_run_id],
                ).fetchone()[0]
            self.assertIn("beta", daily_columns)
            self.assertNotIn("beta_60d", daily_columns)
            self.assertTrue(
                {
                    "beta_window",
                    "correlation_window",
                    "beta_alignment",
                    "missing_policy",
                    "calculation_version",
                    "variance_epsilon",
                    "return_basis",
                }.issubset(snapshot_columns)
            )
            self.assertEqual(migrated_count, residual_count)
            self.assertEqual(cached_count, 0)
            self.assertEqual(epsilon, config.variance_epsilon)

            first_cached = get_snapshot(config, as_of_date, cache=True)
            second_cached = get_snapshot(config, as_of_date, cache=True)
            self.assertEqual(first_cached.snapshot_id, second_cached.snapshot_id)

    def test_snapshot_excludes_an_incomplete_residual_window(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            config = PreprocessingConfig(database_path=database)
            build_preprocessing(config)
            as_of_date = dates[-1].date()
            missing_date = dates[-2].date()

            with duckdb.connect(str(database)) as connection:
                connection.execute(
                    """
                    UPDATE preprocessing.daily_market_residuals
                    SET is_valid = FALSE,
                        market_residual_return = NULL,
                        exclusion_reason = 'missing_stock_return'
                    WHERE trade_date = ? AND ticker = 'BBB'
                    """,
                    [missing_date],
                )
            snapshot = get_snapshot(config, as_of_date, cache=False)

            self.assertEqual(snapshot.tickers, ("CCC", "DDD"))
            self.assertEqual(
                snapshot.exclusions.set_index("ticker").loc["BBB", "reason"],
                "incomplete_residual_window:missing_stock_return",
            )

    def test_successful_rebuild_clears_snapshot_cache(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            config = PreprocessingConfig(database_path=database)
            first_run = build_preprocessing(config)
            get_snapshot(config, dates[-1].date(), cache=True)

            second_run = build_preprocessing(config)

            self.assertNotEqual(first_run, second_run)
            with duckdb.connect(str(database), read_only=True) as connection:
                snapshot_count = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.correlation_snapshots"
                ).fetchone()[0]
                run_ids = connection.execute(
                    "SELECT DISTINCT calculation_run_id FROM preprocessing.daily_market_residuals"
                ).fetchall()
            self.assertEqual(snapshot_count, 0)
            self.assertEqual(run_ids, [(second_run,)])

    def test_as_of_snapshot_does_not_use_as_of_date_residual(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            config = PreprocessingConfig(database_path=database)
            build_preprocessing(config)
            as_of_date = dates[-1].date()
            before = get_snapshot(config, as_of_date, cache=False)

            with duckdb.connect(str(database)) as connection:
                connection.execute(
                    """
                    UPDATE preprocessing.daily_market_residuals
                    SET market_residual_return = 999.0
                    WHERE trade_date = ? AND ticker IN ('BBB', 'CCC')
                    """,
                    [as_of_date],
                )
            after = get_snapshot(config, as_of_date, cache=False)

            self.assertTrue(before.residual_matrix.equals(after.residual_matrix))
            self.assertTrue(before.correlation_matrix.equals(after.correlation_matrix))

    def test_excel_export_has_expected_sheets_and_refuses_silent_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database, dates = build_test_database(root)
            config = PreprocessingConfig(database_path=database)
            build_preprocessing(config)
            snapshot = get_snapshot(config, dates[-1].date())
            output = root / "snapshot.xlsx"

            export_snapshot_workbook(snapshot, output)

            from openpyxl import load_workbook

            workbook = load_workbook(output, data_only=False, read_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary",
                    "Beta_Used",
                    "Stock_Returns",
                    "Residual_Matrix",
                    "Correlation_Matrix",
                    "Excluded_Stocks",
                ],
            )
            stock_returns_sheet = workbook["Stock_Returns"]
            self.assertEqual(stock_returns_sheet["B1"].value, "SPY")
            self.assertEqual(
                stock_returns_sheet.max_column,
                len(snapshot.tickers) + 2,
            )
            for row_number, expected_return in enumerate(
                snapshot.market_returns,
                start=2,
            ):
                self.assertAlmostEqual(
                    stock_returns_sheet.cell(row_number, 2).value,
                    float(expected_return),
                )
            self.assertEqual(workbook["Residual_Matrix"].max_row, 6)
            self.assertEqual(workbook["Correlation_Matrix"].max_row, 4)
            correlation_sheet = workbook["Correlation_Matrix"]
            conditional_ranges = list(correlation_sheet.conditional_formatting)
            self.assertEqual(len(conditional_ranges), 1)
            correlation_rules = correlation_sheet.conditional_formatting[
                conditional_ranges[0]
            ]
            diagonal_rules = [
                rule
                for rule in correlation_rules
                if rule.type == "expression" and rule.stopIfTrue
            ]
            self.assertEqual(len(diagonal_rules), 1)
            summary = workbook["Summary"]
            summary_rows = {
                summary.cell(row, 1).value: row
                for row in range(2, summary.max_row + 1)
                if summary.cell(row, 1).value is not None
            }
            self.assertEqual(
                summary.cell(summary_rows["Overall QC"], 2).value,
                "OK",
            )
            self.assertEqual(summary.max_column, 2)
            workbook.close()
            with self.assertRaises(FileExistsError):
                export_snapshot_workbook(snapshot, output)

    def test_non_default_parameters_and_qc_formats_are_exported_without_hardcoding(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database, dates = build_test_database(root)
            config = PreprocessingConfig(
                database_path=database,
                beta_window=20,
                correlation_window=10,
                calculation_version="paper_baseline_test",
                variance_epsilon=1e-12,
            )
            build_preprocessing(config)
            snapshot = get_snapshot(config, dates[-1].date(), cache=True)
            output = root / "snapshot.xlsx"
            export_snapshot_workbook(snapshot, output)

            self.assertEqual(snapshot.beta_window, 20)
            self.assertEqual(snapshot.correlation_window, 10)
            self.assertEqual(snapshot.calculation_version, "paper_baseline_test")
            self.assertEqual(snapshot.variance_epsilon, 1e-12)

            from openpyxl import load_workbook

            workbook = load_workbook(output, data_only=False, read_only=False)
            parameters = workbook["Summary"]
            rows = {
                parameters.cell(row, 1).value: row
                for row in range(2, parameters.max_row + 1)
                if parameters.cell(row, 1).value is not None
            }
            expected_values = {
                "Return basis": EXPECTED_RETURN_BASIS,
                "Beta window": 20,
                "Correlation window": 10,
                "Beta alignment": config.beta_alignment,
                "Missing policy": config.missing_policy,
                "Calculation version": config.calculation_version,
                "Variance epsilon": config.variance_epsilon,
            }
            for label, expected in expected_values.items():
                self.assertEqual(parameters.cell(rows[label], 2).value, expected)

            for label in (
                "Beta window",
                "Correlation window",
                "Selected stocks",
                "Valid stocks",
                "Excluded stocks",
                "Numerical rank",
            ):
                self.assertEqual(parameters.cell(rows[label], 2).number_format, "0")
            self.assertEqual(
                parameters.cell(rows["Contains non-finite values"], 2).number_format,
                "General",
            )
            self.assertEqual(
                parameters.cell(rows["Contains non-finite values"], 2).value,
                "NO",
            )
            self.assertEqual(
                parameters.cell(rows["Minimum eigenvalue"], 2).number_format,
                "0.000000E+00",
            )
            self.assertEqual(
                parameters.cell(rows["Maximum asymmetry"], 2).number_format,
                "0.000000E+00",
            )
            self.assertEqual(parameters.max_column, 2)
            workbook.close()


def _scalar_snapshot_expectations(
    window_dates: tuple[date, ...],
    membership: pd.DataFrame,
    daily_residuals: pd.DataFrame,
    variance_epsilon: float,
) -> dict[str, object]:
    ordered_membership = membership.sort_values(
        "market_cap_rank"
    ).reset_index(drop=True)
    expected_dates = pd.Index(
        pd.to_datetime(window_dates),
        name="trade_date",
    )
    residuals = daily_residuals.copy()
    residuals["trade_date"] = pd.to_datetime(residuals["trade_date"])
    exclusions: list[dict[str, object]] = []
    tickers: list[str] = []
    ranks: list[int] = []

    for row in ordered_membership.itertuples(index=False):
        ticker = str(row.ticker)
        rank = int(row.market_cap_rank)
        ticker_rows = (
            residuals.loc[residuals["ticker"] == ticker]
            .set_index("trade_date")
            .reindex(expected_dates)
        )
        invalid = ticker_rows.loc[
            (~ticker_rows["is_valid"].eq(True))
            | ticker_rows["market_residual_return"].isna()
        ]
        if not invalid.empty:
            raw_reason = invalid["exclusion_reason"].dropna()
            reason = (
                str(raw_reason.iloc[0])
                if not raw_reason.empty
                else "missing_residual_row"
            )
            exclusions.append(
                {
                    "ticker": ticker,
                    "market_cap_rank": rank,
                    "reason": f"incomplete_residual_window:{reason}",
                }
            )
            continue
        values = ticker_rows[
            "market_residual_return"
        ].astype(float).to_numpy()
        if float(np.std(values, ddof=1)) <= variance_epsilon:
            exclusions.append(
                {
                    "ticker": ticker,
                    "market_cap_rank": rank,
                    "reason": "zero_residual_variance",
                }
            )
            continue
        tickers.append(ticker)
        ranks.append(rank)

    indexed = residuals.set_index(["trade_date", "ticker"]).sort_index()

    def matrix(column: str) -> pd.DataFrame:
        return (
            indexed[column]
            .unstack("ticker")
            .reindex(index=expected_dates, columns=tickers)
            .astype(float)
        )

    market_frame = matrix("market_return")
    market_return = market_frame.iloc[:, 0].rename("SPY")
    residual_matrix = matrix("market_residual_return")
    return {
        "tickers": tuple(tickers),
        "ranks": tuple(ranks),
        "exclusions": pd.DataFrame(
            exclusions,
            columns=("ticker", "market_cap_rank", "reason"),
        ).sort_values("market_cap_rank", ignore_index=True),
        "beta": matrix("beta"),
        "stock_return": matrix("stock_return"),
        "market_residual_return": residual_matrix,
        "market_return": market_return,
        "correlation": residual_matrix.corr(),
    }


def build_test_database(directory: Path) -> tuple[Path, pd.DatetimeIndex]:
    # Match the production catalog shape. DuckDB treats a catalog and schema
    # with the same name as ambiguous when resolving two-part identifiers.
    database = directory / "market.duckdb"
    dates = pd.bdate_range("2020-01-02", periods=70)
    x = np.arange(len(dates), dtype=float)
    market_returns = 0.0005 + 0.0002 * np.sin(x / 4.0) + 0.00001 * x
    stock_returns = {
        "AAA": 2.0 * market_returns,
        "BBB": market_returns + 0.001 * np.sin(x / 2.0),
        "CCC": 0.8 * market_returns + 0.001 * np.cos(x / 3.0),
        "DDD": 1.2 * market_returns + 0.0007 * np.sin(x / 5.0 + 0.3),
    }

    with DuckDBDataset(database) as dataset:
        dataset.initialise()
    with duckdb.connect(str(database)) as connection:
        connection.execute(
            """
            INSERT INTO audit.settings VALUES ('return_basis', ?, 'test return basis')
            """,
            [EXPECTED_RETURN_BASIS],
        )
        market_frame = pd.DataFrame(
            {
                "ticker": "SPY",
                "trade_date": dates.date,
                "market_return": market_returns,
                "source": "test",
            }
        )
        price_frame = pd.concat(
            [
                pd.DataFrame(
                    {
                        "ticker": ticker,
                        "trade_date": dates.date,
                        "price_return": returns,
                        "source": "test",
                    }
                )
                for ticker, returns in stock_returns.items()
            ],
            ignore_index=True,
        )
        membership = pd.DataFrame(
            [
                {
                    "eligible_date": trade_date,
                    "ticker": ticker,
                    "market_cap_rank": rank,
                }
                for trade_date in dates.date
                for rank, ticker in enumerate(("AAA", "BBB", "CCC", "DDD"), start=1)
            ]
        )
        connection.register("market_frame", market_frame)
        connection.register("price_frame", price_frame)
        connection.register("membership_frame", membership)
        try:
            connection.execute(
                """
                INSERT INTO market_data.market_returns (ticker, trade_date, market_return, source)
                SELECT ticker, trade_date, market_return, source FROM market_frame
                """
            )
            connection.execute(
                """
                INSERT INTO market_data.daily_prices (ticker, trade_date, price_return, source)
                SELECT ticker, trade_date, price_return, source FROM price_frame
                """
            )
            connection.execute(
                """
                CREATE TABLE market_data.universe_membership AS
                SELECT * FROM membership_frame
                """
            )
        finally:
            connection.unregister("market_frame")
            connection.unregister("price_frame")
            connection.unregister("membership_frame")
    return database, dates


def downgrade_preprocessing_schema(
    connection: duckdb.DuckDBPyConnection,
    as_of_date: date,
) -> None:
    connection.execute(
        """
        ALTER TABLE audit.preprocessing_runs DROP COLUMN variance_epsilon;
        DROP TABLE preprocessing.snapshot_correlations;
        DROP TABLE preprocessing.snapshot_residuals;
        DROP TABLE preprocessing.snapshot_exclusions;
        DROP TABLE preprocessing.correlation_snapshots;

        CREATE TABLE preprocessing.correlation_snapshots (
            snapshot_id VARCHAR PRIMARY KEY,
            preprocessing_run_id VARCHAR NOT NULL,
            as_of_date DATE NOT NULL UNIQUE,
            window_start DATE NOT NULL,
            window_end DATE NOT NULL,
            selected_stock_count INTEGER NOT NULL,
            valid_stock_count INTEGER NOT NULL,
            excluded_stock_count INTEGER NOT NULL,
            maximum_asymmetry DOUBLE NOT NULL,
            minimum_correlation DOUBLE NOT NULL,
            maximum_correlation DOUBLE NOT NULL,
            minimum_eigenvalue DOUBLE NOT NULL,
            numerical_rank INTEGER NOT NULL,
            has_non_finite_values BOOLEAN NOT NULL,
            created_at TIMESTAMPTZ NOT NULL
        );
        CREATE TABLE preprocessing.snapshot_residuals (
            snapshot_id VARCHAR NOT NULL,
            trade_date DATE NOT NULL,
            column_index INTEGER NOT NULL,
            ticker VARCHAR NOT NULL,
            market_cap_rank INTEGER NOT NULL,
            stock_return DOUBLE NOT NULL,
            market_return DOUBLE NOT NULL,
            beta_60d DOUBLE NOT NULL,
            market_residual_return DOUBLE NOT NULL,
            PRIMARY KEY (snapshot_id, trade_date, column_index)
        );
        CREATE TABLE preprocessing.snapshot_correlations (
            snapshot_id VARCHAR NOT NULL,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            ticker_i VARCHAR NOT NULL,
            ticker_j VARCHAR NOT NULL,
            correlation DOUBLE NOT NULL,
            CHECK (row_index <= column_index),
            PRIMARY KEY (snapshot_id, row_index, column_index)
        );
        CREATE TABLE preprocessing.snapshot_exclusions (
            snapshot_id VARCHAR NOT NULL,
            ticker VARCHAR NOT NULL,
            market_cap_rank INTEGER NOT NULL,
            reason VARCHAR NOT NULL,
            PRIMARY KEY (snapshot_id, ticker)
        );
        """
    )
    run_id = connection.execute(
        """
        SELECT calculation_run_id
        FROM preprocessing.daily_market_residuals
        LIMIT 1
        """
    ).fetchone()[0]
    connection.execute(
        """
        INSERT INTO preprocessing.correlation_snapshots VALUES (
            'legacy-cache', ?, ?, ?, ?, 4, 4, 0,
            0.0, -0.5, 1.0, 0.0, 4, FALSE, current_timestamp
        )
        """,
        [run_id, as_of_date, as_of_date, as_of_date],
    )


def make_new_snapshot_correlations_fail(
    connection: duckdb.DuckDBPyConnection,
    preserved_snapshot_id: str,
) -> None:
    escaped_snapshot_id = preserved_snapshot_id.replace("'", "''")
    connection.execute(
        f"""
        CREATE TABLE preprocessing.snapshot_correlations_guarded (
            snapshot_id VARCHAR NOT NULL,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            ticker_i VARCHAR NOT NULL,
            ticker_j VARCHAR NOT NULL,
            correlation DOUBLE NOT NULL,
            CHECK (row_index <= column_index),
            CHECK (snapshot_id = '{escaped_snapshot_id}' OR correlation > 2.0),
            PRIMARY KEY (snapshot_id, row_index, column_index)
        );
        INSERT INTO preprocessing.snapshot_correlations_guarded
        SELECT * FROM preprocessing.snapshot_correlations;
        DROP TABLE preprocessing.snapshot_correlations;
        ALTER TABLE preprocessing.snapshot_correlations_guarded
        RENAME TO snapshot_correlations;
        """
    )


if __name__ == "__main__":
    unittest.main()
