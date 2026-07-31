from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

import duckdb
import pandas as pd
from openpyxl import load_workbook

from stat_arb_backtest import (
    BacktestConfig,
    BacktestMarketData,
    BacktestTarget,
    TargetWeight,
    calculate_period_performance,
    calculate_performance_metrics,
    export_backtest_workbook,
    run_backtest,
    simulate_backtest,
)
from stat_arb_backtest.cli import build_parser
from stat_arb_preprocessing import PreprocessingConfig, build_preprocessing
from tests.test_preprocessing import build_test_database


class BacktestCalculationTests(unittest.TestCase):
    def test_progress_bar_wraps_backtest_sessions_when_enabled(self) -> None:
        previous = date(2026, 1, 2)
        start = date(2026, 1, 5)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 101.0],
                "SPY": [100.0, 101.0],
            },
            index=pd.Index([previous, start], name="trade_date"),
        )
        target = BacktestTarget(
            start,
            1,
            1,
            (TargetWeight("AAA", 1.0),),
        )

        with patch(
            "stat_arb_backtest.calculations.tqdm",
            side_effect=lambda iterable, **_: iterable,
        ) as progress:
            simulate_backtest(
                BacktestMarketData(previous, (start,), closes),
                BacktestConfig(start, start),
                lambda _: target,
                show_progress=True,
            )

        progress.assert_called_once()
        self.assertEqual(progress.call_args.kwargs["desc"], "Backtest")
        self.assertEqual(progress.call_args.kwargs["unit"], "session")
        self.assertFalse(progress.call_args.kwargs["disable"])

    def test_fixed_positions_drift_between_rebalances(self) -> None:
        previous = date(2026, 1, 2)
        start = date(2026, 1, 5)
        end = date(2026, 1, 6)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 110.0, 121.0],
                "BBB": [100.0, 100.0, 100.0],
                "SPY": [100.0, 101.0, 102.0],
            },
            index=pd.Index([previous, start, end], name="trade_date"),
        )
        market_data = BacktestMarketData(
            previous_session=previous,
            sessions=(start, end),
            closes=closes,
        )
        target = BacktestTarget(
            as_of_date=start,
            cluster_count=1,
            active_cluster_count=1,
            weights=(
                TargetWeight(ticker="AAA", portfolio_weight=0.5),
                TargetWeight(ticker="BBB", portfolio_weight=0.5),
            ),
        )

        result = simulate_backtest(
            market_data,
            BacktestConfig(
                start_date=start,
                end_date=end,
                rebalance_period=3,
                take_profit_threshold=1.0,
            ),
            lambda as_of_date: target,
        )

        self.assertAlmostEqual(result.daily_performance[0].nav, 1.05)
        self.assertAlmostEqual(result.daily_performance[0].strategy_return, 0.05)
        self.assertAlmostEqual(result.daily_performance[1].nav, 1.105)
        self.assertAlmostEqual(
            result.daily_performance[1].strategy_return,
            1.105 / 1.05 - 1.0,
        )
        self.assertEqual(len(result.trades), 2)

    def test_all_cash_target_earns_zero_and_produces_no_infinite_ratios(
        self,
    ) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=2).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 110.0, 90.0],
                "SPY": [100.0, 101.0, 99.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        target = BacktestTarget(
            as_of_date=sessions[0],
            cluster_count=1,
            active_cluster_count=0,
            weights=(TargetWeight("AAA", 0.0),),
        )

        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                start_date=sessions[0],
                end_date=sessions[-1],
            ),
            lambda _: target,
        )

        self.assertEqual(
            [row.strategy_return for row in result.daily_performance],
            [0.0, 0.0],
        )
        self.assertEqual(
            [row.cash_weight for row in result.daily_performance],
            [1.0, 1.0],
        )
        self.assertEqual(result.strategy_metrics.ending_nav, 1.0)
        self.assertIsNone(result.strategy_metrics.sharpe_ratio)
        self.assertIsNone(result.strategy_metrics.sortino_ratio)

    def test_scheduled_rebalance_applies_new_target_after_third_return(self) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(
            pd.bdate_range("2026-01-05", periods=4).date
        )
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 110.0, 121.0, 133.1, 146.41],
                "BBB": [100.0, 100.0, 100.0, 100.0, 110.0],
                "SPY": [100.0, 100.0, 100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        initial = BacktestTarget(
            as_of_date=sessions[0],
            cluster_count=1,
            active_cluster_count=1,
            weights=(TargetWeight("AAA", 1.0),),
        )
        replacement = BacktestTarget(
            as_of_date=sessions[3],
            cluster_count=1,
            active_cluster_count=1,
            weights=(TargetWeight("BBB", 1.0),),
        )

        result = simulate_backtest(
            BacktestMarketData(
                previous_session=previous,
                sessions=sessions,
                closes=closes,
            ),
            BacktestConfig(
                start_date=sessions[0],
                end_date=sessions[-1],
                rebalance_period=3,
                take_profit_threshold=1.0,
            ),
            {sessions[0]: initial, sessions[3]: replacement}.__getitem__,
        )

        self.assertAlmostEqual(result.daily_performance[2].nav, 1.331)
        self.assertEqual(result.daily_performance[2].trigger_reason, "scheduled")
        self.assertAlmostEqual(result.daily_performance[3].strategy_return, 0.10)
        self.assertAlmostEqual(result.daily_performance[3].nav, 1.4641)
        self.assertEqual(
            [
                (event.reason, event.event_date, event.effective_date)
                for event in result.rebalance_events
            ],
            [
                ("initial", previous, sessions[0]),
                ("scheduled", sessions[2], sessions[3]),
            ],
        )

    def test_stop_win_at_exact_threshold_changes_next_session_target(self) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=3).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 105.0, 115.5, 127.05],
                "BBB": [100.0, 100.0, 102.0, 103.0],
                "SPY": [100.0, 100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        initial = BacktestTarget(
            sessions[0],
            1,
            1,
            (TargetWeight("AAA", 1.0),),
        )
        replacement = BacktestTarget(
            sessions[1],
            1,
            1,
            (TargetWeight("BBB", 1.0),),
        )

        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                start_date=sessions[0],
                end_date=sessions[-1],
                take_profit_threshold=0.05,
            ),
            {sessions[0]: initial, sessions[1]: replacement}.__getitem__,
        )

        self.assertEqual(result.daily_performance[0].trigger_reason, "stop_win")
        self.assertEqual(result.daily_performance[1].round_id, 2)
        self.assertEqual(result.daily_performance[1].holding_day, 1)
        self.assertAlmostEqual(result.daily_performance[1].strategy_return, 0.02)
        self.assertEqual(result.rebalance_events[1].event_date, sessions[0])
        self.assertEqual(result.rebalance_events[1].effective_date, sessions[1])
        self.assertEqual(result.rebalance_events[1].reason, "stop_win")

    def test_missing_position_is_frozen_then_liquidated_on_recovery(self) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=3).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, float("nan"), float("nan"), 120.0],
                "BBB": [100.0, 150.0, 150.0, 150.0],
                "CCC": [100.0, 100.0, 100.0, 100.0],
                "SPY": [100.0, 100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        initial = BacktestTarget(
            sessions[0],
            1,
            1,
            (TargetWeight("AAA", 0.5), TargetWeight("BBB", 0.5)),
        )
        replacement = BacktestTarget(
            sessions[1],
            1,
            1,
            (TargetWeight("CCC", 1.0),),
        )

        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                start_date=sessions[0],
                end_date=sessions[-1],
                rebalance_period=3,
                take_profit_threshold=0.20,
            ),
            {sessions[0]: initial, sessions[1]: replacement}.__getitem__,
        )

        self.assertEqual(result.daily_performance[0].trigger_reason, "stop_win")
        self.assertAlmostEqual(result.rebalance_events[1].frozen_value, 0.5)
        self.assertAlmostEqual(result.rebalance_events[1].available_capital, 0.75)
        self.assertAlmostEqual(result.daily_performance[1].nav, 1.25)
        self.assertAlmostEqual(result.daily_performance[2].nav, 1.35)
        recovery_sales = [
            trade
            for trade in result.trades
            if trade.reason == "recovery_liquidation"
        ]
        self.assertEqual(len(recovery_sales), 1)
        self.assertEqual(recovery_sales[0].ticker, "AAA")
        self.assertAlmostEqual(recovery_sales[0].trade_notional, 0.6)
        self.assertEqual(recovery_sales[0].execution_price, 120.0)
        recovery_lot = next(
            lot for lot in result.position_lots if lot.ticker == "AAA"
        )
        self.assertEqual(recovery_lot.status, "CLOSED")
        self.assertEqual(recovery_lot.final_sell_date, sessions[-1])
        self.assertEqual(recovery_lot.final_sell_price, 120.0)
        self.assertAlmostEqual(recovery_lot.lot_return or 0.0, 0.20)
        self.assertIn(
            "recovered_liquidated",
            [audit.action for audit in result.missing_data_audit],
        )
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "recovery_backtest.xlsx"
            export_backtest_workbook(result, output)
            workbook = load_workbook(output, read_only=False)
            actions = workbook["Portfolio_Actions"]
            recovery_row = next(
                row
                for row in actions.iter_rows(min_row=2, values_only=True)
                if row[13] == "recovery_liquidation"
            )
            self.assertEqual(recovery_row[3], "AAA")
            self.assertEqual(recovery_row[8], "SELL")
            workbook.close()

    def test_frozen_position_still_in_target_is_retained_on_recovery(self) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=2).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, float("nan"), 120.0],
                "BBB": [100.0, 150.0, 150.0],
                "CCC": [100.0, 100.0, 100.0],
                "SPY": [100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        targets = {
            sessions[0]: BacktestTarget(
                sessions[0],
                1,
                1,
                (TargetWeight("AAA", 0.5), TargetWeight("BBB", 0.5)),
            ),
            sessions[1]: BacktestTarget(
                sessions[1],
                1,
                1,
                (TargetWeight("AAA", 0.5), TargetWeight("CCC", 0.5)),
            ),
        }

        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                start_date=sessions[0],
                end_date=sessions[-1],
                take_profit_threshold=0.20,
            ),
            targets.__getitem__,
        )

        self.assertEqual(result.daily_performance[0].trigger_reason, "stop_win")
        self.assertAlmostEqual(result.daily_performance[1].nav, 1.35)
        self.assertEqual(result.daily_performance[1].position_count, 2)
        self.assertNotIn(
            "recovery_liquidation",
            [trade.reason for trade in result.trades],
        )
        self.assertIn(
            "frozen_retained_target",
            [audit.action for audit in result.missing_data_audit],
        )
        self.assertIn(
            "recovered_marked",
            [audit.action for audit in result.missing_data_audit],
        )

    def test_scheduled_reason_wins_when_threshold_is_reached_on_day_three(
        self,
    ) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=4).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 100.0, 100.0, 105.0, 105.0],
                "BBB": [100.0, 100.0, 100.0, 100.0, 100.0],
                "SPY": [100.0, 100.0, 100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        targets = {
            sessions[0]: BacktestTarget(
                sessions[0],
                1,
                1,
                (TargetWeight("AAA", 1.0),),
            ),
            sessions[3]: BacktestTarget(
                sessions[3],
                1,
                1,
                (TargetWeight("BBB", 1.0),),
            ),
        }

        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                start_date=sessions[0],
                end_date=sessions[-1],
                take_profit_threshold=0.05,
            ),
            targets.__getitem__,
        )

        self.assertIsNone(result.daily_performance[0].trigger_reason)
        self.assertIsNone(result.daily_performance[1].trigger_reason)
        self.assertEqual(result.daily_performance[2].trigger_reason, "scheduled")
        self.assertEqual(result.rebalance_events[1].reason, "scheduled")

    def test_performance_metrics_use_compounding_and_zero_risk_free_rate(
        self,
    ) -> None:
        returns = (0.02, -0.01, 0.03, -0.02)
        ending_nav = float(pd.Series(returns).add(1.0).prod())

        metrics = calculate_performance_metrics(
            returns,
            starting_nav=1.0,
            ending_nav=ending_nav,
            annualization_sessions=252,
        )

        expected_mean = float(pd.Series(returns).mean())
        expected_std = float(pd.Series(returns).std(ddof=1))
        expected_downside = float(
            pd.Series([-0.01, -0.02]).std(ddof=1)
        )
        self.assertAlmostEqual(metrics.total_return, ending_nav - 1.0)
        self.assertAlmostEqual(
            metrics.annualized_return,
            ending_nav ** (252 / 4) - 1.0,
        )
        self.assertAlmostEqual(
            metrics.sharpe_ratio,
            (252**0.5) * expected_mean / expected_std,
        )
        self.assertAlmostEqual(
            metrics.sortino_ratio,
            (252**0.5) * expected_mean / expected_downside,
        )

    def test_period_performance_compounds_calendar_years_and_months(self) -> None:
        previous = date(2025, 12, 29)
        sessions = (
            date(2025, 12, 30),
            date(2025, 12, 31),
            date(2026, 1, 2),
            date(2026, 2, 2),
        )
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 110.0, 99.0, 108.9, 103.455],
                "SPY": [100.0, 102.0, 100.98, 104.0094, 101.929212],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        target = BacktestTarget(
            sessions[0],
            1,
            1,
            (TargetWeight("AAA", 1.0),),
        )
        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                sessions[0],
                sessions[-1],
                rebalance_period=10,
                take_profit_threshold=10.0,
            ),
            lambda _: target,
        )

        annual = calculate_period_performance(
            result.daily_performance,
            frequency="year",
            annualization_sessions=252,
        )
        monthly = calculate_period_performance(
            result.daily_performance,
            frequency="month",
            annualization_sessions=252,
        )

        self.assertEqual(
            [(row.period_start.year, row.session_count) for row in annual],
            [(2025, 2), (2026, 2)],
        )
        self.assertAlmostEqual(annual[0].strategy_return, -0.01)
        self.assertAlmostEqual(annual[0].spy_return, 0.0098)
        self.assertAlmostEqual(annual[0].excess_return, -0.0198)
        self.assertAlmostEqual(annual[0].strategy_max_drawdown, -0.10)
        self.assertAlmostEqual(annual[1].strategy_return, 0.045)
        self.assertAlmostEqual(annual[1].strategy_max_drawdown, -0.05)
        self.assertEqual(
            [
                (row.period_start.year, row.period_start.month)
                for row in monthly
            ],
            [(2025, 12), (2026, 1), (2026, 2)],
        )
        self.assertAlmostEqual(monthly[1].strategy_return, 0.10)
        self.assertAlmostEqual(monthly[2].strategy_return, -0.05)

    def test_unfilled_initial_buy_stays_cash_and_is_audited(self) -> None:
        previous = date(2026, 1, 2)
        start = date(2026, 1, 5)
        closes = pd.DataFrame(
            {
                "AAA": [float("nan"), 100.0],
                "SPY": [100.0, 101.0],
            },
            index=pd.Index((previous, start), name="trade_date"),
        )
        target = BacktestTarget(
            start,
            1,
            1,
            (TargetWeight("AAA", 1.0),),
        )

        result = simulate_backtest(
            BacktestMarketData(previous, (start,), closes),
            BacktestConfig(start, start),
            lambda _: target,
        )

        self.assertAlmostEqual(result.daily_performance[0].nav, 1.0)
        self.assertAlmostEqual(result.daily_performance[0].cash_weight, 1.0)
        self.assertEqual(len(result.trades), 1)
        self.assertEqual(result.trades[0].side, "BUY")
        self.assertEqual(result.trades[0].status, "unfilled_missing_close")
        self.assertAlmostEqual(result.trades[0].trade_notional, 1.0)
        self.assertIsNone(result.trades[0].execution_price)
        self.assertEqual(result.trades[0].units_traded, 0.0)
        self.assertEqual(result.trades[0].executed_notional, 0.0)
        self.assertEqual(result.position_lots, ())
        self.assertEqual(
            [audit.action for audit in result.missing_data_audit],
            ["buy_unfilled_cash_retained"],
        )

    def test_target_weight_audit_keeps_only_positive_weights(self) -> None:
        previous = date(2026, 1, 2)
        start = date(2026, 1, 5)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 110.0],
                "BBB": [100.0, 90.0],
                "CCC": [100.0, 105.0],
                "SPY": [100.0, 101.0],
            },
            index=pd.Index((previous, start), name="trade_date"),
        )
        target = BacktestTarget(
            start,
            2,
            1,
            (
                TargetWeight(
                    "AAA",
                    0.5,
                    market_cap_rank=1,
                    cluster_id=0,
                    cumulative_deviation=-0.10,
                    classification="previous_loser",
                    local_weight=1.0,
                ),
                TargetWeight(
                    "BBB",
                    0.0,
                    market_cap_rank=2,
                    cluster_id=0,
                    cumulative_deviation=0.10,
                    classification="previous_winner",
                    local_weight=0.0,
                ),
                TargetWeight(
                    "CCC",
                    0.0,
                    market_cap_rank=3,
                    cluster_id=1,
                    cumulative_deviation=0.0,
                    classification="neutral",
                    local_weight=0.0,
                ),
            ),
        )

        result = simulate_backtest(
            BacktestMarketData(previous, (start,), closes),
            BacktestConfig(start, start),
            lambda _: target,
        )

        self.assertEqual(
            [row.ticker for row in result.target_weights],
            ["AAA"],
        )
        self.assertTrue(
            all(row.portfolio_weight > 0.0 for row in result.target_weights)
        )
        self.assertAlmostEqual(result.daily_performance[0].nav, 1.05)
        self.assertAlmostEqual(
            result.rebalance_events[0].target_gross_exposure,
            0.5,
        )
        self.assertEqual(
            result.calculation_version,
            "long_only_close_to_close_stateful_v2",
        )

    def test_fifo_lots_match_partial_and_cross_lot_sales(self) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=4).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 100.0, 200.0, 250.0, 250.0],
                "SPY": [100.0, 100.0, 100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        targets = {
            sessions[0]: BacktestTarget(
                sessions[0],
                1,
                1,
                (TargetWeight("AAA", 0.4),),
            ),
            sessions[1]: BacktestTarget(
                sessions[1],
                1,
                1,
                (TargetWeight("AAA", 0.8),),
            ),
            sessions[2]: BacktestTarget(
                sessions[2],
                1,
                1,
                (TargetWeight("AAA", 0.5),),
            ),
            sessions[3]: BacktestTarget(
                sessions[3],
                1,
                0,
                (TargetWeight("AAA", 0.0),),
            ),
        }

        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                sessions[0],
                sessions[-1],
                rebalance_period=1,
                take_profit_threshold=10.0,
            ),
            targets.__getitem__,
        )

        self.assertEqual(
            [trade.trade_id for trade in result.trades],
            [1, 2, 3, 4],
        )
        self.assertEqual(
            [trade.side for trade in result.trades],
            ["BUY", "BUY", "SELL", "SELL"],
        )
        self.assertTrue(
            all(
                abs(
                    trade.execution_price * trade.units_traded
                    - trade.executed_notional
                )
                < 1e-12
                for trade in result.trades
                if trade.execution_price is not None
            )
        )
        for trade in result.trades:
            expected_units = (
                trade.units_after - trade.units_before
                if trade.side == "BUY"
                else trade.units_before - trade.units_after
            )
            self.assertAlmostEqual(trade.units_traded, expected_units)
        first, second = result.position_lots
        self.assertEqual(first.lot_id, "LOT000001")
        self.assertEqual(first.buy_trade_id, 1)
        self.assertEqual(first.first_sell_date, sessions[1])
        self.assertEqual(first.final_sell_date, sessions[2])
        self.assertEqual(first.final_sell_price, 250.0)
        self.assertAlmostEqual(first.matched_sell_vwap or 0.0, 206.25)
        self.assertAlmostEqual(first.sale_proceeds, 0.825)
        self.assertAlmostEqual(first.realized_pnl, 0.425)
        self.assertAlmostEqual(first.realized_return or 0.0, 1.0625)
        self.assertAlmostEqual(first.lot_return or 0.0, 1.0625)
        self.assertEqual(first.status, "CLOSED")
        self.assertEqual(second.buy_trade_id, 2)
        self.assertEqual(second.first_sell_date, sessions[2])
        self.assertEqual(second.final_sell_date, sessions[2])
        self.assertEqual(second.final_sell_price, 250.0)
        self.assertAlmostEqual(second.lot_return or 0.0, 1.5)
        self.assertEqual(second.status, "CLOSED")
        self.assertEqual(result.fifo_reconciliation_status, "OK")

    def test_fifo_lots_leave_partial_and_open_rows_unclosed_at_end(self) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=3).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 100.0, 200.0, 200.0],
                "SPY": [100.0, 100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        targets = {
            sessions[0]: BacktestTarget(
                sessions[0],
                1,
                1,
                (TargetWeight("AAA", 0.4),),
            ),
            sessions[1]: BacktestTarget(
                sessions[1],
                1,
                1,
                (TargetWeight("AAA", 0.8),),
            ),
            sessions[2]: BacktestTarget(
                sessions[2],
                1,
                1,
                (TargetWeight("AAA", 0.5),),
            ),
        }

        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                sessions[0],
                sessions[-1],
                rebalance_period=1,
                take_profit_threshold=10.0,
            ),
            targets.__getitem__,
        )

        first, second = result.position_lots
        self.assertEqual(first.status, "PARTIALLY_CLOSED")
        self.assertEqual(first.first_sell_date, sessions[1])
        self.assertIsNone(first.final_sell_date)
        self.assertIsNone(first.final_sell_price)
        self.assertAlmostEqual(first.realized_return or 0.0, 1.0)
        self.assertIsNone(first.lot_return)
        self.assertEqual(second.status, "OPEN")
        self.assertEqual(second.sold_units, 0.0)
        self.assertIsNone(second.realized_return)
        self.assertIsNone(second.lot_return)
        net_trade_units = sum(
            trade.units_traded
            if trade.side == "BUY"
            else -trade.units_traded
            for trade in result.trades
            if trade.status == "executed"
        )
        self.assertAlmostEqual(
            sum(lot.remaining_units for lot in result.position_lots),
            net_trade_units,
        )


class BacktestIntegrationTests(unittest.TestCase):
    def test_run_backtest_rolls_non_trading_start_to_next_spy_session(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            add_close_prices(database)
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
                correlation_window=5,
            )
            build_preprocessing(preprocessing_config)
            calendar = tuple(value.date() for value in dates)
            effective_start_index = next(
                index
                for index in range(60, len(calendar) - 3)
                if calendar[index].weekday() == 0
            )
            effective_start = calendar[effective_start_index]
            requested_start = effective_start - timedelta(days=1)
            end = calendar[effective_start_index + 3]

            result = run_backtest(
                preprocessing_config,
                BacktestConfig(
                    start_date=requested_start,
                    end_date=end,
                    take_profit_threshold=1.0,
                ),
            )

            expected_sessions = calendar[
                effective_start_index : effective_start_index + 4
            ]
            self.assertEqual(result.config.start_date, effective_start)
            self.assertEqual(
                tuple(row.trade_date for row in result.daily_performance),
                expected_sessions,
            )
            self.assertEqual(
                result.rebalance_events[0].effective_date,
                effective_start,
            )

    def test_run_backtest_uses_existing_pipeline_without_result_tables(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            add_close_prices(database)
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
                correlation_window=5,
            )
            build_preprocessing(preprocessing_config)
            sessions = tuple(value.date() for value in dates[-4:])

            result = run_backtest(
                preprocessing_config,
                BacktestConfig(
                    start_date=sessions[0],
                    end_date=sessions[-1],
                    take_profit_threshold=1.0,
                ),
            )

            self.assertEqual(
                tuple(row.trade_date for row in result.daily_performance),
                sessions,
            )
            self.assertEqual(
                [event.reason for event in result.rebalance_events],
                ["initial", "scheduled"],
            )
            with duckdb.connect(str(database), read_only=True) as connection:
                backtest_tables = connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM information_schema.tables
                    WHERE table_schema IN ('backtest', 'portfolio_backtest')
                       OR table_name LIKE '%backtest%'
                    """
                ).fetchone()[0]
            self.assertEqual(backtest_tables, 0)

    def test_excel_contains_period_performance_and_refuses_overwrite(self) -> None:
        previous = date(2026, 1, 2)
        start = date(2026, 1, 5)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 105.0],
                "BBB": [100.0, 95.0],
                "SPY": [100.0, 102.0],
            },
            index=pd.Index((previous, start), name="trade_date"),
        )
        target = BacktestTarget(
            start,
            1,
            1,
            (
                TargetWeight(
                    "AAA",
                    1.0,
                    market_cap_rank=1,
                    cluster_id=0,
                    cumulative_deviation=-0.10,
                    classification="previous_loser",
                    local_weight=1.0,
                ),
                TargetWeight(
                    "BBB",
                    0.0,
                    market_cap_rank=2,
                    cluster_id=0,
                    cumulative_deviation=0.10,
                    classification="previous_winner",
                    local_weight=0.0,
                ),
            ),
        )
        result = simulate_backtest(
            BacktestMarketData(previous, (start,), closes),
            BacktestConfig(start, start),
            lambda _: target,
        )

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "backtest.xlsx"
            export_backtest_workbook(result, output)

            workbook = load_workbook(output, data_only=False, read_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary",
                    "Period_Performance",
                    "Daily_Performance",
                    "Rebalance_Events",
                    "Portfolio_Actions",
                    "Position_Lots",
                    "Missing_Data_Audit",
                ],
            )
            summary = workbook["Summary"]
            self.assertAlmostEqual(
                summary["B7"].value,
                result.strategy_metrics.total_return,
            )
            periods = workbook["Period_Performance"]
            self.assertEqual(periods["A4"].value, "Annual Performance")
            self.assertEqual(periods["A5"].value, "Period")
            self.assertEqual(periods["A6"].value.year, 2026)
            self.assertAlmostEqual(periods["C6"].value, 0.05)
            self.assertEqual(
                periods["A9"].value,
                "Monthly Strategy Return Heatmap",
            )
            self.assertAlmostEqual(periods["B11"].value, 0.05)
            self.assertAlmostEqual(periods["N11"].value, 0.05)
            self.assertEqual(len(list(periods.conditional_formatting)), 6)
            self.assertEqual(workbook["Daily_Performance"].max_row, 2)
            self.assertEqual(workbook["Rebalance_Events"]["C2"].value, "initial")
            action_sheet = workbook["Portfolio_Actions"]
            self.assertEqual(action_sheet.max_row, 2)
            self.assertEqual(action_sheet["D2"].value, "AAA")
            self.assertGreater(action_sheet["H2"].value, 0.0)
            self.assertEqual(action_sheet["I2"].value, "BUY")
            self.assertEqual(action_sheet["J2"].value, 100.0)
            self.assertAlmostEqual(action_sheet["K2"].value, 0.01)
            self.assertAlmostEqual(action_sheet["L2"].value, 1.0)
            lot_sheet = workbook["Position_Lots"]
            self.assertEqual(lot_sheet.max_row, 2)
            self.assertEqual(lot_sheet["A2"].value, "LOT000001")
            self.assertEqual(lot_sheet["D2"].value, 100.0)
            self.assertEqual(lot_sheet["L2"].value, "OPEN")
            self.assertEqual(workbook["Missing_Data_Audit"].max_row, 1)
            self.assertEqual(
                summary["B21"].value,
                "1 / 1",
            )
            self.assertEqual(
                summary["D21"].value,
                1,
            )
            self.assertEqual(
                summary["D23"].value,
                "OK",
            )
            self.assertEqual(
                workbook["Daily_Performance"].freeze_panes,
                "A2",
            )
            self.assertEqual(
                workbook["Daily_Performance"].auto_filter.ref,
                "A1:P2",
            )
            self.assertEqual(
                workbook["Daily_Performance"]["A2"].number_format,
                "yyyy-mm-dd",
            )
            self.assertEqual(
                workbook["Daily_Performance"]["B2"].number_format,
                "0.0000%",
            )
            self.assertEqual(action_sheet["C2"].number_format, "yyyy-mm-dd")
            self.assertEqual(action_sheet["L2"].number_format, "0.000000")
            self.assertEqual(action_sheet.column_dimensions["A"].width, 11.0)
            self.assertEqual(action_sheet["A1"].fill.fill_type, "solid")
            self.assertTrue(
                str(action_sheet["A1"].fill.fgColor.rgb).endswith("1F4E78")
            )
            self.assertEqual(action_sheet["A2"].fill.fill_type, "solid")
            self.assertTrue(
                str(action_sheet["A2"].fill.fgColor.rgb).endswith("F8FAFC")
            )
            action_conditional_ranges = list(
                action_sheet.conditional_formatting
            )
            self.assertEqual(len(action_conditional_ranges), 2)
            side_rules = action_sheet.conditional_formatting[
                action_conditional_ranges[0]
            ]
            self.assertEqual(len(side_rules), 2)
            self.assertEqual(
                side_rules[0].formula,
                ['I2="BUY"'],
            )
            self.assertTrue(
                all(
                    action_sheet.column_dimensions[column].hidden
                    for column in "OPQRSTUVW"
                )
            )
            self.assertTrue(
                all(
                    workbook["Daily_Performance"]
                    .column_dimensions[column]
                    .hidden
                    for column in "MNOP"
                )
            )
            self.assertEqual(
                workbook["Missing_Data_Audit"].sheet_state,
                "hidden",
            )
            workbook.close()

            with self.assertRaises(FileExistsError):
                export_backtest_workbook(result, output)

    def test_excel_export_removes_temporary_file_after_save_failure(self) -> None:
        previous = date(2026, 1, 2)
        start = date(2026, 1, 5)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 101.0],
                "SPY": [100.0, 101.0],
            },
            index=pd.Index((previous, start), name="trade_date"),
        )
        target = BacktestTarget(
            start,
            1,
            1,
            (TargetWeight("AAA", 1.0),),
        )
        result = simulate_backtest(
            BacktestMarketData(previous, (start,), closes),
            BacktestConfig(start, start),
            lambda _: target,
        )

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "backtest.xlsx"
            with patch(
                "openpyxl.workbook.workbook.Workbook.save",
                side_effect=RuntimeError("save failed"),
            ):
                with self.assertRaisesRegex(RuntimeError, "save failed"):
                    export_backtest_workbook(result, output)

            self.assertFalse(output.exists())
            self.assertEqual(list(Path(directory).iterdir()), [])

    def test_excel_actions_keep_target_only_and_sell_only_rows(self) -> None:
        previous = date(2026, 1, 2)
        sessions = tuple(pd.bdate_range("2026-01-05", periods=3).date)
        closes = pd.DataFrame(
            {
                "AAA": [100.0, 100.0, 100.0, 100.0],
                "BBB": [100.0, 100.0, 100.0, 100.0],
                "SPY": [100.0, 100.0, 100.0, 100.0],
            },
            index=pd.Index((previous, *sessions), name="trade_date"),
        )
        targets = {
            sessions[0]: BacktestTarget(
                sessions[0],
                1,
                1,
                (TargetWeight("AAA", 1.0),),
            ),
            sessions[1]: BacktestTarget(
                sessions[1],
                1,
                1,
                (TargetWeight("AAA", 1.0),),
            ),
            sessions[2]: BacktestTarget(
                sessions[2],
                1,
                1,
                (TargetWeight("BBB", 1.0),),
            ),
        }
        result = simulate_backtest(
            BacktestMarketData(previous, sessions, closes),
            BacktestConfig(
                sessions[0],
                sessions[-1],
                rebalance_period=1,
                take_profit_threshold=10.0,
            ),
            targets.__getitem__,
        )

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "backtest.xlsx"
            export_backtest_workbook(result, output)
            workbook = load_workbook(output, data_only=False, read_only=False)
            actions = workbook["Portfolio_Actions"]
            rows = [
                tuple(
                    actions.cell(row, column).value
                    for column in range(1, actions.max_column + 1)
                )
                for row in range(2, actions.max_row + 1)
            ]
            target_only = next(
                row
                for row in rows
                if row[0] == 2 and row[3] == "AAA"
            )
            self.assertEqual(target_only[7], 1.0)
            self.assertIsNone(target_only[14])
            self.assertIsNone(target_only[8])
            sell_only = next(
                row
                for row in rows
                if row[0] == 3
                and row[3] == "AAA"
                and row[8] == "SELL"
            )
            self.assertEqual(sell_only[7], 0.0)
            self.assertEqual(sell_only[9], 100.0)
            self.assertEqual(sell_only[11], 1.0)
            workbook.close()

    def test_cli_requires_explicit_range_and_uses_project_defaults(self) -> None:
        parser = build_parser()
        args = parser.parse_args(
            [
                "export",
                "--start-date",
                "2026-01-05",
                "--end-date",
                "2026-01-09",
            ]
        )

        self.assertEqual(args.rebalance_period, 3)
        self.assertAlmostEqual(args.take_profit_threshold, 0.05)
        self.assertAlmostEqual(args.deviation_threshold, 0.05)
        self.assertEqual(args.lookback_window, 5)
        self.assertFalse(args.no_progress)


def add_close_prices(database: Path) -> None:
    with duckdb.connect(str(database)) as connection:
        prices = connection.execute(
            """
            SELECT ticker, trade_date, price_return
            FROM market_data.daily_prices
            ORDER BY ticker, trade_date
            """
        ).fetchdf()
        prices["close"] = (
            prices.assign(growth=1.0 + prices["price_return"])
            .groupby("ticker")["growth"]
            .cumprod()
            .mul(100.0)
        )
        market = connection.execute(
            """
            SELECT trade_date, market_return
            FROM market_data.market_returns
            WHERE ticker = 'SPY'
            ORDER BY trade_date
            """
        ).fetchdf()
        market["close"] = (1.0 + market["market_return"]).cumprod() * 100.0
        connection.register("close_prices", prices)
        connection.register("market_closes", market)
        try:
            connection.execute(
                """
                UPDATE market_data.daily_prices AS target
                SET close = source.close
                FROM close_prices AS source
                WHERE target.ticker = source.ticker
                  AND target.trade_date = source.trade_date
                """
            )
            connection.execute(
                """
                UPDATE market_data.market_returns AS target
                SET close = source.close
                FROM market_closes AS source
                WHERE target.ticker = 'SPY'
                  AND target.trade_date = source.trade_date
                """
            )
        finally:
            connection.unregister("close_prices")
            connection.unregister("market_closes")


if __name__ == "__main__":
    unittest.main()
