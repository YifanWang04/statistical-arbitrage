from __future__ import annotations

from datetime import date
from pathlib import Path
import tempfile
import unittest

import duckdb
import numpy as np
from openpyxl import load_workbook
import pandas as pd

from stat_arb_cluster_count import (
    DEFAULT_CLUSTER_COUNT_ESTIMATION_WINDOW,
    calculate_cluster_count,
    calculate_cluster_count_for_date,
)
from stat_arb_cluster_count.excel import export_cluster_count_workbook
from stat_arb_preprocessing import (
    DEFAULT_CLUSTERING_CORRELATION_WINDOW,
    PreprocessingConfig,
    build_preprocessing,
)
from stat_arb_preprocessing.models import PreprocessingSnapshot, SnapshotQuality
from tests.test_preprocessing import build_test_database


class ClusterCountCalculationTests(unittest.TestCase):
    def test_selects_minimum_k_that_reaches_threshold(self) -> None:
        snapshot = make_snapshot(
            np.array(
                [
                    [1.0, 0.5, 0.5],
                    [0.5, 1.0, 0.5],
                    [0.5, 0.5, 1.0],
                ]
            )
        )

        result = calculate_cluster_count(snapshot, 0.80)

        np.testing.assert_allclose(result.raw_eigenvalues, (2.0, 0.5, 0.5))
        self.assertAlmostEqual(result.total_variance, 3.0)
        self.assertEqual(result.selected_k, 2)
        self.assertLess(result.cumulative_explained_ratio[0], 0.80)
        self.assertGreaterEqual(result.cumulative_explained_ratio[1], 0.80)
        self.assertEqual(result.quality.numerical_rank, 3)

    def test_threshold_hit_exactly_uses_the_first_eigenvalue(self) -> None:
        snapshot = make_snapshot(
            np.array(
                [
                    [1.0, 0.5, 0.5],
                    [0.5, 1.0, 0.5],
                    [0.5, 0.5, 1.0],
                ]
            )
        )

        result = calculate_cluster_count(snapshot, 2.0 / 3.0)

        self.assertEqual(result.selected_k, 1)

    def test_ratio_just_below_threshold_does_not_cross_early(self) -> None:
        correlation = 0.80 - 1e-12
        snapshot = make_snapshot(
            np.array(
                [
                    [1.0, correlation],
                    [correlation, 1.0],
                ]
            )
        )

        result = calculate_cluster_count(snapshot, 0.90)

        self.assertLess(result.cumulative_explained_ratio[0], 0.90)
        self.assertEqual(result.selected_k, 2)

    def test_threshold_one_uses_all_positive_identity_eigenvalues(self) -> None:
        result = calculate_cluster_count(make_snapshot(np.eye(4)), 1.0)

        self.assertEqual(result.selected_k, 4)
        self.assertEqual(result.cumulative_explained_ratio[-1], 1.0)

    def test_tiny_negative_eigenvalue_is_disclosed_and_zeroed(self) -> None:
        correlation = np.full((3, 3), -0.500000000025)
        np.fill_diagonal(correlation, 1.0)

        result = calculate_cluster_count(make_snapshot(correlation), 0.90)

        self.assertLess(result.raw_eigenvalues[-1], 0.0)
        self.assertGreaterEqual(result.raw_eigenvalues[-1], -1e-10)
        self.assertEqual(result.effective_eigenvalues[-1], 0.0)
        self.assertEqual(result.quality.adjusted_negative_eigenvalue_count, 1)

    def test_rejects_invalid_thresholds(self) -> None:
        snapshot = make_snapshot(np.eye(2))
        for threshold in (0.0, -0.1, 1.01, float("nan"), float("inf"), "bad"):
            with self.subTest(threshold=threshold):
                with self.assertRaisesRegex(ValueError, "variance_threshold"):
                    calculate_cluster_count(snapshot, threshold)  # type: ignore[arg-type]

    def test_rejects_invalid_correlation_matrices(self) -> None:
        cases = {
            "row labels": make_snapshot(np.eye(2)).correlation_matrix.rename(
                index={"AAA": "BAD"}
            ),
            "symmetric": pd.DataFrame(
                [[1.0, 0.2], [0.1, 1.0]],
                index=("AAA", "BBB"),
                columns=("AAA", "BBB"),
            ),
            "diagonal": pd.DataFrame(
                [[0.9, 0.0], [0.0, 1.0]],
                index=("AAA", "BBB"),
                columns=("AAA", "BBB"),
            ),
            "finite": pd.DataFrame(
                [[1.0, np.nan], [np.nan, 1.0]],
                index=("AAA", "BBB"),
                columns=("AAA", "BBB"),
            ),
            "positive semidefinite": pd.DataFrame(
                [
                    [1.0, -0.6, -0.6],
                    [-0.6, 1.0, -0.6],
                    [-0.6, -0.6, 1.0],
                ],
                index=("AAA", "BBB", "CCC"),
                columns=("AAA", "BBB", "CCC"),
            ),
        }
        expected_messages = {
            "row labels": "row labels",
            "symmetric": "not symmetric",
            "diagonal": "diagonal",
            "finite": "non-finite",
            "positive semidefinite": "positive semidefinite",
        }

        for name, matrix in cases.items():
            with self.subTest(name=name):
                snapshot = make_snapshot(np.eye(len(matrix)))
                object.__setattr__(snapshot, "correlation_matrix", matrix)
                with self.assertRaisesRegex(ValueError, expected_messages[name]):
                    calculate_cluster_count(snapshot)


class ClusterCountIntegrationTests(unittest.TestCase):
    def test_on_demand_calculation_does_not_persist_snapshot_or_cluster_count(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            # The shared fixture has 70 sessions. A shorter beta window leaves
            # enough valid residual history to exercise the independent
            # 20-session cluster-count window.
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
            )
            build_preprocessing(preprocessing_config)
            as_of_date = dates[-1].date()

            with duckdb.connect(str(database), read_only=True) as connection:
                before = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.correlation_snapshots"
                ).fetchone()[0]

            result = calculate_cluster_count_for_date(
                preprocessing_config,
                as_of_date,
                cluster_count_estimation_window=(
                    DEFAULT_CLUSTER_COUNT_ESTIMATION_WINDOW
                ),
            )

            with duckdb.connect(str(database), read_only=True) as connection:
                after = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.correlation_snapshots"
                ).fetchone()[0]
                cluster_count_tables = connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM information_schema.tables
                    WHERE table_schema = 'cluster_count'
                       OR table_name LIKE 'cluster_count%'
                    """
                ).fetchone()[0]
            self.assertEqual(before, 0)
            self.assertEqual(after, before)
            self.assertEqual(cluster_count_tables, 0)
            self.assertEqual(result.as_of_date, as_of_date)
            self.assertEqual(
                preprocessing_config.correlation_window,
                DEFAULT_CLUSTERING_CORRELATION_WINDOW,
            )
            self.assertEqual(
                result.cluster_count_estimation_window,
                DEFAULT_CLUSTER_COUNT_ESTIMATION_WINDOW,
            )
            self.assertGreaterEqual(result.selected_k, 1)

    def test_excel_has_two_auditable_sheets_and_refuses_silent_overwrite(self) -> None:
        result = calculate_cluster_count(
            make_snapshot(
                np.array(
                    [
                        [1.0, 0.5, 0.5],
                        [0.5, 1.0, 0.5],
                        [0.5, 0.5, 1.0],
                    ]
                )
            ),
            0.80,
        )
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "cluster_count.xlsx"

            export_cluster_count_workbook(result, output)

            workbook = load_workbook(output, data_only=False, read_only=False)
            self.assertEqual(workbook.sheetnames, ["Eigenvalues", "K_Calculation"])
            eigenvalues = workbook["Eigenvalues"]
            calculation = workbook["K_Calculation"]
            self.assertEqual(eigenvalues.max_row, result.stock_count + 1)
            self.assertEqual(
                [eigenvalues.cell(row, 1).value for row in range(2, 5)],
                [1, 2, 3],
            )
            raw = [eigenvalues.cell(row, 2).value for row in range(2, 5)]
            self.assertEqual(raw, sorted(raw, reverse=True))
            self.assertEqual(calculation["B12"].value, 0.80)
            self.assertEqual(
                calculation["A9"].value,
                "Cluster-count estimation window",
            )
            self.assertEqual(
                calculation["B9"].value,
                result.cluster_count_estimation_window,
            )
            self.assertEqual(calculation["B20"].value, result.selected_k)
            self.assertEqual(calculation["C15"].value, "=SUM('Eigenvalues'!B2:B4)")
            self.assertEqual(calculation["B25"].value, "='Eigenvalues'!C2")
            self.assertEqual(calculation["C25"].value, "=SUM($B$25:B25)")
            self.assertEqual(calculation["D25"].value, "=C25/$C$16")
            self.assertEqual(calculation["E25"].value, '=IF(D25>=$B$12,"YES","NO")')
            self.assertEqual(calculation["F25"].value, '=IF(A25<=$B$20,"YES","NO")')
            self.assertTrue(str(calculation["C20"].value).startswith("=COUNTIF"))
            self.assertEqual(calculation.freeze_panes, "A25")
            workbook.close()

            with self.assertRaises(FileExistsError):
                export_cluster_count_workbook(result, output)


def make_snapshot(correlation: np.ndarray) -> PreprocessingSnapshot:
    size = int(correlation.shape[0])
    tickers = tuple(chr(ord("A") + index) * 3 for index in range(size))
    frame = pd.DataFrame(correlation, index=tickers, columns=tickers)
    values = np.asarray(correlation, dtype=float)
    finite_values = values[np.isfinite(values)]
    eigenvalues = (
        np.linalg.eigvalsh((values + values.T) / 2.0)
        if bool(np.isfinite(values).all())
        else np.array([float("nan")])
    )
    quality = SnapshotQuality(
        maximum_asymmetry=float(np.nanmax(np.abs(values - values.T))),
        minimum_correlation=float(np.min(finite_values)) if finite_values.size else 0.0,
        maximum_correlation=float(np.max(finite_values)) if finite_values.size else 0.0,
        minimum_eigenvalue=float(np.nanmin(eigenvalues)),
        numerical_rank=int(np.linalg.matrix_rank(np.nan_to_num(values))),
        has_non_finite_values=not bool(np.isfinite(values).all()),
    )
    empty = pd.DataFrame()
    return PreprocessingSnapshot(
        snapshot_id="in-memory-test-snapshot",
        preprocessing_run_id="test-preprocessing-run",
        as_of_date=date(2026, 7, 17),
        window_start=date(2026, 7, 10),
        window_end=date(2026, 7, 16),
        beta_window=60,
        correlation_window=5,
        beta_alignment="include_current_session",
        missing_policy="complete_window",
        calculation_version="paper_baseline_v1",
        variance_epsilon=1e-15,
        return_basis="split_consistent_close_price_return_excluding_dividends",
        tickers=tickers,
        market_cap_ranks=tuple(range(1, size + 1)),
        beta_matrix=empty,
        stock_return_matrix=empty,
        market_returns=pd.Series(dtype=float),
        residual_matrix=empty,
        correlation_matrix=frame,
        exclusions=empty,
        selected_stock_count=size,
        quality=quality,
    )


if __name__ == "__main__":
    unittest.main()
