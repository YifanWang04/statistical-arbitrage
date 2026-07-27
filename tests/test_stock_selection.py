from __future__ import annotations

from dataclasses import replace
from pathlib import Path
import tempfile
import unittest

import duckdb
import numpy as np
from openpyxl import load_workbook
import pandas as pd

from stat_arb_clustering import cluster_sponge_sym
from stat_arb_preprocessing import PreprocessingConfig, build_preprocessing
from stat_arb_stock_selection import (
    NEUTRAL,
    PREVIOUS_LOSER,
    PREVIOUS_WINNER,
    StockSelectionConfig,
    identify_stocks_for_date,
    identify_stocks_to_trade,
)
from stat_arb_stock_selection.excel import export_stock_selection_workbook
from tests.test_cluster_count import make_snapshot
from tests.test_preprocessing import build_test_database


class StockSelectionCalculationTests(unittest.TestCase):
    def test_classifies_cumulative_raw_return_deviations_with_strict_thresholds(
        self,
    ) -> None:
        snapshot = make_snapshot(np.eye(4))
        clustering = cluster_sponge_sym(snapshot, 1)
        returns = pd.DataFrame(
            {
                "AAA": [0.03] * 5,
                "BBB": [0.01] * 5,
                "CCC": [-0.01] * 5,
                "DDD": [-0.03] * 5,
            },
            index=pd.date_range("2026-07-10", periods=5, freq="B"),
        )

        result = identify_stocks_to_trade(
            clustering,
            returns,
            StockSelectionConfig(deviation_threshold=0.06),
        )

        np.testing.assert_allclose(
            result.cluster_mean_return_matrix.to_numpy(),
            np.zeros((5, 1)),
            atol=1e-15,
        )
        np.testing.assert_allclose(
            result.cumulative_deviations,
            (0.15, 0.05, -0.05, -0.15),
        )
        self.assertEqual(
            result.classifications,
            (PREVIOUS_WINNER, NEUTRAL, NEUTRAL, PREVIOUS_LOSER),
        )
        self.assertEqual(result.winner_tickers, ("AAA",))
        self.assertEqual(result.loser_tickers, ("DDD",))
        self.assertLess(result.quality.maximum_daily_cluster_sum_error, 1e-15)
        self.assertLess(
            result.quality.maximum_cumulative_cluster_sum_error,
            1e-15,
        )

    def test_uses_daily_cluster_means_before_cumulating(self) -> None:
        snapshot = make_snapshot(np.eye(4))
        clustering = replace(
            cluster_sponge_sym(snapshot, 1),
            requested_cluster_count=2,
            cluster_labels=(0, 0, 1, 1),
            cluster_sizes=(2, 2),
        )
        returns = pd.DataFrame(
            {
                "AAA": [0.04, 0.00, 0.02, 0.00, 0.02],
                "BBB": [0.00, 0.00, 0.00, 0.00, 0.00],
                "CCC": [0.01, 0.03, 0.01, 0.03, 0.01],
                "DDD": [0.03, 0.01, 0.03, 0.01, 0.03],
            },
            index=pd.date_range("2026-07-10", periods=5, freq="B"),
        )

        result = identify_stocks_to_trade(clustering, returns)

        np.testing.assert_allclose(
            result.cluster_mean_return_matrix[0],
            (0.02, 0.00, 0.01, 0.00, 0.01),
        )
        np.testing.assert_allclose(
            result.cluster_mean_return_matrix[1],
            (0.02, 0.02, 0.02, 0.02, 0.02),
        )
        np.testing.assert_allclose(
            result.cumulative_deviations,
            (0.04, -0.04, -0.01, 0.01),
        )
        self.assertEqual(
            result.classifications,
            (
                PREVIOUS_WINNER,
                PREVIOUS_LOSER,
                PREVIOUS_LOSER,
                PREVIOUS_WINNER,
            ),
        )

    def test_equal_threshold_and_single_stock_cluster_are_neutral(self) -> None:
        snapshot = make_snapshot(np.eye(3))
        clustering = replace(
            cluster_sponge_sym(snapshot, 1),
            requested_cluster_count=2,
            cluster_labels=(0, 0, 1),
            cluster_sizes=(2, 1),
        )
        returns = pd.DataFrame(
            {
                "AAA": [0.02] * 5,
                "BBB": [0.00] * 5,
                "CCC": [0.03] * 5,
            },
            index=pd.date_range("2026-07-10", periods=5, freq="B"),
        )

        result = identify_stocks_to_trade(
            clustering,
            returns,
            StockSelectionConfig(deviation_threshold=0.05),
        )

        np.testing.assert_allclose(
            result.cumulative_deviations,
            (0.05, -0.05, 0.0),
        )
        self.assertEqual(result.classifications, (NEUTRAL, NEUTRAL, NEUTRAL))

    def test_rejects_invalid_config_and_misaligned_inputs(self) -> None:
        for kwargs, message in (
            ({"lookback_window": 1}, "lookback_window"),
            ({"lookback_window": True}, "lookback_window"),
            ({"deviation_threshold": -0.01}, "deviation_threshold"),
            ({"deviation_threshold": float("nan")}, "deviation_threshold"),
        ):
            with self.subTest(kwargs=kwargs):
                with self.assertRaisesRegex(ValueError, message):
                    StockSelectionConfig(**kwargs)

        snapshot = make_snapshot(np.eye(3))
        clustering = cluster_sponge_sym(snapshot, 1)
        returns = pd.DataFrame(
            np.zeros((5, 3)),
            index=pd.date_range("2026-07-10", periods=5, freq="B"),
            columns=("AAA", "BBB", "CCC"),
        )
        with self.assertRaisesRegex(ValueError, "exactly match"):
            identify_stocks_to_trade(
                clustering,
                returns.rename(columns={"AAA": "ZZZ"}),
            )
        non_finite = returns.copy()
        non_finite.iloc[0, 0] = np.nan
        with self.assertRaisesRegex(ValueError, "non-finite"):
            identify_stocks_to_trade(clustering, non_finite)
        with self.assertRaisesRegex(ValueError, "lookback"):
            identify_stocks_to_trade(
                clustering,
                returns,
                StockSelectionConfig(lookback_window=4),
            )


class StockSelectionIntegrationTests(unittest.TestCase):
    def test_date_orchestration_uses_raw_returns_and_creates_no_result_tables(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
                correlation_window=5,
            )
            build_preprocessing(preprocessing_config)

            result = identify_stocks_for_date(
                preprocessing_config,
                dates[-1].date(),
            )

            self.assertEqual(result.raw_return_matrix.shape, (5, result.stock_count))
            self.assertEqual(
                tuple(result.raw_return_matrix.columns),
                result.tickers,
            )
            self.assertEqual(result.window_end, dates[-2].date())
            self.assertEqual(
                result.quality.winner_count
                + result.quality.loser_count
                + result.quality.neutral_count,
                result.stock_count,
            )
            with duckdb.connect(str(database), read_only=True) as connection:
                result_tables = connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM information_schema.tables
                    WHERE table_schema IN (
                        'stock_selection',
                        'signals',
                        'clustering',
                        'cluster_count'
                    )
                    """
                ).fetchone()[0]
            self.assertEqual(result_tables, 0)

    def test_excel_is_auditable_and_refuses_silent_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database, dates = build_test_database(root)
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
            )
            build_preprocessing(preprocessing_config)
            result = identify_stocks_for_date(
                preprocessing_config,
                dates[-1].date(),
            )
            output = root / "stock_signals.xlsx"

            export_stock_selection_workbook(result, output)

            workbook = load_workbook(output, data_only=False, read_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary",
                    "Raw_Returns",
                    "Cluster_Mean_Returns",
                    "Daily_Deviations",
                    "Trade_Signals",
                ],
            )
            parameters = workbook["Summary"]
            parameter_rows = {
                parameters.cell(row, 1).value: row
                for row in range(2, parameters.max_row + 1)
            }
            self.assertEqual(
                parameters.cell(
                    parameter_rows["Deviation threshold p"],
                    2,
                ).value,
                0.0,
            )
            self.assertEqual(
                parameters.cell(parameter_rows["Return input"], 2).value,
                "raw stock price returns",
            )
            self.assertEqual(
                parameters.cell(parameter_rows["Overall QC"], 2).value,
                "OK",
            )
            self.assertEqual(
                parameters.cell(
                    parameter_rows["All input returns finite"],
                    2,
                ).value,
                "YES",
            )
            signals = workbook["Trade_Signals"]
            self.assertEqual(signals.max_row, result.stock_count + 1)
            self.assertEqual(signals.max_column, 6)
            self.assertAlmostEqual(
                signals["E2"].value,
                result.cumulative_deviations[0],
            )
            self.assertEqual(
                signals["F2"].value,
                result.classifications[0],
            )
            self.assertEqual(
                workbook["Cluster_Mean_Returns"].max_column,
                3,
            )
            self.assertEqual(workbook["Daily_Deviations"].max_column, 6)
            self.assertEqual(
                workbook["Raw_Returns"].max_row,
                result.config.lookback_window * result.stock_count + 1,
            )
            workbook.close()

            with self.assertRaises(FileExistsError):
                export_stock_selection_workbook(result, output)


if __name__ == "__main__":
    unittest.main()
