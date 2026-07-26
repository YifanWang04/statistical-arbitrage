from __future__ import annotations

from dataclasses import replace
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

import duckdb
import numpy as np
from openpyxl import load_workbook
import pandas as pd

from stat_arb_clustering import (
    CALCULATION_VERSION,
    SpongeSymConfig,
    cluster_sponge_sym,
    cluster_stocks_for_date,
)
from stat_arb_clustering.calculations import _signet_symmetric_laplacian
from stat_arb_clustering.excel import export_clustering_workbook
from stat_arb_preprocessing import PreprocessingConfig, build_preprocessing
from tests.test_cluster_count import make_snapshot
from tests.test_preprocessing import build_test_database


class SpongeSymCalculationTests(unittest.TestCase):
    def test_signed_block_matrix_produces_reproducible_two_clusters(self) -> None:
        correlation = signed_block_correlation()
        snapshot = make_snapshot(correlation)

        first = cluster_sponge_sym(snapshot, 2, SpongeSymConfig())
        second = cluster_sponge_sym(snapshot, 2, SpongeSymConfig())

        np.testing.assert_allclose(first.positive_degrees, np.full(6, 1.6))
        np.testing.assert_allclose(first.negative_degrees, np.full(6, 1.2))
        self.assertEqual(first.embedding.shape, (6, 1))
        self.assertEqual(first.generalized_eigenvalues, second.generalized_eigenvalues)
        self.assertEqual(first.cluster_labels, second.cluster_labels)
        self.assertEqual(sorted(first.cluster_sizes), [3, 3])
        self.assertEqual(first.quality.nonempty_cluster_count, 2)
        self.assertEqual(first.calculation_version, CALCULATION_VERSION)
        self.assertLess(
            first.quality.maximum_generalized_eigen_residual,
            1e-10,
        )

    def test_embedding_is_eigenvector_scaled_by_inverse_eigenvalue(self) -> None:
        snapshot = make_snapshot(signed_block_correlation())
        result = cluster_sponge_sym(snapshot, 2)
        adjacency = snapshot.correlation_matrix.to_numpy(copy=True)
        np.fill_diagonal(adjacency, 0.0)
        positive = np.maximum(adjacency, 0.0)
        negative = np.maximum(-adjacency, 0.0)
        positive_laplacian = _signet_symmetric_laplacian(
            positive,
            positive.sum(axis=1),
        )
        negative_laplacian = _signet_symmetric_laplacian(
            negative,
            negative.sum(axis=1),
        )
        numerator = positive_laplacian + np.eye(6)
        denominator = negative_laplacian + np.eye(6)
        eigenvalue = result.generalized_eigenvalues[0]
        scaled = result.embedding.iloc[:, 0].to_numpy()
        raw_eigenvector = scaled * eigenvalue

        residual = (
            numerator @ raw_eigenvector
            - eigenvalue * (denominator @ raw_eigenvector)
        )

        self.assertAlmostEqual(
            result.inverse_eigenvalue_weights[0],
            1.0 / eigenvalue,
        )
        self.assertLess(float(np.linalg.norm(residual)), 1e-10)

    def test_signet_zero_degree_convention_returns_identity_laplacian(self) -> None:
        adjacency = np.zeros((3, 3), dtype=float)

        laplacian = _signet_symmetric_laplacian(
            adjacency,
            np.zeros(3, dtype=float),
        )

        np.testing.assert_array_equal(laplacian, np.eye(3))

    def test_k_one_assigns_every_stock_without_eigenvectors(self) -> None:
        result = cluster_sponge_sym(make_snapshot(np.eye(4)), 1)

        self.assertEqual(result.cluster_labels, (0, 0, 0, 0))
        self.assertEqual(result.cluster_sizes, (4,))
        self.assertEqual(result.embedding.shape, (4, 0))
        self.assertEqual(result.generalized_eigenvalues, ())
        self.assertEqual(result.quality.kmeans_iterations, 0)

    def test_k_equal_to_stock_count_uses_dense_edge_case(self) -> None:
        result = cluster_sponge_sym(make_snapshot(np.eye(4)), 4)

        self.assertEqual(result.embedding.shape, (4, 3))
        self.assertEqual(result.quality.nonempty_cluster_count, 4)
        self.assertEqual(result.cluster_sizes, (1, 1, 1, 1))

    def test_symmetrizes_input_and_clears_diagonal_before_decomposition(self) -> None:
        snapshot = make_snapshot(
            np.array(
                [
                    [1.0, 0.6, -0.2, -0.1],
                    [0.4, 1.0, -0.1, -0.2],
                    [-0.2, -0.1, 1.0, 0.5],
                    [-0.1, -0.2, 0.5, 1.0],
                ]
            )
        )

        result = cluster_sponge_sym(snapshot, 2)

        self.assertAlmostEqual(result.quality.maximum_input_asymmetry, 0.2)
        self.assertEqual(result.quality.maximum_reconstruction_error, 0.0)
        np.testing.assert_allclose(result.positive_degrees, (0.5, 0.5, 0.5, 0.5))
        np.testing.assert_allclose(result.negative_degrees, (0.3, 0.3, 0.3, 0.3))

    def test_rejects_invalid_configuration_and_inputs(self) -> None:
        for kwargs, message in (
            ({"tau_positive": 0.0}, "tau_positive"),
            ({"tau_negative": float("nan")}, "tau_negative"),
            ({"random_seed": -1}, "random_seed"),
            ({"random_seed": 2**32}, "random_seed"),
            ({"kmeans_n_init": 0}, "kmeans_n_init"),
            ({"kmeans_max_iter": 0}, "kmeans_max_iter"),
        ):
            with self.subTest(kwargs=kwargs):
                with self.assertRaisesRegex(ValueError, message):
                    SpongeSymConfig(**kwargs)

        snapshot = make_snapshot(np.eye(3))
        for invalid_k in (0, 4, 1.5, True):
            with self.subTest(k=invalid_k):
                with self.assertRaisesRegex(ValueError, "k must"):
                    cluster_sponge_sym(snapshot, invalid_k)  # type: ignore[arg-type]

        duplicate = replace(snapshot, tickers=("AAA", "AAA", "CCC"))
        with self.assertRaisesRegex(ValueError, "unique"):
            cluster_sponge_sym(duplicate, 2)

        out_of_range = snapshot.correlation_matrix.copy()
        out_of_range.iloc[0, 1] = 1.1
        out_of_range.iloc[1, 0] = 1.1
        with self.assertRaisesRegex(ValueError, r"\[-1, 1\]"):
            cluster_sponge_sym(
                replace(snapshot, correlation_matrix=out_of_range),
                2,
            )

        non_finite = snapshot.correlation_matrix.copy()
        non_finite.iloc[0, 1] = np.nan
        non_finite.iloc[1, 0] = np.nan
        with self.assertRaisesRegex(ValueError, "non-finite"):
            cluster_sponge_sym(
                replace(snapshot, correlation_matrix=non_finite),
                2,
            )

    def test_surfaces_generalized_eigensolver_failure(self) -> None:
        with patch(
            "stat_arb_clustering.calculations._smallest_generalized_eigenpairs",
            side_effect=RuntimeError("LOBPCG generalized eigenproblem did not converge"),
        ):
            with self.assertRaisesRegex(RuntimeError, "did not converge"):
                cluster_sponge_sym(make_snapshot(signed_block_correlation()), 2)


class SpongeSymIntegrationTests(unittest.TestCase):
    def test_date_orchestration_uses_independent_windows_without_new_result_tables(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
                correlation_window=5,
            )
            build_preprocessing(preprocessing_config)
            as_of_date = dates[-1].date()

            with duckdb.connect(str(database), read_only=True) as connection:
                before = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.correlation_snapshots"
                ).fetchone()[0]

            result = cluster_stocks_for_date(
                preprocessing_config,
                as_of_date,
                cluster_count_estimation_window=20,
                variance_threshold=0.90,
            )

            with duckdb.connect(str(database), read_only=True) as connection:
                after = connection.execute(
                    "SELECT COUNT(*) FROM preprocessing.correlation_snapshots"
                ).fetchone()[0]
                result_tables = connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM information_schema.tables
                    WHERE table_schema IN ('clustering', 'cluster_count')
                       OR table_name LIKE 'clustering%'
                       OR table_name LIKE 'cluster_count%'
                    """
                ).fetchone()[0]

            self.assertEqual(before, after)
            self.assertEqual(result_tables, 0)
            self.assertEqual(result.clustering_correlation_window, 5)
            self.assertIsNotNone(result.cluster_count_result)
            self.assertEqual(
                result.cluster_count_result.cluster_count_estimation_window,
                20,
            )
            self.assertEqual(
                result.embedding_dimension,
                max(result.requested_cluster_count - 1, 0),
            )
            self.assertEqual(sum(result.cluster_sizes), result.stock_count)

    def test_excel_has_four_auditable_sheets_and_refuses_silent_overwrite(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database, dates = build_test_database(root)
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
            )
            build_preprocessing(preprocessing_config)
            result = cluster_stocks_for_date(
                preprocessing_config,
                dates[-1].date(),
            )
            output = root / "clusters.xlsx"

            export_clustering_workbook(result, output)

            workbook = load_workbook(output, data_only=False, read_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Parameters_QC",
                    "Eigenvalues",
                    "Spectral_Embedding",
                    "Cluster_Assignments",
                ],
            )
            parameters = workbook["Parameters_QC"]
            parameter_rows = {
                parameters.cell(row, 1).value: row
                for row in range(2, parameters.max_row + 1)
                if parameters.cell(row, 1).value is not None
            }
            self.assertEqual(
                parameters.cell(parameter_rows["Selected K"], 2).value,
                result.requested_cluster_count,
            )
            self.assertEqual(
                parameters.cell(parameter_rows["Random seed"], 2).value,
                0,
            )
            self.assertEqual(
                parameters.cell(parameter_rows["KMeans n_init"], 2).value,
                10,
            )
            self.assertEqual(
                parameters.cell(parameter_rows["Clustering version"], 2).value,
                CALCULATION_VERSION,
            )
            self.assertTrue(
                str(
                    parameters.cell(
                        parameter_rows["Assignments complete"],
                        3,
                    ).value
                ).startswith("=")
            )
            self.assertEqual(
                parameters.cell(parameter_rows["Nonempty clusters"], 3).value,
                (
                    f'=COUNTIF(C44:C{43 + result.requested_cluster_count},'
                    '">0")'
                ),
            )
            self.assertEqual(
                workbook["Eigenvalues"].max_row,
                result.embedding_dimension + 1,
            )
            self.assertEqual(
                workbook["Spectral_Embedding"].max_column,
                result.embedding_dimension + 1,
            )
            self.assertEqual(
                workbook["Cluster_Assignments"].max_row,
                result.stock_count + 1,
            )
            workbook.close()

            with self.assertRaises(FileExistsError):
                export_clustering_workbook(result, output)


def signed_block_correlation() -> np.ndarray:
    matrix = np.full((6, 6), -0.4, dtype=float)
    matrix[:3, :3] = 0.8
    matrix[3:, 3:] = 0.8
    np.fill_diagonal(matrix, 1.0)
    return matrix


if __name__ == "__main__":
    unittest.main()
