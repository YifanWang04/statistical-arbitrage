from __future__ import annotations

from contextlib import redirect_stderr
from dataclasses import replace
from datetime import date
import io
from pathlib import Path
from types import SimpleNamespace
import math
import tempfile
import unittest
from unittest.mock import Mock, patch

import duckdb
from openpyxl import load_workbook
import pandas as pd

from stat_arb_backtest import (
    BacktestConfig,
    BacktestMarketData,
    BacktestTarget,
    TargetWeight,
    simulate_backtest,
)
from stat_arb_clustering import SpongeSymConfig
from stat_arb_grid_backtest import (
    GridBacktestConfig,
    GridBacktestResult,
    GridRunResult,
    GridRunSpec,
    build_grid_run_specs,
    calculate_grid_run_metrics,
    export_grid_backtest_workbook,
    resolve_grid_date_range,
    run_grid_backtest,
)
from stat_arb_grid_backtest.application import (
    _ClusterCountCache,
    _TargetCache,
    _rank_runs,
)
from stat_arb_grid_backtest.cli import build_parser
from stat_arb_preprocessing import PreprocessingConfig


class GridConfigurationTests(unittest.TestCase):
    def test_default_grid_contains_72_stable_combinations(self) -> None:
        config = GridBacktestConfig(
            start_date=date(2025, 1, 1),
            end_date=date(2026, 7, 27),
        )
        specs = build_grid_run_specs(config)

        self.assertEqual(config.combination_count, 72)
        self.assertEqual(len(specs), 72)
        self.assertEqual(specs[0].run_id, "G0001")
        self.assertEqual(specs[-1].run_id, "G0072")
        self.assertEqual(
            specs[0],
            GridRunSpec("G0001", 5, 0.0, 0.85, 3, 0.03),
        )

    def test_grid_values_are_sorted_deduplicated_and_guarded(self) -> None:
        config = GridBacktestConfig(
            start_date=date(2025, 1, 1),
            end_date=date(2026, 7, 27),
            lookback_windows=(20, 5, 5),
            deviation_thresholds=(0.05, 0.0, 0.05),
            variance_thresholds=(0.90, 0.85),
            rebalance_periods=(10, 3),
            take_profit_thresholds=(0.05, 0.03),
        )
        self.assertEqual(config.lookback_windows, (5, 20))
        self.assertEqual(config.deviation_thresholds, (0.0, 0.05))
        self.assertEqual(config.combination_count, 32)
        with self.assertRaisesRegex(ValueError, "exceeding"):
            replace(config, maximum_combinations=31)

    def test_grid_config_requires_explicit_dates(self) -> None:
        with self.assertRaisesRegex(TypeError, "start_date"):
            GridBacktestConfig()

    def test_cli_requires_explicit_dates(self) -> None:
        with (
            redirect_stderr(io.StringIO()),
            self.assertRaises(SystemExit),
        ):
            build_parser().parse_args(["export"])

        args = build_parser().parse_args(
            [
                "export",
                "--start-date",
                "2025-01-01",
                "--end-date",
                "2026-07-27",
            ]
        )
        self.assertEqual(args.start_date, date(2025, 1, 1))
        self.assertEqual(args.end_date, date(2026, 7, 27))

    def test_explicit_start_and_end_dates_are_preserved(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = Path(directory) / "market.duckdb"
            with duckdb.connect(str(database)) as connection:
                connection.execute("CREATE SCHEMA market_data")
                connection.execute(
                    """
                    CREATE TABLE market_data.market_returns (
                        trade_date DATE,
                        ticker VARCHAR
                    )
                    """
                )
                connection.execute(
                    """
                    INSERT INTO market_data.market_returns
                    VALUES (DATE '2026-07-27', 'SPY')
                    """
                )

            observed = resolve_grid_date_range(
                database,
                GridBacktestConfig(
                    start_date=date(2025, 1, 1),
                    end_date=date(2026, 7, 27),
                ),
            )

        self.assertEqual(
            observed,
            (date(2025, 1, 1), date(2026, 7, 27)),
        )

class GridMetricsTests(unittest.TestCase):
    def test_complete_metrics_preserve_step7_metrics_and_boundaries(
        self,
    ) -> None:
        result = _sample_backtest_result()

        metrics = calculate_grid_run_metrics(result)
        returns = pd.Series((0.02, -0.01, 0.03, -0.02))

        self.assertEqual(
            metrics.annualized_return,
            result.strategy_metrics.annualized_return,
        )
        self.assertEqual(
            metrics.sharpe_ratio,
            result.strategy_metrics.sharpe_ratio,
        )
        self.assertEqual(
            metrics.sortino_ratio,
            result.strategy_metrics.sortino_ratio,
        )
        self.assertAlmostEqual(
            metrics.annualized_volatility,
            returns.std(ddof=1) * math.sqrt(252),
        )
        self.assertAlmostEqual(metrics.maximum_drawdown, -0.02)
        self.assertEqual(metrics.drawdown_peak_date, date(2026, 1, 7))
        self.assertEqual(metrics.drawdown_trough_date, date(2026, 1, 8))
        self.assertIsNone(metrics.drawdown_recovery_date)
        self.assertAlmostEqual(
            metrics.daily_var_95,
            returns.quantile(0.05, interpolation="linear"),
        )
        self.assertAlmostEqual(metrics.daily_cvar_95, -0.02)
        self.assertEqual(metrics.positive_session_count, 2)
        self.assertEqual(metrics.negative_session_count, 2)
        self.assertEqual(metrics.win_rate, 0.5)
        self.assertEqual(metrics.initial_event_count, 1)
        self.assertEqual(metrics.scheduled_event_count, 0)
        self.assertEqual(metrics.annualized_two_way_turnover, 0.0)
        self.assertEqual(metrics.fifo_reconciliation_status, "OK")
        self.assertEqual(metrics.overall_qc, "OK")

    def test_ranking_uses_sharpe_then_return_drawdown_and_run_id(
        self,
    ) -> None:
        base = calculate_grid_run_metrics(_sample_backtest_result())
        specs = (
            GridRunSpec("G0001", 5, 0.0, 0.85, 3, 0.03),
            GridRunSpec("G0002", 5, 0.0, 0.85, 3, 0.05),
            GridRunSpec("G0003", 5, 0.0, 0.85, 5, 0.03),
            GridRunSpec("G0004", 5, 0.0, 0.85, 5, 0.05),
        )
        runs = (
            GridRunResult(
                specs[0],
                "SUCCESS",
                replace(
                    base,
                    sharpe_ratio=1.0,
                    annualized_return=0.20,
                    maximum_drawdown=-0.10,
                ),
            ),
            GridRunResult(
                specs[1],
                "SUCCESS",
                replace(
                    base,
                    sharpe_ratio=1.0,
                    annualized_return=0.25,
                    maximum_drawdown=-0.20,
                ),
            ),
            GridRunResult(
                specs[2],
                "SUCCESS",
                replace(
                    base,
                    sharpe_ratio=1.0,
                    annualized_return=0.25,
                    maximum_drawdown=-0.08,
                ),
            ),
            GridRunResult(
                specs[3],
                "SUCCESS",
                replace(base, sharpe_ratio=None),
            ),
        )

        ranked = _rank_runs(runs)

        self.assertEqual(
            {run.spec.run_id: run.rank for run in ranked},
            {"G0001": 3, "G0002": 2, "G0003": 1, "G0004": None},
        )


class GridCacheAndIntegrationTests(unittest.TestCase):
    def test_parallel_lookbacks_preserve_grid_run_order(self) -> None:
        market_data, target = _sample_market_data_and_target()
        grid_config = GridBacktestConfig(
            start_date=market_data.sessions[0],
            end_date=market_data.sessions[-1],
            lookback_windows=(20, 5),
            deviation_thresholds=(0.0,),
            variance_thresholds=(0.85,),
            rebalance_periods=(10,),
            take_profit_thresholds=(1.0,),
        )

        with (
            patch(
                "stat_arb_grid_backtest.application.resolve_grid_date_range",
                return_value=(
                    market_data.sessions[0],
                    market_data.sessions[-1],
                ),
            ),
            patch(
                "stat_arb_grid_backtest.application.BacktestMarketDataRepository",
                return_value=SimpleNamespace(
                    load=lambda _: market_data,
                ),
            ),
            patch(
                "stat_arb_grid_backtest.application._TargetCache.get",
                side_effect=lambda as_of_date, *_: replace(
                    target,
                    as_of_date=as_of_date,
                ),
            ),
        ):
            observed = run_grid_backtest(
                PreprocessingConfig(Path("unused.duckdb")),
                grid_config,
            )

        self.assertEqual(
            [
                (run.spec.run_id, run.spec.lookback_window)
                for run in observed.runs
            ],
            [("G0001", 5), ("G0002", 20)],
        )

    def test_target_cache_reuses_cluster_count_snapshot_for_same_window(
        self,
    ) -> None:
        preprocessing = PreprocessingConfig(
            database_path=Path("unused.duckdb"),
        )
        snapshot = SimpleNamespace(stock_return_matrix="raw")
        cluster_count = SimpleNamespace(variance_threshold=0.90)

        with (
            patch(
                "stat_arb_grid_backtest.application.get_snapshot",
                return_value=snapshot,
            ) as get_snapshot,
            patch(
                "stat_arb_grid_backtest.application.calculate_cluster_counts",
                return_value=(cluster_count,),
            ),
            patch(
                "stat_arb_grid_backtest.application.cluster_stocks_from_snapshot",
                return_value="cluster",
            ),
            patch(
                "stat_arb_grid_backtest.application.identify_stocks_to_trade",
                return_value="selection",
            ),
            patch(
                "stat_arb_grid_backtest.application.assign_portfolio_weights",
                return_value="weights",
            ),
            patch(
                "stat_arb_grid_backtest.application.target_from_portfolio_weights",
                return_value="target",
            ),
        ):
            cluster_counts = _ClusterCountCache(
                preprocessing,
                20,
                (0.90,),
            )
            targets = _TargetCache(
                preprocessing,
                20,
                (0.05,),
                (0.90,),
                cluster_counts,
                SpongeSymConfig(),
            )

            observed = targets.get(
                date(2026, 1, 5),
                0.90,
                0.05,
            )

        self.assertEqual(observed, "target")
        self.assertEqual(get_snapshot.call_count, 1)

    def test_target_cache_reuses_snapshot_and_clustering_across_p(
        self,
    ) -> None:
        preprocessing = PreprocessingConfig(
            database_path=Path("unused.duckdb"),
            correlation_window=5,
        )
        cluster_counts = Mock()
        cluster_counts.get.return_value = {0.85: "k85", 0.90: "k90"}
        snapshot = SimpleNamespace(stock_return_matrix="raw")
        clustering = Mock(side_effect=["c85", "c90"])

        with (
            patch(
                "stat_arb_grid_backtest.application.get_snapshot",
                return_value=snapshot,
            ) as get_snapshot,
            patch(
                "stat_arb_grid_backtest.application.cluster_stocks_from_snapshot",
                clustering,
            ),
            patch(
                "stat_arb_grid_backtest.application.identify_stocks_to_trade",
                side_effect=lambda cluster, raw, config: (
                    cluster,
                    config.deviation_threshold,
                ),
            ) as identify,
            patch(
                "stat_arb_grid_backtest.application.assign_portfolio_weights",
                side_effect=lambda value: value,
            ),
            patch(
                "stat_arb_grid_backtest.application.target_from_portfolio_weights",
                side_effect=lambda value: value,
            ),
        ):
            cache = _TargetCache(
                preprocessing,
                5,
                (0.0, 0.05),
                (0.85, 0.90),
                cluster_counts,
                SpongeSymConfig(),
            )
            for variance in (0.85, 0.90):
                for deviation in (0.0, 0.05):
                    cache.get(date(2026, 1, 5), variance, deviation)
            cache.get(date(2026, 1, 5), 0.85, 0.0)

        self.assertEqual(get_snapshot.call_count, 1)
        self.assertEqual(cluster_counts.get.call_count, 1)
        self.assertEqual(clustering.call_count, 2)
        self.assertEqual(identify.call_count, 4)

    def test_grid_one_run_matches_direct_step7_state_engine(self) -> None:
        market_data, target = _sample_market_data_and_target()
        grid_config = GridBacktestConfig(
            start_date=market_data.sessions[0],
            end_date=market_data.sessions[-1],
            lookback_windows=(5,),
            deviation_thresholds=(0.0,),
            variance_thresholds=(0.85,),
            rebalance_periods=(10,),
            take_profit_thresholds=(1.0,),
        )
        direct = simulate_backtest(
            market_data,
            BacktestConfig(
                market_data.sessions[0],
                market_data.sessions[-1],
                rebalance_period=10,
                take_profit_threshold=1.0,
            ),
            lambda _: target,
        )
        with (
            patch(
                "stat_arb_grid_backtest.application.resolve_grid_date_range",
                return_value=(
                    market_data.sessions[0],
                    market_data.sessions[-1],
                ),
            ),
            patch(
                "stat_arb_grid_backtest.application.BacktestMarketDataRepository",
                return_value=SimpleNamespace(
                    load=lambda _: market_data,
                ),
            ),
            patch(
                "stat_arb_grid_backtest.application._TargetCache.get",
                return_value=target,
            ),
        ):
            observed = run_grid_backtest(
                PreprocessingConfig(Path("unused.duckdb")),
                grid_config,
            )

        metrics = observed.runs[0].metrics
        self.assertEqual(observed.runs[0].status, "SUCCESS")
        self.assertIsNotNone(metrics)
        self.assertEqual(
            metrics.total_return,
            direct.strategy_metrics.total_return,
        )
        self.assertEqual(
            metrics.sharpe_ratio,
            direct.strategy_metrics.sharpe_ratio,
        )

    def test_one_failed_combination_does_not_stop_the_grid(self) -> None:
        market_data, target = _sample_market_data_and_target()
        grid_config = GridBacktestConfig(
            start_date=market_data.sessions[0],
            end_date=market_data.sessions[-1],
            lookback_windows=(5,),
            deviation_thresholds=(0.0, 0.05),
            variance_thresholds=(0.85,),
            rebalance_periods=(10,),
            take_profit_thresholds=(1.0,),
        )

        def target_or_failure(
            as_of_date: date,
            variance_threshold: float,
            deviation_threshold: float,
        ) -> BacktestTarget:
            del variance_threshold
            if deviation_threshold == 0.05:
                raise ValueError("isolated target failure")
            return replace(target, as_of_date=as_of_date)

        with (
            patch(
                "stat_arb_grid_backtest.application.resolve_grid_date_range",
                return_value=(
                    market_data.sessions[0],
                    market_data.sessions[-1],
                ),
            ),
            patch(
                "stat_arb_grid_backtest.application.BacktestMarketDataRepository",
                return_value=SimpleNamespace(
                    load=lambda _: market_data,
                ),
            ),
            patch(
                "stat_arb_grid_backtest.application._TargetCache.get",
                side_effect=target_or_failure,
            ),
        ):
            observed = run_grid_backtest(
                PreprocessingConfig(Path("unused.duckdb")),
                grid_config,
            )

        self.assertEqual(
            [run.status for run in observed.runs],
            ["SUCCESS", "FAILED"],
        )
        self.assertEqual(observed.runs[1].error_type, "ValueError")
        self.assertEqual(observed.best_run_id, "G0001")
        self.assertEqual(observed.overall_qc, "CHECK")


class GridExcelTests(unittest.TestCase):
    def test_excel_has_five_audit_sheets_and_refuses_overwrite(
        self,
    ) -> None:
        metrics = calculate_grid_run_metrics(_sample_backtest_result())
        runs = (
            GridRunResult(
                GridRunSpec("G0001", 5, 0.0, 0.85, 3, 0.03),
                "SUCCESS",
                metrics,
                rank=1,
            ),
            GridRunResult(
                GridRunSpec("G0002", 5, 0.0, 0.85, 3, 0.05),
                "FAILED",
                error_type="ValueError",
                error_message="sample failure",
            ),
        )
        config = GridBacktestConfig(
            start_date=date(2026, 1, 5),
            end_date=date(2026, 1, 8),
            lookback_windows=(5,),
            deviation_thresholds=(0.0,),
            variance_thresholds=(0.85,),
            rebalance_periods=(3,),
            take_profit_thresholds=(0.03, 0.05),
        )
        result = GridBacktestResult(
            config=config,
            requested_start_date=date(2026, 1, 5),
            requested_end_date=date(2026, 1, 8),
            effective_start_date=date(2026, 1, 5),
            effective_end_date=date(2026, 1, 8),
            beta_window=60,
            cluster_count_estimation_window=20,
            sponge_config=SpongeSymConfig(),
            runs=runs,
            best_run_id="G0001",
            overall_qc="CHECK",
        )

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "grid.xlsx"
            export_grid_backtest_workbook(
                result,
                output,
            )
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary",
                    "Grid_Results",
                    "Parameter_Grid",
                    "Metric_Definitions",
                    "Checks",
                ],
            )
            self.assertEqual(workbook["Grid_Results"].max_row, 3)
            self.assertEqual(workbook["Grid_Results"]["B2"].value, "G0001")
            self.assertEqual(workbook["Grid_Results"]["A2"].value, 1)
            self.assertEqual(workbook["Grid_Results"]["C3"].value, "FAILED")
            self.assertEqual(workbook["Summary"]["B8"].value, "G0001")
            self.assertEqual(workbook["Summary"]["D8"].value, "CHECK")
            self.assertEqual(workbook["Grid_Results"].freeze_panes, "K2")
            self.assertGreater(
                workbook["Metric_Definitions"].max_row,
                50,
            )
            workbook.close()
            with self.assertRaises(FileExistsError):
                export_grid_backtest_workbook(
                    result,
                    output,
                )


def _sample_market_data_and_target() -> tuple[
    BacktestMarketData,
    BacktestTarget,
]:
    previous = date(2026, 1, 2)
    sessions = (
        date(2026, 1, 5),
        date(2026, 1, 6),
        date(2026, 1, 7),
        date(2026, 1, 8),
    )
    returns = (0.02, -0.01, 0.03, -0.02)
    spy_returns = (0.01, -0.005, 0.01, 0.0)
    aaa = [100.0]
    spy = [100.0]
    for value in returns:
        aaa.append(aaa[-1] * (1.0 + value))
    for value in spy_returns:
        spy.append(spy[-1] * (1.0 + value))
    closes = pd.DataFrame(
        {"AAA": aaa, "SPY": spy},
        index=pd.Index((previous, *sessions), name="trade_date"),
    )
    target = BacktestTarget(
        sessions[0],
        cluster_count=1,
        active_cluster_count=1,
        weights=(TargetWeight("AAA", 1.0),),
    )
    return BacktestMarketData(previous, sessions, closes), target


def _sample_backtest_result():
    market_data, target = _sample_market_data_and_target()
    return simulate_backtest(
        market_data,
        BacktestConfig(
            market_data.sessions[0],
            market_data.sessions[-1],
            rebalance_period=10,
            take_profit_threshold=1.0,
        ),
        lambda as_of_date: replace(target, as_of_date=as_of_date),
    )


if __name__ == "__main__":
    unittest.main()
