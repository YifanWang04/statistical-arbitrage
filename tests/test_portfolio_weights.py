from __future__ import annotations

from dataclasses import replace
from pathlib import Path
import tempfile
import unittest

import duckdb
import numpy as np
from openpyxl import load_workbook
import pandas as pd

from stat_arb_backtest import target_from_portfolio_weights
from stat_arb_clustering import cluster_sponge_sym
from stat_arb_portfolio_weights import (
    assign_portfolio_weights,
    assign_weights_for_date,
)
from stat_arb_portfolio_weights.excel import (
    export_portfolio_weight_workbook,
)
from stat_arb_preprocessing import PreprocessingConfig, build_preprocessing
from stat_arb_stock_selection import (
    NEUTRAL,
    PREVIOUS_LOSER,
    PREVIOUS_WINNER,
    identify_stocks_to_trade,
)
from tests.test_cluster_count import make_snapshot
from tests.test_preprocessing import build_test_database


class PortfolioWeightCalculationTests(unittest.TestCase):
    def test_equal_long_only_weights_and_equal_cluster_gross(self) -> None:
        selection = make_selection(
            labels=(0, 0, 0, 0, 1, 1, 1, 1),
            classifications=(
                PREVIOUS_WINNER,
                PREVIOUS_LOSER,
                PREVIOUS_LOSER,
                NEUTRAL,
                PREVIOUS_WINNER,
                PREVIOUS_WINNER,
                PREVIOUS_LOSER,
                NEUTRAL,
            ),
        )

        result = assign_portfolio_weights(selection)

        np.testing.assert_allclose(
            result.local_weights,
            (0.0, 0.5, 0.5, 0.0, 0.0, 0.0, 1.0, 0.0),
        )
        np.testing.assert_allclose(
            result.portfolio_weights,
            (0.0, 0.25, 0.25, 0.0, 0.0, 0.0, 0.5, 0.0),
        )
        self.assertEqual(result.quality.active_cluster_count, 2)
        self.assertEqual(result.quality.inactive_cluster_count, 0)
        self.assertAlmostEqual(result.quality.long_exposure, 1.0)
        self.assertAlmostEqual(result.quality.short_exposure, 0.0)
        self.assertAlmostEqual(result.quality.net_exposure, 1.0)
        self.assertAlmostEqual(result.quality.gross_exposure, 1.0)
        self.assertAlmostEqual(result.quality.uninvested_gross_exposure, 0.0)
        for allocation in result.cluster_allocations:
            self.assertTrue(allocation.is_active)
            self.assertAlmostEqual(allocation.local_long_exposure, 1.0)
            self.assertAlmostEqual(allocation.local_short_exposure, 0.0)
            self.assertAlmostEqual(allocation.local_net_exposure, 1.0)
            self.assertAlmostEqual(allocation.local_gross_exposure, 1.0)
            self.assertAlmostEqual(allocation.portfolio_gross_exposure, 0.5)

    def test_inactive_cluster_capital_is_redistributed_to_active_clusters(
        self,
    ) -> None:
        selection = make_selection(
            labels=(0, 0, 1),
            classifications=(PREVIOUS_WINNER, PREVIOUS_LOSER, NEUTRAL),
        )

        result = assign_portfolio_weights(selection)

        np.testing.assert_allclose(result.local_weights, (0.0, 1.0, 0.0))
        np.testing.assert_allclose(
            result.portfolio_weights,
            (0.0, 1.0, 0.0),
        )
        self.assertTrue(result.cluster_allocations[0].is_active)
        self.assertFalse(result.cluster_allocations[1].is_active)
        self.assertAlmostEqual(result.quality.gross_exposure, 1.0)
        self.assertAlmostEqual(result.quality.uninvested_gross_exposure, 0.0)
        self.assertAlmostEqual(
            result.cluster_allocations[0].target_gross_exposure,
            1.0,
        )
        self.assertAlmostEqual(
            result.cluster_allocations[1].uninvested_gross_exposure,
            0.0,
        )

    def test_multiple_active_clusters_split_all_capital_equally(self) -> None:
        selection = make_selection(
            labels=(0, 0, 1, 1, 2, 2, 2),
            classifications=(
                PREVIOUS_LOSER,
                PREVIOUS_WINNER,
                NEUTRAL,
                PREVIOUS_WINNER,
                PREVIOUS_LOSER,
                PREVIOUS_LOSER,
                NEUTRAL,
            ),
        )

        result = assign_portfolio_weights(selection)

        np.testing.assert_allclose(
            result.portfolio_weights,
            (0.5, 0.0, 0.0, 0.0, 0.25, 0.25, 0.0),
        )
        self.assertEqual(result.quality.active_cluster_count, 2)
        self.assertEqual(result.quality.inactive_cluster_count, 1)
        self.assertAlmostEqual(result.quality.gross_exposure, 1.0)
        self.assertAlmostEqual(result.quality.uninvested_gross_exposure, 0.0)
        np.testing.assert_allclose(
            [
                allocation.target_gross_exposure
                for allocation in result.cluster_allocations
            ],
            (0.5, 0.0, 0.5),
        )

    def test_winner_only_cluster_is_inactive(self) -> None:
        selection = make_selection(
            labels=(0, 0, 0),
            classifications=(PREVIOUS_WINNER, NEUTRAL, NEUTRAL),
        )

        result = assign_portfolio_weights(selection)

        np.testing.assert_allclose(result.local_weights, (0.0, 0.0, 0.0))
        np.testing.assert_allclose(result.portfolio_weights, (0.0, 0.0, 0.0))
        self.assertEqual(result.quality.active_cluster_count, 0)
        self.assertEqual(result.quality.inactive_cluster_count, 1)
        self.assertAlmostEqual(result.quality.uninvested_gross_exposure, 1.0)
        self.assertAlmostEqual(
            result.cluster_allocations[0].uninvested_gross_exposure,
            1.0,
        )

    def test_loser_only_cluster_is_active(self) -> None:
        selection = make_selection(
            labels=(0, 0, 0),
            classifications=(PREVIOUS_LOSER, NEUTRAL, NEUTRAL),
        )

        result = assign_portfolio_weights(selection)

        np.testing.assert_allclose(result.local_weights, (1.0, 0.0, 0.0))
        np.testing.assert_allclose(result.portfolio_weights, (1.0, 0.0, 0.0))
        self.assertEqual(result.quality.active_cluster_count, 1)
        self.assertEqual(result.quality.inactive_cluster_count, 0)
        self.assertAlmostEqual(result.quality.long_exposure, 1.0)
        self.assertAlmostEqual(result.quality.short_exposure, 0.0)
        self.assertAlmostEqual(result.quality.net_exposure, 1.0)
        self.assertAlmostEqual(result.quality.gross_exposure, 1.0)
        self.assertAlmostEqual(result.quality.uninvested_gross_exposure, 0.0)

    def test_backtest_target_contains_only_positive_portfolio_weights(self) -> None:
        selection = make_selection(
            labels=(0, 0, 1, 1),
            classifications=(
                PREVIOUS_WINNER,
                PREVIOUS_LOSER,
                NEUTRAL,
                PREVIOUS_LOSER,
            ),
        )
        result = assign_portfolio_weights(selection)

        target = target_from_portfolio_weights(result)

        self.assertEqual(
            tuple(weight.ticker for weight in target.weights),
            (result.tickers[1], result.tickers[3]),
        )
        self.assertTrue(
            all(weight.portfolio_weight > 0.0 for weight in target.weights)
        )
        self.assertAlmostEqual(target.target_gross_exposure, 1.0)

    def test_rejects_malformed_stock_selection_results(self) -> None:
        selection = make_selection(
            labels=(0, 0, 1, 1),
            classifications=(
                PREVIOUS_WINNER,
                PREVIOUS_LOSER,
                PREVIOUS_WINNER,
                PREVIOUS_LOSER,
            ),
        )
        cases = (
            (
                replace(
                    selection,
                    classifications=(
                        "unsupported",
                        *selection.classifications[1:],
                    ),
                ),
                "unsupported stock classifications",
            ),
            (
                replace(
                    selection,
                    classifications=selection.classifications[:-1],
                ),
                "classifications do not match",
            ),
            (
                replace(
                    selection,
                    cumulative_deviations=(
                        float("nan"),
                        *selection.cumulative_deviations[1:],
                    ),
                ),
                "non-finite",
            ),
            (
                replace(
                    selection,
                    clustering_result=replace(
                        selection.clustering_result,
                        cluster_sizes=(3, 1),
                    ),
                ),
                "cluster sizes do not reconcile",
            ),
        )
        for malformed, message in cases:
            with self.subTest(message=message):
                with self.assertRaisesRegex(ValueError, message):
                    assign_portfolio_weights(malformed)


class PortfolioWeightIntegrationTests(unittest.TestCase):
    def test_date_orchestration_uses_previous_sessions_and_persists_nothing(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database, dates = build_test_database(Path(directory))
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
                correlation_window=5,
            )
            build_preprocessing(preprocessing_config)

            result = assign_weights_for_date(
                preprocessing_config,
                dates[-1].date(),
            )

            self.assertEqual(
                result.stock_selection_result.window_end,
                dates[-2].date(),
            )
            self.assertAlmostEqual(result.quality.short_exposure, 0.0)
            self.assertAlmostEqual(
                result.quality.net_exposure,
                result.quality.gross_exposure,
            )
            self.assertAlmostEqual(
                result.quality.gross_exposure
                + result.quality.uninvested_gross_exposure,
                1.0,
            )
            with duckdb.connect(str(database), read_only=True) as connection:
                result_tables = connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM information_schema.tables
                    WHERE table_schema IN (
                        'portfolio_weights',
                        'portfolio',
                        'weights',
                        'stock_selection',
                        'clustering',
                        'cluster_count'
                    )
                    """
                ).fetchone()[0]
            self.assertEqual(result_tables, 0)

    def test_excel_is_auditable_and_refuses_silent_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database, dates = build_test_database(root)
            preprocessing_config = PreprocessingConfig(
                database_path=database,
                beta_window=40,
                correlation_window=5,
            )
            build_preprocessing(preprocessing_config)
            result = assign_weights_for_date(
                preprocessing_config,
                dates[-1].date(),
            )
            output = root / "portfolio_weights.xlsx"

            export_portfolio_weight_workbook(result, output)

            workbook = load_workbook(output, data_only=False, read_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary",
                    "Cluster_Allocations",
                    "Stock_Weights",
                ],
            )
            summary = workbook["Summary"]
            summary_rows = {
                summary.cell(row, 1).value: row
                for row in range(2, summary.max_row + 1)
            }
            self.assertEqual(
                summary.cell(summary_rows["Selected K"], 2).value,
                result.cluster_count,
            )
            self.assertEqual(
                summary.cell(
                    summary_rows["Cluster allocation"],
                    2,
                ).value,
                "equal 1/A share across A active clusters",
            )
            self.assertEqual(
                summary.cell(
                    summary_rows["Position direction"],
                    2,
                ).value,
                "long only: previous losers; winners stay at zero",
            )
            self.assertEqual(
                summary.cell(summary_rows["Overall QC"], 2).value,
                "OK",
            )
            clusters = workbook["Cluster_Allocations"]
            self.assertEqual(clusters.max_row, result.cluster_count + 1)
            self.assertEqual(clusters.max_column, 11)
            self.assertEqual(
                clusters["F2"].value,
                (
                    "Active"
                    if result.cluster_allocations[0].is_active
                    else "Inactive"
                ),
            )
            self.assertAlmostEqual(
                clusters["J2"].value,
                result.cluster_allocations[0].portfolio_gross_exposure,
            )
            stocks = workbook["Stock_Weights"]
            self.assertEqual(stocks.max_row, result.stock_count + 1)
            self.assertEqual(stocks.max_column, 7)
            self.assertEqual(stocks["A2"].value, result.tickers[0])
            self.assertAlmostEqual(
                stocks["G2"].value,
                result.portfolio_weights[0],
            )
            workbook.close()

            with self.assertRaises(FileExistsError):
                export_portfolio_weight_workbook(result, output)


def make_selection(
    *,
    labels: tuple[int, ...],
    classifications: tuple[str, ...],
):
    size = len(labels)
    if len(classifications) != size:
        raise ValueError("test classifications must match labels")
    cluster_count = max(labels) + 1
    snapshot = make_snapshot(np.eye(size))
    clustering = replace(
        cluster_sponge_sym(snapshot, 1),
        requested_cluster_count=cluster_count,
        cluster_labels=labels,
        cluster_sizes=tuple(
            labels.count(cluster_id) for cluster_id in range(cluster_count)
        ),
    )
    returns = pd.DataFrame(
        np.zeros((5, size)),
        index=pd.date_range("2026-07-10", periods=5, freq="B"),
        columns=clustering.tickers,
    )
    selection = identify_stocks_to_trade(clustering, returns)
    deviations = tuple(
        0.1
        if classification == PREVIOUS_WINNER
        else -0.1
        if classification == PREVIOUS_LOSER
        else 0.0
        for classification in classifications
    )
    return replace(
        selection,
        cumulative_deviations=deviations,
        classifications=classifications,
    )


if __name__ == "__main__":
    unittest.main()
