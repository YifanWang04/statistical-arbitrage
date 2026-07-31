from __future__ import annotations

from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from stat_arb_excel import workbook_for_publication

from .models import ClusterCountResult


NAVY = "1F4E78"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"
GREEN = "548235"
RED = "C00000"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
CHECK_FILL = PatternFill("solid", fgColor="FCE4D6")
NEUTRAL_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(color=WHITE, bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
SCIENTIFIC_FORMAT = "0.000000E+00"
QC_TOLERANCE = 1e-10


def export_cluster_count_workbook(
    result: ClusterCountResult,
    output_path: Path,
    *,
    replace_existing: bool = False,
) -> Path:
    with workbook_for_publication(
        output_path,
        replace_existing=replace_existing,
    ) as (workbook, output):
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Summary")
        eigenvalues = workbook.create_sheet("Eigenvalues")

        _write_summary(summary, result)
        _write_eigenvalues(eigenvalues, result)
    return output


def _write_summary(sheet: Worksheet, result: ClusterCountResult) -> None:
    rows = [
        ("Run setup", None),
        ("As-of date", result.as_of_date),
        ("Window start", result.window_start),
        ("Window end", result.window_end),
        (
            "Cluster-count estimation window",
            result.cluster_count_estimation_window,
        ),
        ("Stock count N", result.stock_count),
        ("Variance threshold P", result.variance_threshold),
        ("Selected K", result.selected_k),
        (None, None),
        ("Quality", None),
        ("Trace of C", result.quality.trace),
        ("Total variance used", result.total_variance),
        ("Raw eigenvalue sum", result.quality.raw_eigenvalue_sum),
        ("Trace difference", result.quality.trace_difference),
        ("Minimum raw eigenvalue", result.quality.minimum_raw_eigenvalue),
        (
            "Adjusted negative eigenvalues",
            result.quality.adjusted_negative_eigenvalue_count,
        ),
        ("Numerical rank", result.quality.numerical_rank),
        ("Threshold crossing", _threshold_crossing_status(result)),
        ("Overall QC", _cluster_count_qc_status(result)),
        (None, None),
        ("Method", None),
        ("Return basis", result.return_basis),
        ("Beta window", result.beta_window),
        (None, None),
        ("Provenance", None),
        ("Preprocessing run id", result.preprocessing_run_id),
        ("Snapshot id", result.snapshot_id),
        ("Source calculation version", result.source_calculation_version),
        ("Cluster-count version", result.calculation_version),
    ]
    for row in rows:
        sheet.append(row)

    sheet.insert_rows(1)
    sheet.merge_cells("A1:B1")
    sheet["A1"] = f"Cluster Count — {result.as_of_date.isoformat()}"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    section_rows = (2, 11, 22, 26)
    for row in section_rows:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1).fill = SECTION_FILL
        sheet.cell(row, 1).font = Font(bold=True, color=NAVY)
        sheet.cell(row, 1).border = Border(bottom=THIN_GRAY)

    rows_by_label = {
        sheet.cell(row, 1).value: row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value is not None
    }
    for label in ("As-of date", "Window start", "Window end"):
        sheet.cell(rows_by_label[label], 2).number_format = "yyyy-mm-dd"
    for label in (
        "Cluster-count estimation window",
        "Stock count N",
        "Selected K",
        "Adjusted negative eigenvalues",
        "Numerical rank",
        "Beta window",
    ):
        sheet.cell(rows_by_label[label], 2).number_format = "0"
    sheet.cell(
        rows_by_label["Variance threshold P"],
        2,
    ).number_format = "0.00%"
    for label in ("Trace of C", "Total variance used", "Raw eigenvalue sum"):
        sheet.cell(rows_by_label[label], 2).number_format = "0.000000"
    for label in ("Trace difference", "Minimum raw eigenvalue"):
        sheet.cell(rows_by_label[label], 2).number_format = SCIENTIFIC_FORMAT

    for label in ("Threshold crossing", "Overall QC"):
        cell = sheet.cell(rows_by_label[label], 2)
        cell.font = Font(
            bold=True,
            color=GREEN if cell.value == "OK" else RED,
        )
        cell.fill = OK_FILL if cell.value == "OK" else CHECK_FILL
    for row in range(3, sheet.max_row + 1):
        if sheet.cell(row, 1).value is not None and row not in section_rows:
            sheet.cell(row, 1).font = Font(bold=True)

    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 58


def _write_eigenvalues(sheet: Worksheet, result: ClusterCountResult) -> None:
    sheet.append(
        [
            "Rank",
            "Raw Eigenvalue",
            "Eigenvalue Used",
            "Cumulative Explained Ratio",
            "Included in K",
        ]
    )
    for rank, values in enumerate(
        zip(
            result.raw_eigenvalues,
            result.effective_eigenvalues,
            result.cumulative_explained_ratio,
            strict=True,
        ),
        start=1,
    ):
        raw, effective, ratio = values
        sheet.append(
            [
                rank,
                raw,
                effective,
                ratio,
                "YES" if rank <= result.selected_k else "NO",
            ]
        )

    _style_header(sheet, columns=5)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    widths = (10, 22, 22, 26, 16)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "0"
        for column in (2, 3):
            sheet.cell(row, column).number_format = SCIENTIFIC_FORMAT
        sheet.cell(row, 4).number_format = "0.00%"
        if row % 2 == 0:
            for column in range(1, 6):
                sheet.cell(row, column).fill = PatternFill(
                    "solid",
                    fgColor="F8FAFC",
                )

    included_range = f"E2:E{sheet.max_row}"
    sheet.conditional_formatting.add(
        included_range,
        CellIsRule(operator="equal", formula=['"YES"'], fill=OK_FILL),
    )
    sheet.conditional_formatting.add(
        included_range,
        CellIsRule(operator="equal", formula=['"NO"'], fill=NEUTRAL_FILL),
    )


def _threshold_crossing_status(result: ClusterCountResult) -> str:
    selected_ratio = result.cumulative_explained_ratio[result.selected_k - 1]
    previous_ratio = (
        result.cumulative_explained_ratio[result.selected_k - 2]
        if result.selected_k > 1
        else None
    )
    crossed = selected_ratio + QC_TOLERANCE >= result.variance_threshold
    first_crossing = (
        previous_ratio is None
        or previous_ratio < result.variance_threshold
    )
    return "OK" if crossed and first_crossing else "CHECK"


def _cluster_count_qc_status(result: ClusterCountResult) -> str:
    quality = result.quality
    checks = (
        result.stock_count == len(result.raw_eigenvalues),
        result.stock_count == len(result.effective_eigenvalues),
        result.stock_count == len(result.cumulative_explained_ratio),
        result.total_variance > 0.0,
        abs(quality.trace_difference) < 1e-8,
        quality.minimum_raw_eigenvalue >= -QC_TOLERANCE,
        1 <= result.selected_k <= result.stock_count,
        _threshold_crossing_status(result) == "OK",
    )
    return "OK" if all(checks) else "CHECK"


def _style_header(sheet: Worksheet, *, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(1, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[1].height = 32
