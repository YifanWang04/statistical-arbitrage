from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .models import StockSelectionResult


NAVY = "1F4E78"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"
GREEN = "548235"
RED = "C00000"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
CHECK_FILL = PatternFill("solid", fgColor="FCE4D6")
NEUTRAL_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(color=WHITE, bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
RETURN_FORMAT = "0.0000%"
CUMULATIVE_RETURN_FORMAT = "0.00%"
SCIENTIFIC_FORMAT = "0.000000E+00"


def export_stock_selection_workbook(
    result: StockSelectionResult,
    output_path: Path,
    *,
    replace_existing: bool = False,
) -> Path:
    output = Path(output_path).resolve()
    if output.exists() and not replace_existing:
        raise FileExistsError(
            f"Excel output already exists: {output}. Use --replace to overwrite it."
        )
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Summary")
    raw_returns = workbook.create_sheet("Raw_Returns")
    cluster_means = workbook.create_sheet("Cluster_Mean_Returns")
    deviations = workbook.create_sheet("Daily_Deviations")
    signals = workbook.create_sheet("Trade_Signals")

    _write_summary(summary, result)
    _write_raw_returns(raw_returns, result)
    _write_cluster_means(cluster_means, result)
    _write_deviations(deviations, result)
    _write_signals(signals, result)

    workbook.save(output)
    return output


def _write_summary(sheet: Worksheet, result: StockSelectionResult) -> None:
    clustering = result.clustering_result
    cluster_count = clustering.cluster_count_result
    if cluster_count is None:
        raise ValueError(
            "stock-selection report requires cluster-count provenance"
        )
    rows = [
        ("Run setup", None),
        ("As-of date T", result.as_of_date),
        ("Lookback start T-w", result.window_start),
        ("Lookback end T-1", result.window_end),
        ("Lookback trading days w", result.config.lookback_window),
        ("Deviation threshold p", result.config.deviation_threshold),
        ("Stock count", result.stock_count),
        ("Selected K", clustering.requested_cluster_count),
        (None, None),
        ("Signals and quality", None),
        ("Previous winner count", result.quality.winner_count),
        ("Previous loser count", result.quality.loser_count),
        ("Neutral count", result.quality.neutral_count),
        (
            "Maximum daily cluster sum error",
            result.quality.maximum_daily_cluster_sum_error,
        ),
        (
            "Maximum cumulative cluster sum error",
            result.quality.maximum_cumulative_cluster_sum_error,
        ),
        (
            "All input returns finite",
            "YES" if result.quality.all_inputs_finite else "NO",
        ),
        ("Overall QC", _stock_selection_qc_status(result)),
        (None, None),
        ("Rules", None),
        ("Return input", "raw stock price returns"),
        ("Winner", "cumulative deviation > p"),
        ("Loser", "cumulative deviation < -p"),
        ("Equality", "neutral when -p <= deviation <= p"),
        (None, None),
        ("Provenance", None),
        ("Clustering snapshot id", clustering.clustering_snapshot_id),
        ("Cluster-count snapshot id", cluster_count.snapshot_id),
        ("Preprocessing run id", clustering.preprocessing_run_id),
        ("Preprocessing version", clustering.source_calculation_version),
        ("Clustering version", clustering.calculation_version),
        ("Stock-selection version", result.calculation_version),
    ]
    for row in rows:
        sheet.append(row)

    sheet.insert_rows(1)
    sheet.merge_cells("A1:B1")
    sheet["A1"] = f"Stock Selection — {result.as_of_date.isoformat()}"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    section_rows = (2, 11, 20, 26)
    for row in section_rows:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1).fill = SECTION_FILL
        sheet.cell(row, 1).font = Font(bold=True, color=NAVY)
        sheet.cell(row, 1).border = Border(bottom=THIN_GRAY)

    rows_by_label = {
        sheet.cell(row, 1).value: row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value is not None
    }
    for label in ("As-of date T", "Lookback start T-w", "Lookback end T-1"):
        sheet.cell(rows_by_label[label], 2).number_format = "yyyy-mm-dd"
    for label in (
        "Lookback trading days w",
        "Stock count",
        "Selected K",
        "Previous winner count",
        "Previous loser count",
        "Neutral count",
    ):
        sheet.cell(rows_by_label[label], 2).number_format = "0"
    sheet.cell(
        rows_by_label["Deviation threshold p"],
        2,
    ).number_format = "0.00%"
    for label in (
        "Maximum daily cluster sum error",
        "Maximum cumulative cluster sum error",
    ):
        sheet.cell(rows_by_label[label], 2).number_format = SCIENTIFIC_FORMAT

    qc_cell = sheet.cell(rows_by_label["Overall QC"], 2)
    qc_cell.font = Font(
        bold=True,
        color=GREEN if qc_cell.value == "OK" else RED,
    )
    qc_cell.fill = OK_FILL if qc_cell.value == "OK" else CHECK_FILL
    for row in range(3, sheet.max_row + 1):
        if sheet.cell(row, 1).value is not None and row not in section_rows:
            sheet.cell(row, 1).font = Font(bold=True)

    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 62


def _write_raw_returns(sheet: Worksheet, result: StockSelectionResult) -> None:
    sheet.append(["Trade Date", "Ticker", "Cluster ID", "Raw Return"])
    labels = result.clustering_result.cluster_labels
    for trade_date in result.raw_return_matrix.index:
        for column_index, ticker in enumerate(result.tickers):
            sheet.append(
                [
                    trade_date.date(),
                    ticker,
                    labels[column_index],
                    float(result.raw_return_matrix.loc[trade_date, ticker]),
                ]
            )
    _style_table(sheet, widths=(16, 14, 12, 16))
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row, 3).number_format = "0"
        sheet.cell(row, 4).number_format = RETURN_FORMAT


def _write_cluster_means(
    sheet: Worksheet,
    result: StockSelectionResult,
) -> None:
    sheet.append(["Trade Date", "Cluster ID", "Cluster Mean Return"])
    for trade_date in result.cluster_mean_return_matrix.index:
        for cluster_id in result.cluster_mean_return_matrix.columns:
            sheet.append(
                [
                    trade_date.date(),
                    int(cluster_id),
                    float(
                        result.cluster_mean_return_matrix.loc[
                            trade_date,
                            cluster_id,
                        ]
                    ),
                ]
            )
    _style_table(sheet, widths=(16, 12, 22))
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row, 2).number_format = "0"
        sheet.cell(row, 3).number_format = RETURN_FORMAT


def _write_deviations(
    sheet: Worksheet,
    result: StockSelectionResult,
) -> None:
    sheet.append(
        [
            "Trade Date",
            "Ticker",
            "Cluster ID",
            "Raw Return",
            "Cluster Mean Return",
            "Daily Deviation",
        ]
    )
    labels = result.clustering_result.cluster_labels
    for trade_date in result.daily_deviation_matrix.index:
        for column_index, ticker in enumerate(result.tickers):
            cluster_id = labels[column_index]
            sheet.append(
                [
                    trade_date.date(),
                    ticker,
                    cluster_id,
                    float(result.raw_return_matrix.loc[trade_date, ticker]),
                    float(
                        result.cluster_mean_return_matrix.loc[
                            trade_date,
                            cluster_id,
                        ]
                    ),
                    float(result.daily_deviation_matrix.loc[trade_date, ticker]),
                ]
            )
    _style_table(sheet, widths=(16, 14, 12, 16, 22, 18))
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row, 3).number_format = "0"
        for column in range(4, 7):
            sheet.cell(row, column).number_format = RETURN_FORMAT


def _write_signals(sheet: Worksheet, result: StockSelectionResult) -> None:
    sheet.append(
        [
            "Ticker",
            "Market Cap Rank",
            "Cluster ID",
            "Cluster Size",
            "Cumulative Deviation",
            "Classification",
        ]
    )
    clustering = result.clustering_result
    for index, ticker in enumerate(result.tickers):
        cluster_id = clustering.cluster_labels[index]
        sheet.append(
            [
                ticker,
                clustering.market_cap_ranks[index],
                cluster_id,
                clustering.cluster_sizes[cluster_id],
                result.cumulative_deviations[index],
                result.classifications[index],
            ]
        )
    _style_table(sheet, widths=(14, 18, 12, 14, 22, 22))
    for row in range(2, sheet.max_row + 1):
        for column in (2, 3, 4):
            sheet.cell(row, column).number_format = "0"
        sheet.cell(row, 5).number_format = CUMULATIVE_RETURN_FORMAT

    classification_range = f"F2:F{sheet.max_row}"
    sheet.conditional_formatting.add(
        classification_range,
        FormulaRule(
            formula=['F2="previous_loser"'],
            fill=OK_FILL,
        ),
    )
    sheet.conditional_formatting.add(
        classification_range,
        FormulaRule(
            formula=['F2="previous_winner"'],
            fill=CHECK_FILL,
        ),
    )
    sheet.conditional_formatting.add(
        classification_range,
        FormulaRule(
            formula=['F2="neutral"'],
            fill=NEUTRAL_FILL,
        ),
    )


def _stock_selection_qc_status(result: StockSelectionResult) -> str:
    quality = result.quality
    checks = (
        quality.all_inputs_finite,
        quality.winner_count
        + quality.loser_count
        + quality.neutral_count
        == result.stock_count,
        quality.maximum_daily_cluster_sum_error < 1e-12,
        quality.maximum_cumulative_cluster_sum_error < 1e-12,
        len(result.cumulative_deviations) == result.stock_count,
        len(result.classifications) == result.stock_count,
    )
    return "OK" if all(checks) else "CHECK"


def _style_table(sheet: Worksheet, *, widths: tuple[int, ...]) -> None:
    _style_header(sheet, columns=len(widths))
    sheet.freeze_panes = "A2"
    last_column = sheet.cell(1, len(widths)).column_letter
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            sheet.cell(1, column).column_letter
        ].width = width
    for row in range(2, sheet.max_row + 1):
        if row % 2 == 0:
            for column in range(1, len(widths) + 1):
                sheet.cell(row, column).fill = PatternFill(
                    "solid",
                    fgColor="F8FAFC",
                )


def _style_header(sheet: Worksheet, *, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(1, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[1].height = 32
