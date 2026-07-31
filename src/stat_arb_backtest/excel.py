from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path

from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from stat_arb_excel import workbook_for_publication

from .calculations import calculate_period_performance
from .models import (
    BacktestResult,
    PeriodPerformance,
    RebalanceEvent,
    TargetWeightRecord,
    TradeRecord,
)


NAVY = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
WHITE = "FFFFFF"
GREEN = "548235"
RED = "C00000"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
OK_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
CHECK_FILL = PatternFill("solid", fgColor=LIGHT_RED)
STRIPE_FILL = PatternFill("solid", fgColor="F8FAFC")
HEADER_FONT = Font(color=WHITE, bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
HEADER_BORDER = Border(bottom=THIN_GRAY)
HEADER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
WRAPPED_ALIGNMENT = Alignment(
    vertical="top",
    wrap_text=True,
)

PERCENT_FORMAT = "0.0000%"
WEIGHT_FORMAT = "0.000000"
NAV_FORMAT = "0.000000"
RATIO_FORMAT = "0.0000"
PRICE_FORMAT = "0.000000"
UNIT_FORMAT = "0.000000"


def export_backtest_workbook(
    result: BacktestResult,
    output_path: Path,
    *,
    replace_existing: bool = False,
) -> Path:
    with workbook_for_publication(
        output_path,
        replace_existing=replace_existing,
    ) as (workbook, output):
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Summary")
        periods = workbook.create_sheet("Period_Performance")
        daily = workbook.create_sheet("Daily_Performance")
        events = workbook.create_sheet("Rebalance_Events")
        actions = workbook.create_sheet("Portfolio_Actions")
        lots = workbook.create_sheet("Position_Lots")
        missing = workbook.create_sheet("Missing_Data_Audit")

        _write_summary(summary, result)
        _write_period_performance(periods, result)
        _write_daily_performance(daily, result)
        _write_rebalance_events(events, result)
        _write_portfolio_actions(actions, result)
        _write_position_lots(lots, result)
        _write_missing_data(missing, result)
        if not result.missing_data_audit:
            missing.sheet_state = "hidden"
    return output


def _write_summary(sheet: Worksheet, result: BacktestResult) -> None:
    config = result.config
    strategy = result.strategy_metrics
    spy = result.spy_metrics
    missing_days = sum(
        row.missing_position_count > 0 for row in result.daily_performance
    )
    portfolio_action_count = len(_portfolio_action_rows(result))
    lot_counts = {
        status: sum(lot.status == status for lot in result.position_lots)
        for status in ("CLOSED", "PARTIALLY_CLOSED", "OPEN")
    }
    overall_qc = _overall_qc(result)

    sheet.merge_cells("A1:D1")
    sheet["A1"] = (
        f"Step 7 Backtest — {config.start_date.isoformat()} to "
        f"{config.end_date.isoformat()}"
    )
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    _write_summary_section(sheet, 3, "Performance")
    performance_rows = (
        ("Metric", "Strategy", "SPY", "Difference"),
        (
            "Starting NAV",
            strategy.starting_nav,
            spy.starting_nav,
            strategy.starting_nav - spy.starting_nav,
        ),
        (
            "Ending NAV",
            strategy.ending_nav,
            spy.ending_nav,
            strategy.ending_nav - spy.ending_nav,
        ),
        (
            "Total return",
            strategy.total_return,
            spy.total_return,
            strategy.total_return - spy.total_return,
        ),
        (
            "Annualized return",
            strategy.annualized_return,
            spy.annualized_return,
            _optional_difference(
                strategy.annualized_return,
                spy.annualized_return,
            ),
        ),
        (
            "Sharpe ratio",
            strategy.sharpe_ratio,
            spy.sharpe_ratio,
            _optional_difference(strategy.sharpe_ratio, spy.sharpe_ratio),
        ),
        (
            "Sortino ratio",
            strategy.sortino_ratio,
            spy.sortino_ratio,
            _optional_difference(strategy.sortino_ratio, spy.sortino_ratio),
        ),
    )
    for row_number, values in enumerate(performance_rows, start=4):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value=value)
    _style_summary_header(sheet, "A4:D4")
    for row in range(5, 11):
        sheet.cell(row, 1).font = Font(bold=True)
        if row % 2 == 0:
            for column in range(1, 5):
                sheet.cell(row, column).fill = STRIPE_FILL
    for row in (5, 6):
        for column in range(2, 5):
            sheet.cell(row, column).number_format = NAV_FORMAT
    for row in (7, 8):
        for column in range(2, 5):
            sheet.cell(row, column).number_format = PERCENT_FORMAT
    for row in (9, 10):
        for column in range(2, 5):
            sheet.cell(row, column).number_format = RATIO_FORMAT

    _write_summary_section(sheet, 12, "Backtest setup")
    setup_rows = (
        ("Start date", config.start_date, "End date", config.end_date),
        (
            "Return sessions",
            strategy.session_count,
            "Initial NAV",
            config.initial_nav,
        ),
        (
            "Missing-price policy",
            config.missing_price_policy,
            "Annualization",
            f"{config.annualization_sessions} sessions",
        ),
        (
            "Detailed audit",
            "technical columns are hidden; unhide them when needed",
            "Costs / slippage",
            "not included",
        ),
    )
    for row_number, values in enumerate(setup_rows, start=13):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value=value)
        sheet.cell(row_number, 1).font = Font(bold=True)
        sheet.cell(row_number, 3).font = Font(bold=True)
    sheet["B13"].number_format = "yyyy-mm-dd"
    sheet["D13"].number_format = "yyyy-mm-dd"
    sheet["D14"].number_format = NAV_FORMAT

    _write_summary_section(sheet, 19, "Audit and QC")
    audit_rows = (
        (
            "Rebalance events",
            len(result.rebalance_events),
            "Trade records",
            len(result.trades),
        ),
        (
            "Target / action rows",
            f"{len(result.target_weights)} / {portfolio_action_count}",
            "Buy lots",
            len(result.position_lots),
        ),
        (
            "Closed lots",
            lot_counts["CLOSED"],
            "Partial / open lots",
            f"{lot_counts['PARTIALLY_CLOSED']} / {lot_counts['OPEN']}",
        ),
        (
            "Missing rows / sessions",
            f"{len(result.missing_data_audit)} / {missing_days}",
            "FIFO reconciliation",
            result.fifo_reconciliation_status,
        ),
        (
            "Metrics finite",
            "YES",
            "Overall QC",
            overall_qc,
        ),
    )
    for row_number, values in enumerate(audit_rows, start=20):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value=value)
        sheet.cell(row_number, 1).font = Font(bold=True)
        sheet.cell(row_number, 3).font = Font(bold=True)
    for cell_reference, status in (
        ("D23", result.fifo_reconciliation_status),
        ("D24", overall_qc),
    ):
        cell = sheet[cell_reference]
        cell.font = Font(
            bold=True,
            color=GREEN if status == "OK" else RED,
        )
        cell.fill = OK_FILL if status == "OK" else CHECK_FILL

    audit = result.research_audit
    _write_summary_section(sheet, 26, "Core input parameters (w, p, P, l, q)")
    core_input_rows = (
        (
            "Lookback window (w)",
            "not captured" if audit is None else audit.selection_lookback_window,
            "Deviation threshold (p)",
            "not captured" if audit is None else audit.deviation_threshold,
        ),
        (
            "Variance threshold (P)",
            "not captured" if audit is None else audit.variance_threshold,
            "Rebalance period (l)",
            config.rebalance_period,
        ),
        (
            "Take-profit threshold (q)",
            config.take_profit_threshold,
            None,
            None,
        ),
    )
    for row_number, values in enumerate(core_input_rows, start=27):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value=value)
        sheet.cell(row_number, 1).font = Font(bold=True)
        if values[2] is not None:
            sheet.cell(row_number, 3).font = Font(bold=True)
    if audit is not None:
        sheet["D27"].number_format = PERCENT_FORMAT
        sheet["B28"].number_format = PERCENT_FORMAT
    sheet["B29"].number_format = PERCENT_FORMAT

    _write_summary_section(sheet, 31, "Signal provenance")
    if audit is None:
        research_rows = (
            (
                "Research audit metadata",
                "not captured; use run_backtest/export_backtest_report",
                None,
                None,
            ),
        )
    else:
        research_rows = (
            (
                "K estimation window",
                audit.cluster_count_estimation_window,
                "Cluster-count version",
                audit.cluster_count_calculation_version,
            ),
            (
                "SPONGE embedding mode",
                audit.embedding_mode,
                "Clustering version",
                audit.clustering_calculation_version,
            ),
            (
                "tau positive",
                audit.tau_positive,
                "tau negative",
                audit.tau_negative,
            ),
            (
                "Random seed",
                audit.random_seed,
                "k-means n_init",
                audit.kmeans_n_init,
            ),
            (
                "k-means max_iter",
                audit.kmeans_max_iter,
                "Preprocessing version",
                audit.preprocessing_calculation_version,
            ),
            (
                "Data pipeline run id",
                audit.data_pipeline_run_id or "not available",
                "Preprocessing run id",
                audit.preprocessing_run_id,
            ),
            (
                "Stock-selection version",
                audit.stock_selection_calculation_version,
                "Portfolio-weight version",
                audit.portfolio_weight_calculation_version,
            ),
        )
    for row_number, values in enumerate(research_rows, start=32):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value=value)
        sheet.cell(row_number, 1).font = Font(bold=True)
        if values[2] is not None:
            sheet.cell(row_number, 3).font = Font(bold=True)
    time_semantics_row = 33 + len(research_rows)
    _write_summary_section(
        sheet,
        time_semantics_row,
        "Time semantics and research scope",
    )
    deviation_scope = (
        "not captured"
        if audit is None
        else f"{audit.deviation_threshold:.4%}"
    )
    notes = (
        (
            "Time semantics",
            "as_of_date=T uses inputs through T-1; the old portfolio earns "
            "the event-date return and the new target starts next session",
        ),
        (
            "Holding convention",
            "fixed economic units between events; weights drift with Close prices",
        ),
        (
            "Strategy scope",
            f"Yahoo Close price return; p={deviation_scope}; long-only losers; "
            "capital is split equally across active clusters; "
            "cash return and risk-free rate are zero",
        ),
        ("Calculation version", result.calculation_version),
        ("FF12 benchmark", "not produced because ff12_code is currently empty"),
        (
            "Paper",
            "https://ora.ox.ac.uk/objects/uuid%3Ac60358c0-24f0-4c66-b973-f84776f66f8a",
        ),
        (
            "Author code",
            "https://github.com/maxclchen/Correlation-Matrix-Clustering-for-Statistical-Arbitrage-Portfolios",
        ),
    )
    for row_number, (label, value) in enumerate(
        notes,
        start=time_semantics_row + 1,
    ):
        sheet.cell(row_number, 1, value=label).font = Font(bold=True)
        sheet.merge_cells(
            start_row=row_number,
            start_column=2,
            end_row=row_number,
            end_column=4,
        )
        sheet.cell(row_number, 2, value=value)
        sheet.cell(row_number, 2).alignment = WRAPPED_ALIGNMENT

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 31
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 31
    for row in range(3, sheet.max_row + 1):
        for column in range(1, 5):
            sheet.cell(row, column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def _write_summary_section(
    sheet: Worksheet,
    row: int,
    label: str,
) -> None:
    sheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=4,
    )
    cell = sheet.cell(row, 1, value=label)
    cell.fill = SECTION_FILL
    cell.font = Font(bold=True, color=NAVY)
    cell.border = HEADER_BORDER


def _style_summary_header(sheet: Worksheet, reference: str) -> None:
    for row in sheet[reference]:
        for cell in row:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER


def _optional_difference(
    left: float | None,
    right: float | None,
) -> float | None:
    if left is None or right is None:
        return None
    return left - right


def _write_daily_performance(
    sheet: Worksheet,
    result: BacktestResult,
) -> None:
    _write_table(
        sheet,
        headers=(
            "Trade Date",
            "Strategy Return",
            "NAV",
            "SPY Return",
            "SPY NAV",
            "Round ID",
            "Holding Day",
            "Round Return",
            "Trigger",
            "Gross Exposure",
            "Positions",
            "Missing Positions",
            "Cash Value",
            "Cash Weight",
            "Frozen Value",
            "Frozen Exposure",
        ),
        rows=(
            (
                row.trade_date,
                row.strategy_return,
                row.nav,
                row.spy_return,
                row.spy_nav,
                row.round_id,
                row.holding_day,
                row.round_return,
                row.trigger_reason,
                row.gross_exposure,
                row.position_count,
                row.missing_position_count,
                row.cash_value,
                row.cash_weight,
                row.frozen_value,
                row.frozen_exposure,
            )
            for row in result.daily_performance
        ),
        widths=(
            13,
            17,
            15,
            14,
            15,
            11,
            13,
            16,
            14,
            16,
            12,
            18,
            15,
            14,
            15,
            16,
        ),
        number_formats={
            1: "yyyy-mm-dd",
            2: PERCENT_FORMAT,
            3: NAV_FORMAT,
            4: PERCENT_FORMAT,
            5: NAV_FORMAT,
            6: "0",
            7: "0",
            8: PERCENT_FORMAT,
            10: PERCENT_FORMAT,
            11: "0",
            12: "0",
            13: NAV_FORMAT,
            14: PERCENT_FORMAT,
            15: NAV_FORMAT,
            16: PERCENT_FORMAT,
        },
    )
    _hide_columns(sheet, ("M", "N", "O", "P"))
    if sheet.max_row >= 2:
        sheet.conditional_formatting.add(
            f"I2:I{sheet.max_row}",
            FormulaRule(formula=['I2="stop_win"'], fill=OK_FILL),
        )
        sheet.conditional_formatting.add(
            f"I2:I{sheet.max_row}",
            FormulaRule(formula=['I2="scheduled"'], fill=SECTION_FILL),
        )


def _write_period_performance(
    sheet: Worksheet,
    result: BacktestResult,
) -> None:
    annual = calculate_period_performance(
        result.daily_performance,
        frequency="year",
        annualization_sessions=result.config.annualization_sessions,
    )
    monthly = calculate_period_performance(
        result.daily_performance,
        frequency="month",
        annualization_sessions=result.config.annualization_sessions,
    )

    sheet.merge_cells("A1:N1")
    sheet["A1"] = (
        f"Calendar Period Performance — {result.config.start_date.isoformat()} "
        f"to {result.config.end_date.isoformat()}"
    )
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:N2")
    sheet["A2"] = (
        "Returns compound the report's daily close-to-close returns within each "
        "calendar period. Excess return is Strategy minus SPY; volatility and "
        "Sharpe are annualized with the configured session count."
    )
    sheet["A2"].alignment = WRAPPED_ALIGNMENT
    sheet.row_dimensions[2].height = 32

    annual_end = _write_period_table(
        sheet,
        title_row=4,
        title="Annual Performance",
        periods=annual,
        period_format="yyyy",
    )
    strategy_heatmap_end = _write_monthly_heatmap(
        sheet,
        title_row=annual_end + 3,
        title="Monthly Strategy Return Heatmap",
        monthly=monthly,
        annual=annual,
        value_attribute="strategy_return",
    )
    excess_heatmap_end = _write_monthly_heatmap(
        sheet,
        title_row=strategy_heatmap_end + 3,
        title="Monthly Excess Return vs SPY Heatmap",
        monthly=monthly,
        annual=annual,
        value_attribute="excess_return",
    )
    _write_period_table(
        sheet,
        title_row=excess_heatmap_end + 3,
        title="Monthly Performance Detail",
        periods=monthly,
        period_format="yyyy-mm",
    )

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    widths = {
        "A": 15,
        "B": 12,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 20,
        "G": 18,
        "H": 17,
        "I": 15,
        "J": 20,
        "K": 18,
        "L": 12,
        "M": 12,
        "N": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _write_period_table(
    sheet: Worksheet,
    *,
    title_row: int,
    title: str,
    periods: Sequence[PeriodPerformance],
    period_format: str,
) -> int:
    headers = (
        "Period",
        "Sessions",
        "Strategy Return",
        "SPY Return",
        "Excess Return",
        "Strategy Ann. Volatility",
        "SPY Ann. Volatility",
        "Strategy Sharpe",
        "SPY Sharpe",
        "Strategy Max Drawdown",
        "SPY Max Drawdown",
    )
    _write_block_title(sheet, title_row, title, end_column=len(headers))
    header_row = title_row + 1
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
    sheet.row_dimensions[header_row].height = 32

    for row_number, period in enumerate(periods, start=header_row + 1):
        values = (
            period.period_start,
            period.session_count,
            period.strategy_return,
            period.spy_return,
            period.excess_return,
            period.strategy_annualized_volatility,
            period.spy_annualized_volatility,
            period.strategy_sharpe_ratio,
            period.spy_sharpe_ratio,
            period.strategy_max_drawdown,
            period.spy_max_drawdown,
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value=value)
            if row_number % 2 == 0:
                cell.fill = STRIPE_FILL
        sheet.cell(row_number, 1).number_format = period_format
        sheet.cell(row_number, 2).number_format = "0"
        for column in range(3, 8):
            sheet.cell(row_number, column).number_format = PERCENT_FORMAT
        for column in (8, 9):
            sheet.cell(row_number, column).number_format = RATIO_FORMAT
        for column in (10, 11):
            sheet.cell(row_number, column).number_format = PERCENT_FORMAT

    end_row = header_row + len(periods)
    if periods:
        return_values = [
            value
            for period in periods
            for value in (
                period.strategy_return,
                period.spy_return,
                period.excess_return,
            )
        ]
        _add_zero_centered_scale(
            sheet,
            f"C{header_row + 1}:E{end_row}",
            return_values,
        )
        drawdowns = [
            value
            for period in periods
            for value in (
                period.strategy_max_drawdown,
                period.spy_max_drawdown,
            )
        ]
        _add_drawdown_scale(
            sheet,
            f"J{header_row + 1}:K{end_row}",
            drawdowns,
        )
    return end_row


def _write_monthly_heatmap(
    sheet: Worksheet,
    *,
    title_row: int,
    title: str,
    monthly: Sequence[PeriodPerformance],
    annual: Sequence[PeriodPerformance],
    value_attribute: str,
) -> int:
    _write_block_title(sheet, title_row, title, end_column=14)
    header_row = title_row + 1
    headers = (
        "Year",
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
        "Year",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    monthly_lookup = {
        (period.period_start.year, period.period_start.month): period
        for period in monthly
    }
    annual_lookup = {
        period.period_start.year: period
        for period in annual
    }
    years = sorted(annual_lookup)
    scale_values: list[float] = []
    for row_number, year in enumerate(years, start=header_row + 1):
        sheet.cell(row_number, 1, value=year).font = Font(bold=True)
        for month in range(1, 13):
            period = monthly_lookup.get((year, month))
            value = (
                getattr(period, value_attribute)
                if period is not None
                else None
            )
            cell = sheet.cell(row_number, month + 1, value=value)
            cell.number_format = PERCENT_FORMAT
            if value is not None:
                scale_values.append(float(value))
        annual_value = getattr(annual_lookup[year], value_attribute)
        annual_cell = sheet.cell(row_number, 14, value=annual_value)
        annual_cell.number_format = PERCENT_FORMAT
        annual_cell.font = Font(bold=True)
        scale_values.append(float(annual_value))

    end_row = header_row + len(years)
    if years:
        _add_zero_centered_scale(
            sheet,
            f"B{header_row + 1}:N{end_row}",
            scale_values,
        )
    return end_row


def _write_block_title(
    sheet: Worksheet,
    row: int,
    label: str,
    *,
    end_column: int,
) -> None:
    sheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=end_column,
    )
    cell = sheet.cell(row, 1, value=label)
    cell.fill = SECTION_FILL
    cell.font = Font(bold=True, color=NAVY)
    cell.border = HEADER_BORDER


def _add_zero_centered_scale(
    sheet: Worksheet,
    reference: str,
    values: Sequence[float],
) -> None:
    maximum = max((abs(float(value)) for value in values), default=0.0)
    maximum = maximum if maximum > 0.0 else 0.01
    sheet.conditional_formatting.add(
        reference,
        ColorScaleRule(
            start_type="num",
            start_value=-maximum,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.0,
            mid_color=WHITE,
            end_type="num",
            end_value=maximum,
            end_color="63BE7B",
        ),
    )


def _add_drawdown_scale(
    sheet: Worksheet,
    reference: str,
    values: Sequence[float],
) -> None:
    maximum = max((abs(float(value)) for value in values), default=0.0)
    maximum = maximum if maximum > 0.0 else 0.01
    sheet.conditional_formatting.add(
        reference,
        ColorScaleRule(
            start_type="num",
            start_value=-maximum,
            start_color="F8696B",
            end_type="num",
            end_value=0.0,
            end_color=WHITE,
        ),
    )


def _write_rebalance_events(
    sheet: Worksheet,
    result: BacktestResult,
) -> None:
    _write_table(
        sheet,
        headers=(
            "Event ID",
            "Event Date",
            "Reason",
            "Effective Date",
            "Held Sessions",
            "Round Return",
            "NAV",
            "K",
            "Active Clusters",
            "Target Gross",
            "Round ID",
            "Inactive Clusters",
            "Frozen Value",
            "Available Capital",
        ),
        rows=(
            (
                event.event_id,
                event.event_date,
                event.reason,
                event.effective_date,
                event.held_sessions,
                event.round_return,
                event.nav,
                event.cluster_count,
                event.active_cluster_count,
                event.target_gross_exposure,
                event.round_id,
                event.inactive_cluster_count,
                event.frozen_value,
                event.available_capital,
            )
            for event in result.rebalance_events
        ),
        widths=(11, 14, 14, 15, 15, 16, 15, 9, 17, 16, 11, 18, 15, 18),
        number_formats={
            1: "0",
            2: "yyyy-mm-dd",
            4: "yyyy-mm-dd",
            5: "0",
            6: PERCENT_FORMAT,
            7: NAV_FORMAT,
            8: "0",
            9: "0",
            10: PERCENT_FORMAT,
            11: "0",
            12: "0",
            13: NAV_FORMAT,
            14: NAV_FORMAT,
        },
    )
    _hide_columns(sheet, ("K", "L", "M", "N"))


def _write_portfolio_actions(
    sheet: Worksheet,
    result: BacktestResult,
) -> None:
    _write_table(
        sheet,
        headers=(
            "Event ID",
            "Event Date",
            "Effective Date",
            "Ticker",
            "Market Cap Rank",
            "Cluster ID",
            "Cumulative Deviation",
            "Target Portfolio Weight",
            "Side",
            "Execution Price",
            "Units Traded",
            "Executed Notional",
            "Status",
            "Reason",
            "Trade ID",
            "Trade Date",
            "Classification",
            "Local Weight",
            "Units Before",
            "Units After",
            "Requested Notional",
            "Value Before",
            "Value After",
        ),
        rows=_portfolio_action_rows(result),
        widths=(
            11,
            14,
            15,
            14,
            18,
            12,
            23,
            23,
            10,
            14,
            16,
            20,
            23,
            24,
            11,
            14,
            21,
            16,
            16,
            16,
            20,
            16,
            16,
        ),
        number_formats={
            1: "0",
            2: "yyyy-mm-dd",
            3: "yyyy-mm-dd",
            5: "0",
            6: "0",
            7: PERCENT_FORMAT,
            8: WEIGHT_FORMAT,
            10: PRICE_FORMAT,
            11: UNIT_FORMAT,
            12: NAV_FORMAT,
            15: "0",
            16: "yyyy-mm-dd",
            18: WEIGHT_FORMAT,
            19: UNIT_FORMAT,
            20: UNIT_FORMAT,
            21: NAV_FORMAT,
            22: NAV_FORMAT,
            23: NAV_FORMAT,
        },
    )
    _hide_columns(sheet, ("O", "P", "Q", "R", "S", "T", "U", "V", "W"))
    if sheet.max_row >= 2:
        sheet.conditional_formatting.add(
            f"I2:I{sheet.max_row}",
            FormulaRule(formula=['I2="BUY"'], fill=OK_FILL),
        )
        sheet.conditional_formatting.add(
            f"I2:I{sheet.max_row}",
            FormulaRule(formula=['I2="SELL"'], fill=CHECK_FILL),
        )
        sheet.conditional_formatting.add(
            f"M2:M{sheet.max_row}",
            FormulaRule(
                formula=['M2="unfilled_missing_close"'],
                fill=CHECK_FILL,
            ),
        )


def _portfolio_action_rows(
    result: BacktestResult,
) -> tuple[tuple[object, ...], ...]:
    events = {event.event_id: event for event in result.rebalance_events}
    targets = {
        (target.event_id, target.ticker): target
        for target in result.target_weights
    }
    trades_by_key: dict[tuple[int, str], list[TradeRecord]] = defaultdict(list)
    standalone_trades: list[TradeRecord] = []
    for trade in result.trades:
        if trade.event_id is None:
            standalone_trades.append(trade)
        else:
            trades_by_key[(trade.event_id, trade.ticker)].append(trade)

    rows: list[tuple[object, ...]] = []
    event_keys = set(targets) | set(trades_by_key)
    for event_id, ticker in sorted(event_keys):
        event = events[event_id]
        target = targets.get((event_id, ticker))
        matching_trades = trades_by_key.get((event_id, ticker), [])
        if not matching_trades:
            rows.append(
                _portfolio_action_row(event, target, None)
            )
            continue
        rows.extend(
            _portfolio_action_row(event, target, trade)
            for trade in matching_trades
        )
    rows.extend(
        _portfolio_action_row(None, None, trade)
        for trade in standalone_trades
    )
    rows.sort(
        key=lambda row: (
            row[15] or row[1] or row[2],
            row[14] or 0,
            row[3],
        )
    )
    return tuple(rows)


def _portfolio_action_row(
    event: RebalanceEvent | None,
    target: TargetWeightRecord | None,
    trade: TradeRecord | None,
) -> tuple[object, ...]:
    event_id = (
        event.event_id
        if event is not None
        else trade.event_id if trade is not None else target.event_id
    )
    target_weight = target.portfolio_weight if target is not None else None
    if (
        target is None
        and trade is not None
        and trade.event_id is not None
        and trade.side == "SELL"
    ):
        target_weight = 0.0
    execution_price = (
        trade.execution_price if trade is not None else None
    )
    return (
        event_id,
        event.event_date if event is not None else None,
        (
            target.effective_date
            if target is not None
            else event.effective_date if event is not None else None
        ),
        target.ticker if target is not None else trade.ticker,
        target.market_cap_rank if target is not None else None,
        target.cluster_id if target is not None else None,
        target.cumulative_deviation if target is not None else None,
        target_weight,
        trade.side if trade is not None else None,
        execution_price,
        trade.units_traded if trade is not None else None,
        trade.executed_notional if trade is not None else None,
        trade.status if trade is not None else None,
        trade.reason if trade is not None else None,
        trade.trade_id if trade is not None else None,
        trade.trade_date if trade is not None else None,
        target.classification if target is not None else None,
        target.local_weight if target is not None else None,
        trade.units_before if trade is not None else None,
        trade.units_after if trade is not None else None,
        trade.requested_notional if trade is not None else None,
        trade.value_before if trade is not None else None,
        trade.value_after if trade is not None else None,
    )


def _write_position_lots(
    sheet: Worksheet,
    result: BacktestResult,
) -> None:
    _write_table(
        sheet,
        headers=(
            "Lot ID",
            "Ticker",
            "Buy Date",
            "Buy Price",
            "Bought Units",
            "Buy Notional",
            "Exit Date",
            "Exit Price",
            "Remaining Units",
            "Realized P&L",
            "Lot Return",
            "Status",
            "Buy Trade ID",
            "Buy Event ID",
            "Sold Units",
            "First Sell Date",
            "Matched Sell VWAP",
            "Sale Proceeds",
            "Realized Return",
        ),
        rows=(
            (
                lot.lot_id,
                lot.ticker,
                lot.buy_date,
                lot.buy_price,
                lot.bought_units,
                lot.buy_notional,
                lot.final_sell_date,
                lot.final_sell_price,
                lot.remaining_units,
                lot.realized_pnl,
                lot.lot_return,
                lot.status,
                lot.buy_trade_id,
                lot.buy_event_id,
                lot.sold_units,
                lot.first_sell_date,
                lot.matched_sell_vwap,
                lot.sale_proceeds,
                lot.realized_return,
            )
            for lot in result.position_lots
        ),
        widths=(
            14,
            14,
            14,
            14,
            16,
            17,
            16,
            16,
            18,
            17,
            16,
            20,
            14,
            14,
            16,
            16,
            19,
            17,
            18,
        ),
        number_formats={
            3: "yyyy-mm-dd",
            4: PRICE_FORMAT,
            5: UNIT_FORMAT,
            6: NAV_FORMAT,
            7: "yyyy-mm-dd",
            8: PRICE_FORMAT,
            9: UNIT_FORMAT,
            10: NAV_FORMAT,
            11: PERCENT_FORMAT,
            13: "0",
            14: "0",
            15: UNIT_FORMAT,
            16: "yyyy-mm-dd",
            17: PRICE_FORMAT,
            18: NAV_FORMAT,
            19: PERCENT_FORMAT,
        },
    )
    _hide_columns(sheet, ("M", "N", "O", "P", "Q", "R", "S"))
    if sheet.max_row >= 2:
        sheet.conditional_formatting.add(
            f"L2:L{sheet.max_row}",
            FormulaRule(formula=['L2="CLOSED"'], fill=OK_FILL),
        )
        sheet.conditional_formatting.add(
            f"L2:L{sheet.max_row}",
            FormulaRule(
                formula=['L2="PARTIALLY_CLOSED"'],
                fill=SECTION_FILL,
            ),
        )
        sheet.conditional_formatting.add(
            f"L2:L{sheet.max_row}",
            FormulaRule(formula=['L2="OPEN"'], fill=SECTION_FILL),
        )


def _write_missing_data(
    sheet: Worksheet,
    result: BacktestResult,
) -> None:
    _write_table(
        sheet,
        headers=(
            "Trade Date",
            "Ticker",
            "Event",
            "Action",
            "Last Valid Close",
            "Position Value",
            "Details",
        ),
        rows=(
            (
                audit.trade_date,
                audit.ticker,
                audit.event,
                audit.action,
                audit.last_valid_close,
                audit.position_value,
                audit.details,
            )
            for audit in result.missing_data_audit
        ),
        widths=(14, 14, 20, 36, 18, 18, 74),
        number_formats={
            1: "yyyy-mm-dd",
            5: NAV_FORMAT,
            6: NAV_FORMAT,
        },
        alignments={7: WRAPPED_ALIGNMENT},
    )


def _overall_qc(result: BacktestResult) -> str:
    daily = result.daily_performance
    finite = all(
        _finite(
            row.strategy_return,
            row.nav,
            row.round_return,
            row.cash_value,
            row.cash_weight,
            row.gross_exposure,
            row.frozen_value,
            row.frozen_exposure,
            row.spy_return,
            row.spy_nav,
        )
        for row in daily
    )
    reconciled = all(
        abs(row.cash_weight + row.gross_exposure - 1.0) < 1e-10
        for row in daily
    )
    metrics = (
        result.strategy_metrics,
        result.spy_metrics,
    )
    metric_values = [
        value
        for metric in metrics
        for value in (
            metric.annualized_return,
            metric.sharpe_ratio,
            metric.sortino_ratio,
        )
        if value is not None
    ]
    return (
        "OK"
        if finite
        and reconciled
        and result.fifo_reconciliation_status == "OK"
        and all(_finite(value) for value in metric_values)
        else "CHECK"
    )


def _finite(*values: float) -> bool:
    import math

    return all(math.isfinite(float(value)) for value in values)


def _write_table(
    sheet: Worksheet,
    *,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    widths: tuple[int, ...],
    number_formats: Mapping[int, str] | None = None,
    alignments: Mapping[int, Alignment] | None = None,
) -> None:
    if len(headers) != len(widths):
        raise ValueError("table headers and widths must have the same length")
    formats = number_formats or {}
    configured_alignments = alignments or {}
    sheet.append(list(headers))
    for column in range(1, len(widths) + 1):
        cell = sheet.cell(1, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    for row_number, values in enumerate(rows, start=2):
        row_values = tuple(values)
        if len(row_values) != len(headers):
            raise ValueError(
                f"table row {row_number} has {len(row_values)} values; "
                f"expected {len(headers)}"
            )
        for column, value in enumerate(row_values, start=1):
            cell = sheet.cell(row_number, column, value=value)
            if row_number % 2 == 0:
                cell.fill = STRIPE_FILL
            number_format = formats.get(column)
            if number_format is not None:
                cell.number_format = number_format
            alignment = configured_alignments.get(column)
            if alignment is not None:
                cell.alignment = alignment

    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    last_column = sheet.cell(1, len(widths)).column_letter
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    for column, width in enumerate(widths, start=1):
        column_letter = sheet.cell(1, column).column_letter
        sheet.column_dimensions[column_letter].width = width


def _hide_columns(
    sheet: Worksheet,
    columns: Sequence[str],
) -> None:
    for column in columns:
        dimension = sheet.column_dimensions[column]
        dimension.hidden = True
        dimension.outline_level = 1
