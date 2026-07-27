from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .models import PortfolioWeightResult


NAVY = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
GREEN = "548235"
RED = "C00000"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
OK_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
CHECK_FILL = PatternFill("solid", fgColor=LIGHT_RED)
NEUTRAL_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
HEADER_FONT = Font(color=WHITE, bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")

WEIGHT_FORMAT = "0.000000"
RETURN_FORMAT = "0.00%"
QC_TOLERANCE = 1e-12


def export_portfolio_weight_workbook(
    result: PortfolioWeightResult,
    output_path: Path,
    *,
    replace_existing: bool = False,
) -> Path:
    """Export a concise result workbook without duplicating audit calculations."""
    output = Path(output_path).resolve()
    if output.exists() and not replace_existing:
        raise FileExistsError(
            f"Excel output already exists: {output}. Use --replace to overwrite it."
        )
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Summary")
    clusters = workbook.create_sheet("Cluster_Allocations")
    stocks = workbook.create_sheet("Stock_Weights")

    _write_summary(summary, result)
    _write_cluster_allocations(clusters, result)
    _write_stock_weights(stocks, result)

    workbook.save(output)
    return output


def _write_summary(sheet: Worksheet, result: PortfolioWeightResult) -> None:
    selection = result.stock_selection_result
    clustering = selection.clustering_result
    quality = result.quality

    sheet.merge_cells("A1:B1")
    sheet["A1"] = f"Portfolio Weights — {result.as_of_date.isoformat()}"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    rows = [
        ("Run setup", None),
        ("As-of date T", result.as_of_date),
        ("Signal window start T-w", selection.window_start),
        ("Signal window end T-1", selection.window_end),
        ("Lookback trading days w", selection.config.lookback_window),
        ("Deviation threshold p", selection.config.deviation_threshold),
        ("Stock count", result.stock_count),
        ("Selected K", result.cluster_count),
        (None, None),
        ("Portfolio exposure", None),
        ("Active clusters", quality.active_cluster_count),
        ("Inactive clusters", quality.inactive_cluster_count),
        ("Long exposure", quality.long_exposure),
        ("Short exposure", quality.short_exposure),
        ("Net exposure", quality.net_exposure),
        ("Gross exposure", quality.gross_exposure),
        ("Uninvested gross exposure", quality.uninvested_gross_exposure),
        ("Overall QC", _overall_qc_status(result)),
        (None, None),
        ("Method", None),
        ("Return input", "raw stock price returns"),
        ("Position direction", "long only: previous losers; winners stay at zero"),
        ("Cluster allocation", "equal 1/K share of total gross exposure"),
        (
            "Inactive cluster",
            "zero weights; its gross allocation remains uninvested",
        ),
        (None, None),
        ("Calculation versions", None),
        ("Preprocessing", clustering.source_calculation_version),
        ("Clustering", clustering.calculation_version),
        ("Stock selection", selection.calculation_version),
        ("Portfolio weights", result.calculation_version),
    ]
    for row in rows:
        sheet.append(row)

    for row in (2, 11, 21, 27):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = sheet.cell(row, 1)
        cell.fill = SECTION_FILL
        cell.font = Font(bold=True, color=NAVY)
        cell.border = Border(bottom=THIN_GRAY)

    for row in (3, 4, 5):
        sheet.cell(row, 2).number_format = "yyyy-mm-dd"
    sheet.cell(7, 2).number_format = RETURN_FORMAT
    for row in (8, 9, 12, 13):
        sheet.cell(row, 2).number_format = "0"
    for row in range(14, 19):
        sheet.cell(row, 2).number_format = RETURN_FORMAT

    qc_cell = sheet["B19"]
    qc_cell.font = Font(bold=True, color=GREEN if qc_cell.value == "OK" else RED)
    qc_cell.fill = OK_FILL if qc_cell.value == "OK" else CHECK_FILL

    for row in range(3, sheet.max_row + 1):
        if sheet.cell(row, 1).value is not None and row not in (11, 21, 27):
            sheet.cell(row, 1).font = Font(bold=True)

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 52
    sheet["B24"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["B25"].alignment = Alignment(wrap_text=True, vertical="top")


def _write_cluster_allocations(
    sheet: Worksheet,
    result: PortfolioWeightResult,
) -> None:
    headers = [
        "Cluster ID",
        "Stocks",
        "Winners",
        "Losers",
        "Neutral",
        "Status",
        "Target Gross",
        "Long",
        "Short",
        "Gross",
        "Uninvested",
    ]
    sheet.append(headers)
    for allocation in result.cluster_allocations:
        sheet.append(
            [
                allocation.cluster_id,
                allocation.stock_count,
                allocation.winner_count,
                allocation.loser_count,
                allocation.neutral_count,
                "Active" if allocation.is_active else "Inactive",
                allocation.target_gross_exposure,
                allocation.portfolio_long_exposure,
                allocation.portfolio_short_exposure,
                allocation.portfolio_gross_exposure,
                allocation.uninvested_gross_exposure,
            ]
        )

    _style_table(
        sheet,
        widths=(12, 10, 11, 11, 11, 12, 15, 14, 14, 14, 14),
    )
    for row in range(2, sheet.max_row + 1):
        for column in range(1, 6):
            sheet.cell(row, column).number_format = "0"
        for column in range(7, 12):
            sheet.cell(row, column).number_format = RETURN_FORMAT

    status_range = f"F2:F{sheet.max_row}"
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Active"'], fill=OK_FILL),
    )
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Inactive"'], fill=CHECK_FILL),
    )


def _write_stock_weights(
    sheet: Worksheet,
    result: PortfolioWeightResult,
) -> None:
    headers = [
        "Ticker",
        "Market Cap Rank",
        "Cluster ID",
        "Cumulative Deviation",
        "Classification",
        "Local Weight",
        "Portfolio Weight",
    ]
    sheet.append(headers)
    selection = result.stock_selection_result
    clustering = selection.clustering_result
    for index, ticker in enumerate(result.tickers):
        sheet.append(
            [
                ticker,
                clustering.market_cap_ranks[index],
                clustering.cluster_labels[index],
                selection.cumulative_deviations[index],
                selection.classifications[index],
                result.local_weights[index],
                result.portfolio_weights[index],
            ]
        )

    _style_table(
        sheet,
        widths=(14, 18, 12, 22, 22, 16, 18),
    )
    for row in range(2, sheet.max_row + 1):
        for column in (2, 3):
            sheet.cell(row, column).number_format = "0"
        sheet.cell(row, 4).number_format = RETURN_FORMAT
        for column in (6, 7):
            sheet.cell(row, column).number_format = WEIGHT_FORMAT

    classification_range = f"E2:E{sheet.max_row}"
    sheet.conditional_formatting.add(
        classification_range,
        FormulaRule(
            formula=['E2="previous_loser"'],
            fill=OK_FILL,
        ),
    )
    sheet.conditional_formatting.add(
        classification_range,
        FormulaRule(
            formula=['E2="previous_winner"'],
            fill=CHECK_FILL,
        ),
    )
    sheet.conditional_formatting.add(
        classification_range,
        FormulaRule(
            formula=['E2="neutral"'],
            fill=NEUTRAL_FILL,
        ),
    )


def _overall_qc_status(result: PortfolioWeightResult) -> str:
    quality = result.quality
    expected_invested_gross = quality.active_cluster_count / result.cluster_count
    checks = (
        quality.all_weights_finite,
        quality.maximum_active_cluster_local_net_error < QC_TOLERANCE,
        quality.maximum_active_cluster_local_gross_error < QC_TOLERANCE,
        quality.maximum_cluster_portfolio_gross_error < QC_TOLERANCE,
        abs(quality.short_exposure) < QC_TOLERANCE,
        abs(quality.net_exposure - quality.gross_exposure) < QC_TOLERANCE,
        abs(quality.gross_exposure - expected_invested_gross) < QC_TOLERANCE,
        abs(
            quality.gross_exposure
            + quality.uninvested_gross_exposure
            - 1.0
        )
        < QC_TOLERANCE,
    )
    return "OK" if all(checks) else "CHECK"


def _style_table(sheet: Worksheet, *, widths: tuple[int, ...]) -> None:
    _style_header(sheet, columns=len(widths))
    sheet.freeze_panes = "A2"
    last_column = sheet.cell(1, len(widths)).column_letter
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    for column, width in enumerate(widths, start=1):
        column_letter = sheet.cell(1, column).column_letter
        sheet.column_dimensions[column_letter].width = width
    for row in range(2, sheet.max_row + 1):
        if row % 2 == 0:
            for column in range(1, len(widths) + 1):
                sheet.cell(row, column).fill = PatternFill(
                    "solid",
                    fgColor="F8FAFC",
                )


def _style_header(
    sheet: Worksheet,
    *,
    columns: int,
) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(1, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[1].height = 32
