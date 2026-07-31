from __future__ import annotations

from collections.abc import Callable, Iterable, Sequence
from dataclasses import dataclass, fields
from pathlib import Path

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from stat_arb_excel import workbook_for_publication

from .models import GridBacktestResult, GridRunMetrics, GridRunResult


NAVY = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
LIGHT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
GREEN = "548235"
RED = "C00000"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
OK_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
CHECK_FILL = PatternFill("solid", fgColor=LIGHT_RED)
BEST_FILL = PatternFill("solid", fgColor=LIGHT_YELLOW)
STRIPE_FILL = PatternFill("solid", fgColor="F8FAFC")
HEADER_FONT = Font(color=WHITE, bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
HEADER_BORDER = Border(bottom=THIN_GRAY)
HEADER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
WRAPPED_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

PERCENT_FORMAT = "0.00%"
NAV_FORMAT = "0.0000"
RATIO_FORMAT = "0.00"
COUNT_FORMAT = "0"
DATE_FORMAT = "yyyy-mm-dd"


def _metric(attribute: str) -> Callable[[GridRunResult], object]:
    return lambda run: (
        getattr(run.metrics, attribute) if run.metrics is not None else None
    )


@dataclass(frozen=True)
class ResultColumn:
    header: str
    accessor: Callable[[GridRunResult], object]
    number_format: str | None
    width: int
    visible: bool = True


CORE_RESULT_COLUMNS = (
    ResultColumn("Rank", lambda run: run.rank, COUNT_FORMAT, 8),
    ResultColumn("Run ID", lambda run: run.spec.run_id, None, 11),
    ResultColumn("Status", lambda run: run.status, None, 12),
    ResultColumn("w", lambda run: run.spec.lookback_window, COUNT_FORMAT, 7),
    ResultColumn(
        "p",
        lambda run: run.spec.deviation_threshold,
        PERCENT_FORMAT,
        9,
    ),
    ResultColumn(
        "P",
        lambda run: run.spec.variance_threshold,
        PERCENT_FORMAT,
        9,
    ),
    ResultColumn("l", lambda run: run.spec.rebalance_period, COUNT_FORMAT, 7),
    ResultColumn(
        "q",
        lambda run: run.spec.take_profit_threshold,
        PERCENT_FORMAT,
        9,
    ),
    ResultColumn(
        "Annualized Return",
        _metric("annualized_return"),
        PERCENT_FORMAT,
        17,
    ),
    ResultColumn(
        "Annualized Volatility",
        _metric("annualized_volatility"),
        PERCENT_FORMAT,
        19,
    ),
    ResultColumn("Sharpe", _metric("sharpe_ratio"), RATIO_FORMAT, 11),
    ResultColumn("Sortino", _metric("sortino_ratio"), RATIO_FORMAT, 11),
    ResultColumn(
        "Maximum Drawdown",
        _metric("maximum_drawdown"),
        PERCENT_FORMAT,
        18,
    ),
    ResultColumn("Calmar", _metric("calmar_ratio"), RATIO_FORMAT, 11),
    ResultColumn("Win Rate", _metric("win_rate"), PERCENT_FORMAT, 12),
    ResultColumn(
        "SPY Annualized Return",
        _metric("spy_annualized_return"),
        PERCENT_FORMAT,
        20,
    ),
    ResultColumn(
        "Excess Annualized Return",
        _metric("excess_annualized_return"),
        PERCENT_FORMAT,
        22,
    ),
    ResultColumn(
        "Information Ratio",
        _metric("information_ratio"),
        RATIO_FORMAT,
        17,
    ),
    ResultColumn(
        "Average Gross Exposure",
        _metric("average_gross_exposure"),
        PERCENT_FORMAT,
        21,
    ),
    ResultColumn(
        "Average Cash Weight",
        _metric("average_cash_weight"),
        PERCENT_FORMAT,
        19,
    ),
    ResultColumn(
        "Annualized Two-way Turnover",
        _metric("annualized_two_way_turnover"),
        PERCENT_FORMAT,
        25,
    ),
    ResultColumn("Run QC", _metric("overall_qc"), None, 11),
)

AUDIT_RESULT_COLUMNS = (
    ResultColumn(
        "Error Type",
        lambda run: run.error_type,
        None,
        18,
        visible=False,
    ),
    ResultColumn(
        "Error Message",
        lambda run: run.error_message,
        None,
        42,
        visible=False,
    ),
    ResultColumn(
        "Sessions",
        _metric("session_count"),
        COUNT_FORMAT,
        12,
        visible=False,
    ),
    ResultColumn(
        "Starting NAV",
        _metric("starting_nav"),
        NAV_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "Ending NAV",
        _metric("ending_nav"),
        NAV_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "Total Return",
        _metric("total_return"),
        PERCENT_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "Mean Daily Return",
        _metric("mean_daily_return"),
        PERCENT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Annualized Downside Volatility",
        _metric("annualized_downside_volatility"),
        PERCENT_FORMAT,
        28,
        visible=False,
    ),
    ResultColumn(
        "DD Peak Date",
        _metric("drawdown_peak_date"),
        DATE_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "DD Trough Date",
        _metric("drawdown_trough_date"),
        DATE_FORMAT,
        16,
        visible=False,
    ),
    ResultColumn(
        "DD Recovery Date",
        _metric("drawdown_recovery_date"),
        DATE_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Positive Sessions",
        _metric("positive_session_count"),
        COUNT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Negative Sessions",
        _metric("negative_session_count"),
        COUNT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Zero Sessions",
        _metric("zero_session_count"),
        COUNT_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "Average Positive Return",
        _metric("average_positive_return"),
        PERCENT_FORMAT,
        23,
        visible=False,
    ),
    ResultColumn(
        "Average Negative Return",
        _metric("average_negative_return"),
        PERCENT_FORMAT,
        23,
        visible=False,
    ),
    ResultColumn(
        "Payoff Ratio",
        _metric("payoff_ratio"),
        RATIO_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "Profit Factor",
        _metric("profit_factor"),
        RATIO_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "Best Daily Return",
        _metric("best_daily_return"),
        PERCENT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Worst Daily Return",
        _metric("worst_daily_return"),
        PERCENT_FORMAT,
        19,
        visible=False,
    ),
    ResultColumn(
        "Skewness",
        _metric("skewness"),
        RATIO_FORMAT,
        13,
        visible=False,
    ),
    ResultColumn(
        "Excess Kurtosis",
        _metric("excess_kurtosis"),
        RATIO_FORMAT,
        17,
        visible=False,
    ),
    ResultColumn(
        "Daily VaR 95%",
        _metric("daily_var_95"),
        PERCENT_FORMAT,
        16,
        visible=False,
    ),
    ResultColumn(
        "Daily CVaR 95%",
        _metric("daily_cvar_95"),
        PERCENT_FORMAT,
        17,
        visible=False,
    ),
    ResultColumn(
        "SPY Total Return",
        _metric("spy_total_return"),
        PERCENT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "SPY Annualized Volatility",
        _metric("spy_annualized_volatility"),
        PERCENT_FORMAT,
        24,
        visible=False,
    ),
    ResultColumn(
        "SPY Sharpe",
        _metric("spy_sharpe_ratio"),
        RATIO_FORMAT,
        14,
        visible=False,
    ),
    ResultColumn(
        "SPY Sortino",
        _metric("spy_sortino_ratio"),
        RATIO_FORMAT,
        14,
        visible=False,
    ),
    ResultColumn(
        "SPY Maximum Drawdown",
        _metric("spy_maximum_drawdown"),
        PERCENT_FORMAT,
        23,
        visible=False,
    ),
    ResultColumn(
        "SPY Calmar",
        _metric("spy_calmar_ratio"),
        RATIO_FORMAT,
        14,
        visible=False,
    ),
    ResultColumn(
        "Excess Total Return",
        _metric("excess_total_return"),
        PERCENT_FORMAT,
        20,
        visible=False,
    ),
    ResultColumn(
        "SPY Correlation",
        _metric("spy_correlation"),
        RATIO_FORMAT,
        17,
        visible=False,
    ),
    ResultColumn(
        "SPY Beta",
        _metric("spy_beta"),
        RATIO_FORMAT,
        12,
        visible=False,
    ),
    ResultColumn(
        "Annualized Alpha",
        _metric("annualized_alpha"),
        PERCENT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Tracking Error",
        _metric("tracking_error"),
        PERCENT_FORMAT,
        16,
        visible=False,
    ),
    ResultColumn(
        "Initial Events",
        _metric("initial_event_count"),
        COUNT_FORMAT,
        15,
        visible=False,
    ),
    ResultColumn(
        "Scheduled Events",
        _metric("scheduled_event_count"),
        COUNT_FORMAT,
        17,
        visible=False,
    ),
    ResultColumn(
        "Stop-win Events",
        _metric("stop_win_event_count"),
        COUNT_FORMAT,
        16,
        visible=False,
    ),
    ResultColumn(
        "Average Held Sessions",
        _metric("average_held_sessions"),
        "0.00",
        21,
        visible=False,
    ),
    ResultColumn(
        "Average K",
        _metric("average_cluster_count"),
        "0.00",
        13,
        visible=False,
    ),
    ResultColumn(
        "Average Active Clusters",
        _metric("average_active_cluster_count"),
        "0.00",
        22,
        visible=False,
    ),
    ResultColumn(
        "Average Inactive Clusters",
        _metric("average_inactive_cluster_count"),
        "0.00",
        23,
        visible=False,
    ),
    ResultColumn(
        "Average Target Gross",
        _metric("average_target_gross_exposure"),
        PERCENT_FORMAT,
        21,
        visible=False,
    ),
    ResultColumn(
        "Average Frozen Exposure",
        _metric("average_frozen_exposure"),
        PERCENT_FORMAT,
        23,
        visible=False,
    ),
    ResultColumn(
        "Average Positions",
        _metric("average_position_count"),
        "0.00",
        18,
        visible=False,
    ),
    ResultColumn(
        "Minimum Positions",
        _metric("minimum_position_count"),
        COUNT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Maximum Positions",
        _metric("maximum_position_count"),
        COUNT_FORMAT,
        18,
        visible=False,
    ),
    ResultColumn(
        "Missing Sessions",
        _metric("missing_session_count"),
        COUNT_FORMAT,
        17,
        visible=False,
    ),
    ResultColumn(
        "Missing Audit Rows",
        _metric("missing_audit_count"),
        COUNT_FORMAT,
        19,
        visible=False,
    ),
    ResultColumn(
        "FIFO Status",
        _metric("fifo_reconciliation_status"),
        None,
        14,
        visible=False,
    ),
)

RESULT_COLUMNS = CORE_RESULT_COLUMNS + AUDIT_RESULT_COLUMNS


def export_grid_backtest_workbook(
    result: GridBacktestResult,
    output_path: Path,
    *,
    replace_existing: bool = False,
) -> Path:
    with workbook_for_publication(
        output_path,
        replace_existing=replace_existing,
    ) as (workbook, output):
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Summary")
        results = workbook.create_sheet("Grid_Results")
        audit = workbook.create_sheet("Audit")

        _write_summary(summary, result)
        _write_grid_results(results, result)
        _write_audit(audit, result)
    return output


def _write_summary(sheet: Worksheet, result: GridBacktestResult) -> None:
    sheet.merge_cells("A1:L1")
    sheet["A1"] = (
        "Step 8 Grid Backtest — "
        f"{result.effective_start_date.isoformat()} to "
        f"{result.effective_end_date.isoformat()}"
    )
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    _section(sheet, 3, "Run overview", 12)
    overview = (
        (
            "Requested start",
            result.requested_start_date,
            "Requested end",
            result.requested_end_date,
            "Effective start",
            result.effective_start_date,
            "Effective end",
            result.effective_end_date,
        ),
        (
            "Combinations",
            len(result.runs),
            "Successful",
            result.successful_run_count,
            "Failed",
            result.failed_run_count,
            "Overall QC",
            result.overall_qc,
        ),
        (
            "Best run",
            result.best_run_id,
            "Ranking",
            "Sharpe descending",
            "Persistence",
            "Excel only",
            "Technical detail",
            "hidden in Grid_Results",
        ),
    )
    for row_number, values in enumerate(overview, start=4):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value=value)
        for column in (1, 3, 5, 7):
            sheet.cell(row_number, column).font = Font(bold=True)
    for cell in ("B4", "D4", "F4", "H4"):
        sheet[cell].number_format = DATE_FORMAT
    status_cell = sheet["H5"]
    status_cell.font = Font(
        bold=True,
        color=GREEN if result.overall_qc == "OK" else RED,
    )
    status_cell.fill = (
        OK_FILL if result.overall_qc == "OK" else CHECK_FILL
    )

    best = result.best_run
    _section(sheet, 9, "Best run parameters", 12)
    _write_table(
        sheet,
        start_row=10,
        headers=("Run ID", "Rank", "w", "p", "P", "l", "q", "Status"),
        rows=(
            (
                best.spec.run_id if best else None,
                best.rank if best else None,
                best.spec.lookback_window if best else None,
                best.spec.deviation_threshold if best else None,
                best.spec.variance_threshold if best else None,
                best.spec.rebalance_period if best else None,
                best.spec.take_profit_threshold if best else None,
                best.status if best else None,
            ),
        ),
        number_formats={
            2: COUNT_FORMAT,
            3: COUNT_FORMAT,
            4: PERCENT_FORMAT,
            5: PERCENT_FORMAT,
            6: COUNT_FORMAT,
            7: PERCENT_FORMAT,
        },
        auto_filter=False,
    )
    _section(sheet, 13, "Best run performance", 12)
    _write_table(
        sheet,
        start_row=14,
        headers=("Metric", "Strategy", "SPY", "Difference"),
        rows=_best_performance_rows(best),
        number_formats={
            2: PERCENT_FORMAT,
            3: PERCENT_FORMAT,
            4: PERCENT_FORMAT,
        },
        auto_filter=False,
    )
    for row in (18, 19, 21):
        for column in (2, 3, 4):
            sheet.cell(row, column).number_format = RATIO_FORMAT

    _section(sheet, 23, "Best run operations and exposure", 12)
    operations = (
        (
            "Annualized turnover",
            _best_metric(best, "annualized_two_way_turnover"),
            "Average held sessions",
            _best_metric(best, "average_held_sessions"),
        ),
        (
            "Average K",
            _best_metric(best, "average_cluster_count"),
            "Average active clusters",
            _best_metric(best, "average_active_cluster_count"),
        ),
        (
            "Average gross exposure",
            _best_metric(best, "average_gross_exposure"),
            "Average cash weight",
            _best_metric(best, "average_cash_weight"),
        ),
        (
            "Average positions",
            _best_metric(best, "average_position_count"),
            "Missing sessions",
            _best_metric(best, "missing_session_count"),
        ),
    )
    for row_number, values in enumerate(operations, start=24):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value=value)
        sheet.cell(row_number, 1).font = Font(bold=True)
        sheet.cell(row_number, 3).font = Font(bold=True)
    for cell in ("B24", "B26", "D26"):
        sheet[cell].number_format = PERCENT_FORMAT
    for cell in ("D24", "B25", "D25", "B27"):
        sheet[cell].number_format = "0.00"
    sheet["D27"].number_format = COUNT_FORMAT

    _section(sheet, 30, "Top 10 by Sharpe", 12)
    top_headers = (
        "Rank",
        "Run ID",
        "w",
        "p",
        "P",
        "l",
        "q",
        "Annualized Return",
        "Sharpe",
        "Max Drawdown",
        "Excess Ann. Return",
        "Turnover",
    )
    top_runs = sorted(
        (run for run in result.runs if run.rank is not None),
        key=lambda run: int(run.rank),
    )[:10]
    _write_table(
        sheet,
        start_row=31,
        headers=top_headers,
        rows=(
            (
                run.rank,
                run.spec.run_id,
                run.spec.lookback_window,
                run.spec.deviation_threshold,
                run.spec.variance_threshold,
                run.spec.rebalance_period,
                run.spec.take_profit_threshold,
                run.metrics.annualized_return if run.metrics else None,
                run.metrics.sharpe_ratio if run.metrics else None,
                run.metrics.maximum_drawdown if run.metrics else None,
                (
                    run.metrics.excess_annualized_return
                    if run.metrics
                    else None
                ),
                (
                    run.metrics.annualized_two_way_turnover
                    if run.metrics
                    else None
                ),
            )
            for run in top_runs
        ),
        number_formats={
            1: COUNT_FORMAT,
            3: COUNT_FORMAT,
            4: PERCENT_FORMAT,
            5: PERCENT_FORMAT,
            6: COUNT_FORMAT,
            7: PERCENT_FORMAT,
            8: PERCENT_FORMAT,
            9: RATIO_FORMAT,
            10: PERCENT_FORMAT,
            11: PERCENT_FORMAT,
            12: PERCENT_FORMAT,
        },
    )
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    widths = (24, 16, 24, 18, 20, 16, 22, 20, 12, 17, 19, 16)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_grid_results(
    sheet: Worksheet,
    result: GridBacktestResult,
) -> None:
    headers = tuple(column.header for column in RESULT_COLUMNS)
    rows = (
        tuple(column.accessor(run) for column in RESULT_COLUMNS)
        for run in result.runs
    )
    formats = {
        index: column.number_format
        for index, column in enumerate(RESULT_COLUMNS, start=1)
        if column.number_format is not None
    }
    _write_table(
        sheet,
        start_row=1,
        headers=headers,
        rows=rows,
        number_formats=formats,
    )
    hidden_columns: list[str] = []
    for index, column in enumerate(RESULT_COLUMNS, start=1):
        column_letter = get_column_letter(index)
        sheet.column_dimensions[column_letter].width = column.width
        if not column.visible:
            hidden_columns.append(column_letter)
    _hide_columns(sheet, hidden_columns)
    sheet.freeze_panes = "I2"
    sheet.sheet_view.showGridLines = False
    last_column = get_column_letter(len(RESULT_COLUMNS))
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    if sheet.max_row >= 2:
        sheet.conditional_formatting.add(
            f"C2:C{sheet.max_row}",
            FormulaRule(formula=['C2="SUCCESS"'], fill=OK_FILL),
        )
        sheet.conditional_formatting.add(
            f"C2:C{sheet.max_row}",
            FormulaRule(formula=['C2="FAILED"'], fill=CHECK_FILL),
        )
        sheet.conditional_formatting.add(
            f"A2:V{sheet.max_row}",
            FormulaRule(formula=["$A2=1"], fill=BEST_FILL),
        )
        sheet.conditional_formatting.add(
            f"V2:V{sheet.max_row}",
            FormulaRule(formula=['V2="OK"'], fill=OK_FILL),
        )
        sheet.conditional_formatting.add(
            f"V2:V{sheet.max_row}",
            FormulaRule(formula=['V2="CHECK"'], fill=CHECK_FILL),
        )


def _write_audit(sheet: Worksheet, result: GridBacktestResult) -> None:
    _title(sheet, "Grid Backtest Audit", 7)
    _section(sheet, 3, "Quality checks", 7)
    successful_sessions = {
        run.metrics.session_count
        for run in result.runs
        if run.metrics is not None
    }
    unique_ids = len({run.spec.run_id for run in result.runs})
    summary_checks = (
        ("Combination count", len(result.runs), result.config.combination_count, len(result.runs) - result.config.combination_count, 0, "OK" if len(result.runs) == result.config.combination_count else "CHECK", "One result row per Cartesian-product combination"),
        ("Unique run IDs", unique_ids, len(result.runs), unique_ids - len(result.runs), 0, "OK" if unique_ids == len(result.runs) else "CHECK", "Stable run IDs must be unique"),
        ("Successful session alignment", len(successful_sessions), 1, len(successful_sessions) - 1, 0, "OK" if len(successful_sessions) <= 1 else "CHECK", "All successful combinations must use the same return sessions"),
        ("Failed runs", result.failed_run_count, 0, result.failed_run_count, 0, "OK" if result.failed_run_count == 0 else "CHECK", "Failures remain visible and are excluded from ranking"),
        ("Overall QC", result.overall_qc, "OK", None, None, result.overall_qc, "Aggregates run status, NAV/exposure reconciliation and FIFO"),
    )
    _write_table(
        sheet,
        start_row=4,
        headers=("Check", "Actual", "Expected", "Difference", "Tolerance", "Status", "Notes"),
        rows=summary_checks,
        auto_filter=False,
    )

    _section(sheet, 12, "Search dimensions", 7)
    dimension_rows = (
        ("Lookback window", "w", _joined(result.config.lookback_windows), len(result.config.lookback_windows), "Shared clustering and stock-selection sessions"),
        ("Deviation threshold", "p", _joined(result.config.deviation_thresholds), len(result.config.deviation_thresholds), "Winner/loser cumulative raw-return deviation threshold"),
        ("Variance threshold", "P", _joined(result.config.variance_thresholds), len(result.config.variance_thresholds), "Cumulative explained variance used to select K"),
        ("Rebalance period", "l", _joined(result.config.rebalance_periods), len(result.config.rebalance_periods), "Earned return sessions before scheduled rebalance"),
        ("Take-profit threshold", "q", _joined(result.config.take_profit_thresholds), len(result.config.take_profit_thresholds), "Compounded round return for early rebalance"),
    )
    _write_table(
        sheet,
        start_row=13,
        headers=("Parameter", "Symbol", "Candidate values", "Count", "Meaning"),
        rows=dimension_rows,
        number_formats={4: COUNT_FORMAT},
        auto_filter=False,
    )

    _section(sheet, 21, "Fixed settings", 7)
    fixed_rows = (
        ("Requested start", result.requested_start_date, "date", "GridBacktestConfig", "Explicit included return date"),
        ("Requested end", result.requested_end_date, "date", "GridBacktestConfig", "Explicit included return date"),
        ("Effective start", result.effective_start_date, "SPY session", "BacktestMarketDataRepository", "Requested start rolls forward"),
        ("Effective end", result.effective_end_date, "SPY session", "BacktestMarketDataRepository", "Must be an SPY session"),
        ("Beta window", result.beta_window, "sessions", "PreprocessingConfig", "Includes historical session t"),
        ("K estimation window", result.cluster_count_estimation_window, "sessions", "Grid fixed setting", "Independent of actual clustering w"),
        ("Initial NAV", result.config.initial_nav, "NAV units", "GridBacktestConfig", "Scale only"),
        ("Annualization", result.config.annualization_sessions, "sessions", "GridBacktestConfig", "Risk-free and cash return are zero"),
        ("tau positive", result.sponge_config.tau_positive, "SPONGE", "SpongeSymConfig", "Fixed across runs"),
        ("tau negative", result.sponge_config.tau_negative, "SPONGE", "SpongeSymConfig", "Fixed across runs"),
        ("Random seed", result.sponge_config.random_seed, "integer", "SpongeSymConfig", "Fixed across runs"),
        ("k-means n_init", result.sponge_config.kmeans_n_init, "runs", "SpongeSymConfig", "Fixed across runs"),
        ("Combinations", result.config.combination_count, "runs", "Cartesian product", "After sorting and de-duplication"),
        ("Maximum combinations", result.config.maximum_combinations, "runs", "Safety guard", "Must be explicitly raised when exceeded"),
        ("Calculation version", result.calculation_version, "version", "GridBacktestResult", "Defines the grid calculation contract"),
    )
    _write_table(
        sheet,
        start_row=22,
        headers=("Setting", "Value", "Unit / policy", "Source", "Notes"),
        rows=fixed_rows,
        auto_filter=False,
    )
    for row in (23, 24, 25, 26):
        sheet.cell(row, 2).number_format = DATE_FORMAT

    exception_rows = _exception_rows(result)
    _section(sheet, 40, "Exceptions only", 7)
    _write_table(
        sheet,
        start_row=41,
        headers=(
            "Run ID",
            "Status",
            "Rank",
            "Run QC",
            "FIFO",
            "Error Type",
            "Error Message",
        ),
        rows=exception_rows,
        number_formats={3: COUNT_FORMAT},
        auto_filter=False,
    )

    definitions_section_row = 44 + len(exception_rows)
    _section(sheet, definitions_section_row, "Metric definitions", 7)
    _write_table(
        sheet,
        start_row=definitions_section_row + 1,
        headers=("Metric", "Category", "Definition", "Unit", "Undefined rule"),
        rows=_metric_definition_rows(),
    )

    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    widths = (31, 34, 62, 32, 46, 16, 58)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(1, sheet.max_row + 1):
        for column in (3, 5, 7):
            sheet.cell(row, column).alignment = WRAPPED_ALIGNMENT
    if sheet.max_row >= 5:
        sheet.conditional_formatting.add(
            "F5:F9",
            FormulaRule(formula=['F5="OK"'], fill=OK_FILL),
        )
        sheet.conditional_formatting.add(
            "F5:F9",
            FormulaRule(formula=['F5="CHECK"'], fill=CHECK_FILL),
        )
    if exception_rows:
        sheet.conditional_formatting.add(
            f"B42:B{41 + len(exception_rows)}",
            FormulaRule(formula=['B42="FAILED"'], fill=CHECK_FILL),
        )
        sheet.conditional_formatting.add(
            f"D42:D{41 + len(exception_rows)}",
            FormulaRule(formula=['D42="CHECK"'], fill=CHECK_FILL),
        )


def _write_table(
    sheet: Worksheet,
    *,
    start_row: int,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    number_formats: dict[int, str] | None = None,
    auto_filter: bool = True,
) -> None:
    formats = number_formats or {}
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
    for row_number, values in enumerate(rows, start=start_row + 1):
        row_values = tuple(values)
        if len(row_values) != len(headers):
            raise ValueError("Excel row does not match its header count")
        for column, value in enumerate(row_values, start=1):
            cell = sheet.cell(row_number, column, value=value)
            if (row_number - start_row) % 2 == 1:
                cell.fill = STRIPE_FILL
            if column in formats:
                cell.number_format = formats[column]
    sheet.row_dimensions[start_row].height = 34
    if auto_filter:
        last_column = sheet.cell(start_row, len(headers)).column_letter
        sheet.auto_filter.ref = (
            f"A{start_row}:{last_column}{max(sheet.max_row, start_row)}"
        )


def _title(sheet: Worksheet, value: str, columns: int) -> None:
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=columns,
    )
    cell = sheet.cell(1, 1, value=value)
    cell.fill = HEADER_FILL
    cell.font = Font(color=WHITE, bold=True, size=15)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28


def _section(
    sheet: Worksheet,
    row: int,
    label: str,
    columns: int,
) -> None:
    sheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=columns,
    )
    cell = sheet.cell(row, 1, value=label)
    cell.fill = SECTION_FILL
    cell.font = Font(bold=True, color=NAVY)
    cell.border = HEADER_BORDER


def _hide_columns(
    sheet: Worksheet,
    columns: Sequence[str],
) -> None:
    if not columns:
        return
    sheet.column_dimensions.group(
        columns[0],
        columns[-1],
        hidden=True,
        outline_level=1,
    )


def _best_performance_rows(
    run: GridRunResult | None,
) -> tuple[tuple[str, object, object, object], ...]:
    pairs = (
        ("Total return", "total_return", "spy_total_return"),
        (
            "Annualized return",
            "annualized_return",
            "spy_annualized_return",
        ),
        (
            "Annualized volatility",
            "annualized_volatility",
            "spy_annualized_volatility",
        ),
        ("Sharpe", "sharpe_ratio", "spy_sharpe_ratio"),
        ("Sortino", "sortino_ratio", "spy_sortino_ratio"),
        (
            "Maximum drawdown",
            "maximum_drawdown",
            "spy_maximum_drawdown",
        ),
        ("Calmar", "calmar_ratio", "spy_calmar_ratio"),
    )
    rows: list[tuple[str, object, object, object]] = []
    for label, strategy_attribute, spy_attribute in pairs:
        strategy_value = _best_metric(run, strategy_attribute)
        spy_value = _best_metric(run, spy_attribute)
        difference = (
            None
            if strategy_value is None or spy_value is None
            else float(strategy_value) - float(spy_value)
        )
        rows.append((label, strategy_value, spy_value, difference))
    return tuple(rows)


def _exception_rows(
    result: GridBacktestResult,
) -> tuple[tuple[object, ...], ...]:
    rows = tuple(
        (
            run.spec.run_id,
            run.status,
            run.rank,
            run.metrics.overall_qc if run.metrics else None,
            (
                run.metrics.fifo_reconciliation_status
                if run.metrics
                else None
            ),
            run.error_type,
            run.error_message,
        )
        for run in result.runs
        if (
            run.status != "SUCCESS"
            or run.metrics is None
            or run.metrics.overall_qc != "OK"
            or run.metrics.fifo_reconciliation_status != "OK"
        )
    )
    if rows:
        return rows
    return (
        (
            "None",
            "—",
            None,
            "OK",
            "OK",
            None,
            "All runs passed status, run QC, and FIFO checks.",
        ),
    )


def _best_metric(
    run: GridRunResult | None,
    attribute: str,
) -> object:
    if run is None or run.metrics is None:
        return None
    return getattr(run.metrics, attribute)


def _joined(values: Sequence[object]) -> str:
    return ", ".join(map(str, values))


def _metric_definition_rows() -> tuple[tuple[str, str, str, str, str], ...]:
    explicit = {
        "annualized_return": ("Return", "(ending NAV / starting NAV)^(annualization sessions / observed sessions) - 1", "annual rate", "Blank unless both NAV values are positive"),
        "annualized_volatility": ("Risk", "Sample standard deviation of daily returns × sqrt(annualization sessions)", "annual rate", "Blank when daily volatility is zero or fewer than two observations"),
        "sharpe_ratio": ("Risk-adjusted", "Mean daily return / sample daily volatility × sqrt(annualization sessions); risk-free rate is zero", "ratio", "Blank when daily volatility is zero"),
        "annualized_downside_volatility": ("Risk", "Sample standard deviation of negative daily returns × sqrt(annualization sessions)", "annual rate", "Blank with fewer than two negative returns or zero downside volatility"),
        "sortino_ratio": ("Risk-adjusted", "Mean daily return / sample standard deviation of negative daily returns × sqrt(annualization sessions)", "ratio", "Blank with fewer than two negative returns or zero downside volatility"),
        "maximum_drawdown": ("Risk", "Minimum signed NAV / running peak NAV - 1, including starting NAV", "return", "Zero when no drawdown occurs"),
        "calmar_ratio": ("Risk-adjusted", "Annualized return / absolute maximum drawdown", "ratio", "Blank when maximum drawdown is zero or annualized return is undefined"),
        "daily_var_95": ("Tail risk", "Historical 5th percentile of daily returns using linear interpolation", "daily return", "Always defined for a successful run"),
        "daily_cvar_95": ("Tail risk", "Mean daily return at or below the historical 5th percentile", "daily return", "Always defined for a successful run"),
        "spy_beta": ("Benchmark", "Sample covariance(strategy, SPY) / sample variance(SPY)", "ratio", "Blank when SPY variance is zero"),
        "annualized_alpha": ("Benchmark", "252 × [mean(strategy return) - beta × mean(SPY return)]; risk-free rate is zero", "annual rate", "Blank when beta is undefined"),
        "tracking_error": ("Benchmark", "Sample standard deviation of strategy minus SPY daily return × sqrt(annualization sessions)", "annual rate", "Blank when active-return volatility is zero"),
        "information_ratio": ("Benchmark", "Mean active daily return / sample active-return volatility × sqrt(annualization sessions)", "ratio", "Blank when active-return volatility is zero"),
        "annualized_two_way_turnover": ("Operations", "0.5 × sum(executed non-initial trade notional / trade-date NAV) × annualization sessions / observed sessions", "annual rate", "Zero when no post-initial trades execute"),
        "overall_qc": ("Audit", "OK only when daily values are finite, cash plus gross exposure reconciles to one, and FIFO status is OK", "status", "CHECK identifies a run requiring review"),
    }
    rows: list[tuple[str, str, str, str, str]] = []
    for metric_field in fields(GridRunMetrics):
        name = metric_field.name
        category, definition, unit, undefined = explicit.get(
            name,
            _default_metric_definition(name),
        )
        rows.append(
            (
                name,
                category,
                definition,
                unit,
                undefined,
            )
        )
    return tuple(rows)


def _default_metric_definition(
    name: str,
) -> tuple[str, str, str, str]:
    definitions = {
        "session_count": ("Setup", "Number of aligned strategy and SPY return sessions", "count", "Never blank for a successful run"),
        "starting_nav": ("Return", "Configured initial NAV", "NAV", "Never blank"),
        "ending_nav": ("Return", "Final strategy NAV after the last included return", "NAV", "Never blank"),
        "total_return": ("Return", "Ending NAV / starting NAV - 1", "return", "Never blank"),
        "mean_daily_return": ("Return", "Arithmetic mean of daily strategy returns", "daily return", "Never blank"),
        "drawdown_peak_date": ("Risk", "Date of the NAV peak preceding maximum drawdown", "date", "Blank when no drawdown occurs"),
        "drawdown_trough_date": ("Risk", "Date of the maximum-drawdown trough", "date", "Blank when no drawdown occurs"),
        "drawdown_recovery_date": ("Risk", "First later date NAV recovers the prior peak", "date", "Blank when not recovered or no drawdown occurs"),
        "win_rate": ("Distribution", "Positive-return sessions / all sessions", "rate", "Never blank"),
        "payoff_ratio": ("Distribution", "Average positive return / absolute average negative return", "ratio", "Blank without both positive and negative sessions"),
        "profit_factor": ("Distribution", "Sum of positive returns / absolute sum of negative returns", "ratio", "Blank without negative sessions"),
        "skewness": ("Distribution", "Bias-corrected sample skewness of daily returns", "ratio", "Blank with fewer than three sessions"),
        "excess_kurtosis": ("Distribution", "Bias-corrected sample excess kurtosis of daily returns", "ratio", "Blank with fewer than four sessions"),
        "spy_correlation": ("Benchmark", "Sample daily return correlation between strategy and SPY", "ratio", "Blank when either series has zero volatility"),
        "excess_total_return": ("Benchmark", "Strategy total return minus SPY total return", "return", "Never blank"),
        "excess_annualized_return": ("Benchmark", "Strategy annualized return minus SPY annualized return", "annual rate", "Blank when either annualized return is undefined"),
        "average_held_sessions": ("Operations", "Mean held_sessions across non-initial rebalance events", "sessions", "Blank when no later rebalance occurs"),
        "fifo_reconciliation_status": ("Audit", "FIFO buy-lot and sell-unit reconciliation status from step 7", "status", "Never blank for a successful run"),
    }
    if name in definitions:
        return definitions[name]
    if name.startswith("spy_"):
        return ("SPY benchmark", f"SPY counterpart of {name[4:].replace('_', ' ')}", "same as strategy metric", "Uses the same undefined rule as the strategy metric")
    if "count" in name or name.endswith("_sessions") or name.endswith("_events"):
        return ("Operations", name.replace("_", " ").capitalize(), "count", "Never blank for a successful run")
    if any(token in name for token in ("return", "volatility", "drawdown", "rate", "exposure", "weight", "turnover", "alpha", "error")):
        return ("Calculated output", name.replace("_", " ").capitalize(), "rate", "See related denominator or sample-size rule")
    return ("Calculated output", name.replace("_", " ").capitalize(), "value", "Never blank unless the required sample is unavailable")
