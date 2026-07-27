from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import PreprocessingSnapshot


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
DIAGONAL_FILL = PatternFill("solid", fgColor="E7E6E6")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
CHECK_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
QC_TOLERANCE = 1e-10


def export_snapshot_workbook(
    snapshot: PreprocessingSnapshot,
    output_path: Path,
    *,
    replace_existing: bool = False,
) -> Path:
    output = Path(output_path).resolve()
    if output.exists() and not replace_existing:
        raise FileExistsError(
            f"Excel output already exists: {output}. Use --replace to overwrite it."
        )
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    parameters = workbook.create_sheet("Summary")
    beta = workbook.create_sheet("Beta_Used")
    returns = workbook.create_sheet("Stock_Returns")
    residual = workbook.create_sheet("Residual_Matrix")
    correlation = workbook.create_sheet("Correlation_Matrix")
    excluded = workbook.create_sheet("Excluded_Stocks")

    _write_parameters(parameters, snapshot)
    _write_wide_matrix(
        beta,
        snapshot.beta_matrix,
        "Trade Date",
        number_format="0.0000",
    )
    _write_stock_returns(returns, snapshot)
    _write_wide_matrix(
        residual,
        snapshot.residual_matrix,
        "Trade Date",
        number_format="0.0000%",
    )
    _write_correlation(correlation, snapshot)
    _write_exclusions(excluded, snapshot)

    workbook.save(output)
    return output


def _write_parameters(sheet: Worksheet, snapshot: PreprocessingSnapshot) -> None:
    rows = [
        ("Run setup", None),
        ("As-of date", snapshot.as_of_date),
        ("Residual window start", snapshot.window_start),
        ("Residual window end", snapshot.window_end),
        ("Beta window", snapshot.beta_window),
        ("Correlation window", snapshot.correlation_window),
        ("Selected stocks", snapshot.selected_stock_count),
        ("Valid stocks", snapshot.valid_stock_count),
        ("Excluded stocks", snapshot.excluded_stock_count),
        (None, None),
        ("Quality", None),
        ("Maximum asymmetry", snapshot.quality.maximum_asymmetry),
        ("Minimum correlation", snapshot.quality.minimum_correlation),
        ("Maximum correlation", snapshot.quality.maximum_correlation),
        ("Minimum eigenvalue", snapshot.quality.minimum_eigenvalue),
        ("Numerical rank", snapshot.quality.numerical_rank),
        (
            "Contains non-finite values",
            "YES" if snapshot.quality.has_non_finite_values else "NO",
        ),
        ("Overall QC", _preprocessing_qc_status(snapshot)),
        (None, None),
        ("Method", None),
        ("Return basis", snapshot.return_basis),
        ("Beta alignment", snapshot.beta_alignment),
        ("Missing policy", snapshot.missing_policy),
        ("Variance epsilon", snapshot.variance_epsilon),
        (None, None),
        ("Provenance", None),
        ("Preprocessing run id", snapshot.preprocessing_run_id),
        ("Snapshot id", snapshot.snapshot_id),
        ("Calculation version", snapshot.calculation_version),
    ]
    for row in rows:
        sheet.append(row)

    sheet.insert_rows(1)
    sheet.merge_cells("A1:B1")
    sheet["A1"] = f"Preprocessing Snapshot — {snapshot.as_of_date.isoformat()}"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    section_rows = (2, 12, 21, 27)
    for row in section_rows:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1).fill = SECTION_FILL
        sheet.cell(row, 1).font = Font(bold=True, color="1F4E78")
        sheet.cell(row, 1).border = Border(bottom=THIN_GRAY)

    row_by_label = {
        str(sheet.cell(row_number, 1).value): row_number
        for row_number in range(2, sheet.max_row + 1)
        if sheet.cell(row_number, 1).value is not None
    }
    for label in ("As-of date", "Residual window start", "Residual window end"):
        sheet.cell(row_by_label[label], 2).number_format = "yyyy-mm-dd"
    for label in (
        "Beta window",
        "Correlation window",
        "Selected stocks",
        "Valid stocks",
        "Excluded stocks",
        "Numerical rank",
    ):
        sheet.cell(row_by_label[label], 2).number_format = "0"
    for label in ("Minimum correlation", "Maximum correlation"):
        sheet.cell(row_by_label[label], 2).number_format = "0.0000"
    sheet.cell(
        row_by_label["Maximum asymmetry"],
        2,
    ).number_format = "0.000000E+00"
    sheet.cell(row_by_label["Variance epsilon"], 2).number_format = "0.000000E+00"
    sheet.cell(row_by_label["Minimum eigenvalue"], 2).number_format = "0.000000E+00"
    sheet.cell(row_by_label["Contains non-finite values"], 2).number_format = "General"
    qc_cell = sheet.cell(row_by_label["Overall QC"], 2)
    qc_cell.font = Font(
        bold=True,
        color="548235" if qc_cell.value == "OK" else "C00000",
    )
    qc_cell.fill = OK_FILL if qc_cell.value == "OK" else CHECK_FILL
    for row in range(3, sheet.max_row + 1):
        if sheet.cell(row, 1).value is not None and row not in section_rows:
            sheet.cell(row, 1).font = Font(bold=True)

    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 58


def _preprocessing_qc_status(snapshot: PreprocessingSnapshot) -> str:
    quality = snapshot.quality
    checks = (
        not quality.has_non_finite_values,
        quality.maximum_asymmetry < QC_TOLERANCE,
        quality.minimum_correlation >= -1.0 - QC_TOLERANCE,
        quality.maximum_correlation <= 1.0 + QC_TOLERANCE,
        quality.minimum_eigenvalue >= -QC_TOLERANCE,
        snapshot.valid_stock_count + snapshot.excluded_stock_count
        == snapshot.selected_stock_count,
    )
    return "OK" if all(checks) else "CHECK"


def _write_wide_matrix(
    sheet: Worksheet,
    frame,
    index_label: str,
    *,
    number_format: str,
) -> None:
    sheet.append([index_label, *map(str, frame.columns)])
    for index, row in frame.iterrows():
        sheet.append([index.to_pydatetime(), *[float(value) for value in row]])
    _style_header(sheet, 1, len(frame.columns) + 1)
    sheet.freeze_panes = "B2"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 13
    for column in range(2, len(frame.columns) + 2):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = 11
        for row in range(2, len(frame.index) + 2):
            sheet.cell(row, column).number_format = number_format
    for row in range(2, len(frame.index) + 2):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"


def _write_stock_returns(sheet: Worksheet, snapshot: PreprocessingSnapshot) -> None:
    sheet.append(["Trade Date", "SPY", *snapshot.tickers])
    for trade_date in snapshot.stock_return_matrix.index:
        sheet.append(
            [
                trade_date.to_pydatetime(),
                float(snapshot.market_returns.loc[trade_date]),
                *[
                    float(snapshot.stock_return_matrix.loc[trade_date, ticker])
                    for ticker in snapshot.tickers
                ],
            ]
        )
    _style_header(sheet, 1, len(snapshot.tickers) + 2)
    sheet.freeze_panes = "C2"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 13
    for column in range(2, len(snapshot.tickers) + 3):
        sheet.column_dimensions[get_column_letter(column)].width = 11
        for row in range(2, len(snapshot.stock_return_matrix.index) + 2):
            sheet.cell(row, column).number_format = "0.0000%"
    for row in range(2, len(snapshot.stock_return_matrix.index) + 2):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"


def _write_correlation(sheet: Worksheet, snapshot: PreprocessingSnapshot) -> None:
    sheet.append(["Ticker", *snapshot.tickers])
    for ticker, row in snapshot.correlation_matrix.iterrows():
        sheet.append([str(ticker), *[float(value) for value in row]])
    size = len(snapshot.tickers)
    _style_header(sheet, 1, size + 1)
    for row in range(2, size + 2):
        sheet.cell(row, 1).fill = SECTION_FILL
        sheet.cell(row, 1).font = Font(bold=True)
        for column in range(2, size + 2):
            sheet.cell(row, column).number_format = "0.0000"
    matrix_range = f"B2:{get_column_letter(size + 1)}{size + 1}"
    sheet.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=["COLUMN()-COLUMN($B$2)=ROW()-ROW($B$2)"],
            fill=DIAGONAL_FILL,
            stopIfTrue=True,
        ),
    )
    sheet.conditional_formatting.add(
        matrix_range,
        ColorScaleRule(
            start_type="num",
            start_value=-1,
            start_color="2166AC",
            mid_type="num",
            mid_value=0,
            mid_color="FFFFFF",
            end_type="num",
            end_value=1,
            end_color="B2182B",
        ),
    )
    sheet.freeze_panes = "B2"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 13
    for column in range(2, size + 2):
        sheet.column_dimensions[get_column_letter(column)].width = 10


def _write_exclusions(sheet: Worksheet, snapshot: PreprocessingSnapshot) -> None:
    sheet.append(["Market Cap Rank", "Ticker", "Reason"])
    for row in snapshot.exclusions.itertuples(index=False):
        sheet.append([int(row.market_cap_rank), str(row.ticker), str(row.reason)])
    _style_header(sheet, 1, 3)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A1:C{max(1, sheet.max_row)}"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 48


def _style_header(sheet: Worksheet, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(row, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[row].height = 24
