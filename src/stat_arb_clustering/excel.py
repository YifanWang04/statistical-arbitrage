from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import SpongeSymResult


NAVY = "1F4E78"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"
GREEN = "548235"
RED = "C00000"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
CHECK_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FONT = Font(color=WHITE, bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
SCIENTIFIC_FORMAT = "0.000000E+00"


def export_clustering_workbook(
    result: SpongeSymResult,
    output_path: Path,
    *,
    replace_existing: bool = False,
) -> Path:
    if result.cluster_count_result is None:
        raise ValueError(
            "clustering report requires cluster-count provenance from "
            "cluster_stocks_for_date"
        )
    output = Path(output_path).resolve()
    if output.exists() and not replace_existing:
        raise FileExistsError(
            f"Excel output already exists: {output}. Use --replace to overwrite it."
        )
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Summary")
    eigenvalues = workbook.create_sheet("Eigenvalues")
    embedding = workbook.create_sheet("Spectral_Embedding")
    assignments = workbook.create_sheet("Cluster_Assignments")

    _write_summary(summary, result)
    _write_eigenvalues(eigenvalues, result)
    _write_embedding(embedding, result)
    _write_assignments(assignments, result)

    workbook.save(output)
    return output


def _write_summary(sheet: Worksheet, result: SpongeSymResult) -> None:
    cluster_count = result.cluster_count_result
    assert cluster_count is not None
    rows = [
        ("Run setup", None),
        ("As-of date", result.as_of_date),
        ("Clustering window start", result.clustering_window_start),
        ("Clustering window end", result.clustering_window_end),
        ("Cluster-count window start", cluster_count.window_start),
        ("Cluster-count window end", cluster_count.window_end),
        ("Stock count N", result.stock_count),
        ("Selected K", result.requested_cluster_count),
        ("Embedding dimension", result.embedding_dimension),
        (None, None),
        ("Configuration", None),
        ("Variance threshold P", cluster_count.variance_threshold),
        ("Tau positive", result.config.tau_positive),
        ("Tau negative", result.config.tau_negative),
        ("Random seed", result.config.random_seed),
        ("KMeans n_init", result.config.kmeans_n_init),
        ("KMeans max_iter", result.config.kmeans_max_iter),
        (None, None),
        ("Diagnostics", None),
        ("KMeans iterations used", result.quality.kmeans_iterations),
        ("KMeans inertia", result.quality.kmeans_inertia),
        (
            "Maximum generalized eigen residual",
            result.quality.maximum_generalized_eigen_residual,
        ),
        (
            "Zero positive-degree stocks",
            result.quality.zero_positive_degree_count,
        ),
        (
            "Zero negative-degree stocks",
            result.quality.zero_negative_degree_count,
        ),
        ("Maximum input asymmetry", result.quality.maximum_input_asymmetry),
        (
            "Maximum adjacency reconstruction error",
            result.quality.maximum_reconstruction_error,
        ),
        ("Nonempty clusters", result.quality.nonempty_cluster_count),
        ("Minimum cluster size", result.quality.minimum_cluster_size),
        ("Maximum cluster size", result.quality.maximum_cluster_size),
        ("Overall QC", _clustering_qc_status(result)),
        (None, None),
        ("Method", None),
        ("Return basis", result.return_basis),
        (
            "Embedding convention",
            "K-1 eigenvectors scaled by inverse generalized eigenvalue",
        ),
        (None, None),
        ("Provenance", None),
        ("Clustering snapshot id", result.clustering_snapshot_id),
        ("Cluster-count snapshot id", cluster_count.snapshot_id),
        ("Preprocessing run id", result.preprocessing_run_id),
        ("Source calculation version", result.source_calculation_version),
        ("Cluster-count version", cluster_count.calculation_version),
        ("Clustering version", result.calculation_version),
    ]
    for row in rows:
        sheet.append(row)

    sheet.insert_rows(1)
    sheet.merge_cells("A1:B1")
    sheet["A1"] = f"SPONGE_sym Clustering — {result.as_of_date.isoformat()}"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    section_rows = (2, 12, 20, 33, 37)
    for row in section_rows:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1).fill = SECTION_FILL
        sheet.cell(row, 1).font = Font(bold=True, color=NAVY)
        sheet.cell(row, 1).border = Border(bottom=THIN_GRAY)

    rows_by_label = {
        sheet.cell(row, 1).value: row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value is not None
    }
    for label in (
        "As-of date",
        "Clustering window start",
        "Clustering window end",
        "Cluster-count window start",
        "Cluster-count window end",
    ):
        sheet.cell(rows_by_label[label], 2).number_format = "yyyy-mm-dd"
    for label in (
        "Stock count N",
        "Selected K",
        "Embedding dimension",
        "Random seed",
        "KMeans n_init",
        "KMeans max_iter",
        "KMeans iterations used",
        "Zero positive-degree stocks",
        "Zero negative-degree stocks",
        "Nonempty clusters",
        "Minimum cluster size",
        "Maximum cluster size",
    ):
        sheet.cell(rows_by_label[label], 2).number_format = "0"
    sheet.cell(
        rows_by_label["Variance threshold P"],
        2,
    ).number_format = "0.00%"
    for label in ("Tau positive", "Tau negative", "KMeans inertia"):
        sheet.cell(rows_by_label[label], 2).number_format = "0.0000"
    for label in (
        "Maximum generalized eigen residual",
        "Maximum input asymmetry",
        "Maximum adjacency reconstruction error",
    ):
        sheet.cell(rows_by_label[label], 2).number_format = SCIENTIFIC_FORMAT

    qc_cell = sheet.cell(rows_by_label["Overall QC"], 2)
    qc_cell.font = Font(
        bold=True,
        color=GREEN if qc_cell.value == "OK" else RED,
    )
    qc_cell.fill = OK_FILL if qc_cell.value == "OK" else CHECK_FILL
    for row in range(3, sheet.max_row + 1):
        if sheet.cell(row, 1).value is not None and row not in section_rows:
            sheet.cell(row, 1).font = Font(bold=True)
    sheet.cell(rows_by_label["Embedding convention"], 2).alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 62


def _write_eigenvalues(sheet: Worksheet, result: SpongeSymResult) -> None:
    sheet.append(
        [
            "Rank",
            "Generalized Eigenvalue",
            "Inverse Eigenvalue Weight",
            "Final Residual Norm",
        ]
    )
    for rank, values in enumerate(
        zip(
            result.generalized_eigenvalues,
            result.inverse_eigenvalue_weights,
            result.generalized_eigen_residuals,
            strict=True,
        ),
        start=1,
    ):
        sheet.append([rank, *values])

    _style_table(sheet, widths=(10, 24, 24, 24))
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "0"
        for column in range(2, 5):
            sheet.cell(row, column).number_format = SCIENTIFIC_FORMAT


def _write_embedding(sheet: Worksheet, result: SpongeSymResult) -> None:
    sheet.append(["Ticker", *result.embedding.columns])
    for ticker, row in result.embedding.iterrows():
        sheet.append([ticker, *map(float, row)])

    widths = (14,) + (14,) * result.embedding_dimension
    _style_table(sheet, widths=widths, freeze_panes="B2")
    for column in range(2, result.embedding_dimension + 2):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = "0.000000"


def _write_assignments(sheet: Worksheet, result: SpongeSymResult) -> None:
    sheet.append(
        [
            "Ticker",
            "Market Cap Rank",
            "Cluster ID",
            "Cluster Size",
            "Positive Degree",
            "Negative Degree",
        ]
    )
    rows = zip(
        result.tickers,
        result.market_cap_ranks,
        result.cluster_labels,
        result.positive_degrees,
        result.negative_degrees,
        strict=True,
    )
    for ticker, rank, label, positive_degree, negative_degree in rows:
        sheet.append(
            [
                ticker,
                rank,
                label,
                result.cluster_sizes[label],
                positive_degree,
                negative_degree,
            ]
        )

    _style_table(sheet, widths=(14, 18, 12, 14, 18, 18))
    for row in range(2, sheet.max_row + 1):
        for column in (2, 3, 4):
            sheet.cell(row, column).number_format = "0"
        for column in (5, 6):
            sheet.cell(row, column).number_format = "0.0000"


def _clustering_qc_status(result: SpongeSymResult) -> str:
    quality = result.quality
    checks = (
        len(result.cluster_labels) == result.stock_count,
        len(result.cluster_sizes) == result.requested_cluster_count,
        sum(result.cluster_sizes) == result.stock_count,
        quality.nonempty_cluster_count == result.requested_cluster_count,
        quality.minimum_cluster_size > 0,
        quality.maximum_input_asymmetry < 1e-10,
        quality.maximum_reconstruction_error < 1e-10,
        quality.maximum_generalized_eigen_residual < 1e-4,
    )
    return "OK" if all(checks) else "CHECK"


def _style_table(
    sheet: Worksheet,
    *,
    widths: tuple[int, ...],
    freeze_panes: str = "A2",
) -> None:
    _style_header(sheet, columns=len(widths))
    sheet.freeze_panes = freeze_panes
    last_column = get_column_letter(len(widths))
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(2, sheet.max_row + 1):
        if row % 2 == 0:
            for column in range(1, len(widths) + 1):
                sheet.cell(row, column).fill = PatternFill(
                    "solid",
                    fgColor="F8FAFC",
                )


def _style_header(sheet: Worksheet, *, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(1, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[1].height = 32
